# ==========================================
# IMPORTS COMPLETS AU DÉBUT DU FICHIER
# ==========================================
import logging
import json
import pandas as pd
import csv
from datetime import datetime, timedelta

# Django imports de base
from django.urls import reverse_lazy, reverse
from django.views.generic import CreateView, TemplateView, ListView, DeleteView
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse, HttpResponseNotFound, HttpResponseServerError
from django.contrib import messages
from django.utils import timezone
from django.db.models import Max, Count, Sum, Avg, Q
from django.core.cache import cache
from django.views.defaults import page_not_found, server_error
from django.db import transaction
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt

# Imports des modèles et formulaires
#from .forms import EtablissementRegistrationForm, EtablissementLoginForm
from .models import (
    FichierImport, Etablissement, Credit_Amortissables, Decouverts, 
    Affacturage, Cautions, Effets_commerces, Spot, TokenInscription, 
    ActionUtilisateur, Etablissement, User, HistoriqueEmail
)

from .utils import (
    traiter_fichier_excel, 
    extraire_et_calculer_teg, 
    generer_statistiques_teg,
)

from .email_utils import (
    envoyer_email_invitation,
    envoyer_email_validation,
    envoyer_email_rejet,
    envoyer_email_notification_acnef,
    renvoyer_email
)

logger = logging.getLogger(__name__)

# ==========================================
# VUE DE CONNEXION UNIVERSELLE
# ==========================================

from django.contrib.auth.views import LoginView, LogoutView
from django.urls import reverse_lazy
from django.shortcuts import redirect
from django.contrib import messages
from .forms import ConnexionUniverselleForm
from .models import ActionUtilisateur

class ConnexionUniverselleView(LoginView):
    """
    Vue de connexion universelle pour tous les rôles
    Redirige automatiquement selon le rôle de l'utilisateur
    """
    template_name = 'accounts/connexion.html'
    form_class = ConnexionUniverselleForm
    redirect_authenticated_user = True
    
    def get_form_kwargs(self):
        """S'assurer que le formulaire reçoit bien le request"""
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def form_valid(self, form):
        """
        ✅ CORRECTION : Enregistrer l'action ICI (appelé UNE SEULE FOIS)
        Cette méthode est exécutée quand le formulaire de connexion est valide
        """
        # D'abord appeler le parent pour connecter l'utilisateur
        response = super().form_valid(form)
        
        # Maintenant self.request.user est authentifié
        user = self.request.user
        
        # Mettre à jour la dernière connexion
        user.derniere_connexion = timezone.now()
        user.save(update_fields=['derniere_connexion'])
        
        # ✅ Enregistrer la connexion dans le journal UNE SEULE FOIS
        ActionUtilisateur.enregistrer_action(
            utilisateur=user,
            type_action='CONNEXION',
            description=f"Connexion réussie - Rôle: {user.get_role_display()}",
            etablissement=user.etablissement if user.etablissement else None,
            request=self.request
        )
        
        return response
    
    def get_success_url(self):
        """
        ✅ SIMPLIFIÉ : Juste retourner l'URL de redirection selon le rôle
        Plus de journalisation ici pour éviter les doublons
        """
        user = self.request.user
        
        # Redirection selon le rôle
        if user.role == 'ACNEF':
            messages.success(self.request, f"Bienvenue {user.get_full_name()} - Administrateur CNEF")
            return reverse_lazy('interface_chef')  # admin/interface_chef.html
        
        elif user.role == 'UCNEF':
            messages.success(self.request, f"Bienvenue {user.get_full_name()} - Utilisateur CNEF")
            return reverse_lazy('interface_chef')  # admin/interface_chef.html
        
        elif user.role == 'AEF':
            messages.success(self.request, f"Bienvenue {user.get_full_name()} - Administrateur {user.etablissement.Nom_etablissement}")
            return reverse_lazy('interface_aef')  
        
        elif user.role == 'UEF':
            messages.success(self.request, f"Bienvenue {user.get_full_name()} - {user.etablissement.Nom_etablissement}")
            return reverse_lazy('tableau_de_bord')  # accounts/tableau_de_bord.html
        
        # Par défaut (ne devrait jamais arriver)
        return reverse_lazy('tableau_de_bord')
    
    def form_invalid(self, form):
        """Gérer les erreurs de connexion"""
        messages.error(self.request, "Email ou mot de passe incorrect.")
        return super().form_invalid(form)
    
    def get_context_data(self, **kwargs):
        """Ajouter des données au contexte du template"""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Connexion - Plateforme CNEF'
        return context

class DeconnexionView(LogoutView):
    """
    Vue de déconnexion universelle
    """
    next_page = reverse_lazy('connexion')
    
    def dispatch(self, request, *args, **kwargs):
        """Enregistrer la déconnexion avant de déconnecter l'utilisateur"""
        if request.user.is_authenticated:
            ActionUtilisateur.enregistrer_action(
                utilisateur=request.user,
                type_action='DECONNEXION',
                description=f"Déconnexion - Rôle: {request.user.get_role_display()}",
                etablissement=request.user.etablissement if request.user.etablissement else None,
                request=request
            )
            messages.info(request, "Vous avez été déconnecté avec succès.")
        
        return super().dispatch(request, *args, **kwargs)


# ==========================================
# FONCTION HELPER POUR VÉRIFIER LES RÔLES
# ==========================================

def verifier_role_et_rediriger(user):
    """
    Fonction utilitaire pour vérifier le rôle et obtenir l'URL de redirection
    Peut être utilisée dans d'autres vues si nécessaire
    """
    redirections = {
        'ACNEF': 'interface_chef',
        'UCNEF': 'interface_chef',
        'AEF': 'interface_aef',  
        'UEF': 'tableau_de_bord',
    }
    
    return redirections.get(user.role, 'tableau_de_bord')
# ==========================================
# TESTS DE PERMISSIONS UNIFIÉS
# ==========================================

def is_aef(user):
    """Vérifie si l'utilisateur est AEF"""
    return user.is_authenticated and hasattr(user, 'role') and user.role == 'AEF'

def is_uef(user):
    """Vérifie si l'utilisateur est AEF"""
    return user.is_authenticated and hasattr(user, 'role') and user.role == 'UEF'


def is_acnef(user):
    """Vérifie si l'utilisateur est un administrateur CNEF"""
    if not user.is_authenticated:
        return False
    if hasattr(user, 'role'):
        return user.role == 'ACNEF'
    return user.is_staff or user.is_superuser

def is_ucnef(user):
    """Vérifie si l'utilisateur est UCNEF"""
    return user.is_authenticated and user.role == 'UCNEF'


def is_chef(user):
    """Vérifie si l'utilisateur est un chef/admin"""
    return user.is_authenticated and (user.is_staff or user.is_superuser or (hasattr(user, 'role') and user.role == 'ACNEF'))


def is_cnef_user(user):
    """Vérifie si l'utilisateur est du CNEF (ACNEF ou UCNEF)"""
    if not user.is_authenticated:
        return False
    if hasattr(user, 'role'):
        return user.role in ['ACNEF', 'UCNEF']
    return user.is_staff or user.is_superuser

def is_etablissement_user(user):
    """Vérifie si l'utilisateur fait partie d'un établissement"""
    return user.is_authenticated and hasattr(user, 'role') and user.role in ['AEF', 'UEF']


class TableauDeBordView(LoginRequiredMixin, TemplateView):
    template_name = 'accounts/tableau_de_bord.html'
    login_url = reverse_lazy('connexion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        user = self.request.user
        
        # Vérifie que l'utilisateur a un établissement (UEF/AEF)
        if hasattr(user, 'etablissement') and user.etablissement:
            context['imports_recents'] = FichierImport.objects.filter(
                etablissement_cnef=user.etablissement  # LE BON NOM DU CHAMP
            ).order_by('-date_import')[:5]
        else:
            context['imports_recents'] = FichierImport.objects.none()
        
        return context

class FichiersListAPIView(ListView):
    model = FichierImport
    queryset = FichierImport.objects.all().select_related('etablissement_cnef')

    def render_to_response(self, context, **response_kwargs):
        statut_filter = self.request.GET.get('statut', 'tous')
        etablissement_filter = self.request.GET.get('etablissement', 'tous')

        queryset = self.get_queryset()
        if statut_filter != 'tous':
            queryset = queryset.filter(statut=statut_filter)
        if etablissement_filter != 'tous':
            queryset = queryset.filter(etablissement_cnef__type_etablissement=etablissement_filter)

        data = [
            {
                'id': f.id,
                'name': f.nom_fichier,
                'user': f.etablissement_cnef.Nom_etablissement if f.etablissement_cnef else 'Inconnu',
                'date': f.date_import.isoformat(),
                'size': f"{f.fichier.size / 1024:.1f} KB" if f.fichier else "0 KB",
                'statut': f.get_statut_display(),
                'statut_code': f.statut,
            }
            for f in queryset
        ]
        return JsonResponse({'success': True, 'fichiers': data})

class EtablissementsListAPIView(ListView):
    model = Etablissement
    queryset = Etablissement.objects.all()

    def render_to_response(self, context, **response_kwargs):
        """Accepte uniquement GET pour lister les établissements"""
        
        if self.request.method != 'GET':
            return JsonResponse({
                'success': False,
                'message': 'Méthode non autorisée. Utilisez POST sur /api/etablissements/'
            }, status=405)
        
        # Debug: afficher tous les établissements d'abord
        all_etabs = self.get_queryset()
        logger.info(f"📊 Total établissements: {all_etabs.count()}")
        
        # Filtrer les non-chefs
        queryset = all_etabs.annotate(
            dernier_upload=Max('fichiers_imports__date_import')
        )
        
        data = [
            {
                'id': etab.id,
                'nom': etab.Nom_etablissement,
                'code': getattr(etab, 'code_etablissement', ''),
                'type': getattr(etab, 'type_etablissement', ''),
                'categorie': getattr(etab, 'categorie_emf', None) or '-',
                'signup': etab.date_creation.isoformat(),
                'lastUpload': etab.dernier_upload.isoformat() if etab.dernier_upload else '',
                'status': 'Actif' if etab.is_active else 'Inactif',
                'actif': etab.is_active
            }
            for etab in queryset
        ]
        
        return JsonResponse({
            'success': True,
            'etablissements': data,
            'utilisateurs': data  # Compatibilité ancienne API
        })

class FichierDeleteAPIView(LoginRequiredMixin, DeleteView):
    model = FichierImport
    pk_url_kwarg = 'fichier_id'

    def delete(self, request, *args, **kwargs):
        fichier = self.get_object()
        nom_fichier = fichier.nom_fichier
        fichier.delete()
        logger.info(f"Fichier {nom_fichier} supprimé par {request.user}")
        
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='SUPPRESSION_FICHIER',
            description=f"Suppression du fichier {nom_fichier}",
            etablissement=fichier.etablissement_cnef,
            request=request,
            donnees_supplementaires={
                'fichier_id': fichier.id,
                'nom_fichier': nom_fichier,
                'etablissement': fichier.etablissement_cnef.Nom_etablissement if fichier.etablissement_cnef else None
            }
        )
        return JsonResponse({
            'success': True,
            'message': f'Fichier "{nom_fichier}" supprimé avec succès'
        })

    def handle_no_permission(self):
        return JsonResponse({
            'success': False,
            'message': 'Non autorisé'
        }, status=403)

class EtablissementDeleteAPIView(LoginRequiredMixin, DeleteView):
    model = Etablissement
    pk_url_kwarg = 'etablissement_id'

    def delete(self, request, *args, **kwargs):
        etablissement = self.get_object()
        Nom_etablissement = etablissement.Nom_etablissement
        etablissement.delete()
        logger.info(f"Établissement {Nom_etablissement} supprimé par {request.user}")
        
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='SUPPRESSION_ETABLISSEMENT',
            description=f"Suppression de l'établissement {Nom_etablissement}",
            etablissement=etablissement,
            request=request,
            donnees_supplementaires={
                'etablissement_id': etablissement.id,
                'code_etablissement': etablissement.code_etablissement,
                'type_etablissement': etablissement.get_type_etablissement_display()
            }
        )
        return JsonResponse({
            'success': True,
            'message': f'Établissement "{Nom_etablissement}" supprimé avec succès'
        })

    def handle_no_permission(self):
        return JsonResponse({
            'success': False,
            'message': 'Non autorisé'
        }, status=403)
        
@user_passes_test(is_cnef_user)
@login_required
def interface_chef(request):
    """Interface principale du chef pour visualiser les soumissions"""
    statut_filter = request.GET.get('statut', 'tous')
    etablissement_filter = request.GET.get('etablissement', 'tous')
    
    fichiers = FichierImport.objects.all().order_by('-date_import')
    
    if statut_filter != 'tous':
        fichiers = fichiers.filter(statut=statut_filter)
    
    if etablissement_filter != 'tous':
        fichiers = fichiers.filter(etablissement_cnef__type_etablissement=etablissement_filter)
    
    stats = {
        'total': FichierImport.objects.count(),
        'en_attente': FichierImport.objects.filter(statut='EN_COURS').count(),
        'reussis': FichierImport.objects.filter(statut='REUSSI').count(),
        'erreur': FichierImport.objects.filter(statut='ERREUR').count(),
        'rejetes': FichierImport.objects.filter(statut='REJETE').count(),  # NOUVEAU
    }
    
    etablissements = Etablissement.objects.all()
    
    context = {
        'fichiers': fichiers,
        'stats': stats,
        'etablissements': etablissements,
        'statut_filter': statut_filter,
        'etablissement_filter': etablissement_filter,
        'title': 'Interface Chef - Gestion des soumissions'
    }
    return render(request, 'admin/interface_chef.html', context)

@user_passes_test(is_cnef_user)
@login_required
def detail_soumission(request, fichier_id):
    """Détail d'une soumission spécifique avec prévisualisation"""
    fichier = get_object_or_404(FichierImport, id=fichier_id)
    
    # Import de la fonction de prévisualisation
    from .utils import previsualiser_fichier_excel
    
    # Obtenir la prévisualisation du fichier
    preview_result = previsualiser_fichier_excel(fichier)
    
    donnees_importees = {
        'credits': fichier.nb_credits_importes if fichier.statut == 'REUSSI' else preview_result.get('credits', 0),
        'decouverts': fichier.nb_decouverts_importes if fichier.statut == 'REUSSI' else preview_result.get('decouverts', 0),
        'affacturages': fichier.nb_affacturages_importes if fichier.statut == 'REUSSI' else preview_result.get('affacturages', 0),
        'cautions': fichier.nb_cautions_importes if fichier.statut == 'REUSSI' else preview_result.get('cautions', 0),
        'effets': fichier.nb_effets_importes if fichier.statut == 'REUSSI' else preview_result.get('effets', 0),
        'spot': getattr(fichier, 'nb_spots_importes', 0) if fichier.statut == 'REUSSI' else preview_result.get('spot', 0),
    }
    
    # Analyser les erreurs pour créer un rapport d'incohérences
    incoherences = {
        'total': len(preview_result.get('erreurs', [])),
        'par_type': {},
        'details': preview_result.get('erreurs', [])
    }
    
    # Catégoriser les erreurs
    for erreur in preview_result.get('erreurs', []):
        if 'Date' in erreur or 'date' in erreur:
            incoherences['par_type']['dates'] = incoherences['par_type'].get('dates', 0) + 1
        elif 'colonnes' in erreur or 'colonne' in erreur:
            incoherences['par_type']['structure'] = incoherences['par_type'].get('structure', 0) + 1
        elif 'Type de feuille' in erreur:
            incoherences['par_type']['feuilles'] = incoherences['par_type'].get('feuilles', 0) + 1
        else:
            incoherences['par_type']['autres'] = incoherences['par_type'].get('autres', 0) + 1
    
    # Ajouter l'URL pour la visualisation des TEG
    teg_verification_url = reverse('visualisation_teg', kwargs={'fichier_id': fichier_id})

    context = {
        'fichier': fichier,
        'donnees_importees': donnees_importees,
        'preview_result': preview_result,
        'incoherences': incoherences,
        'total_lignes': preview_result.get('total_lignes', 0),
        'teg_verification_url': teg_verification_url,  
        'title': f'Détail - {fichier.nom_fichier}'
    }
    
    return render(request, 'admin/detail_soumission.html', context)

@login_required
def get_stats_ajax(request):
    """Récupérer les statistiques en AJAX"""
    stats = {
        'total': FichierImport.objects.count(),
        'en_attente': FichierImport.objects.filter(statut='EN_COURS').count(),
        'reussis': FichierImport.objects.filter(statut='REUSSI').count(),
    }
    
    top_etablissements = Etablissement.objects.annotate(
        nb_fichiers=Count('fichiers_imports')
    ).order_by('-nb_fichiers')[:5]
    
    etablissements_data = [
        {'nom': etab.type_etablissement, 'count': etab.nb_fichiers}
        for etab in top_etablissements
    ]
    
    return JsonResponse({
        'success': True,
        'stats': stats,
        'top_etablissements': etablissements_data
    })

@login_required
def upload_fichier_utilisateur(request):
    """Vue pour l'upload des fichiers par les utilisateurs"""

    if not request.FILES.get('fichier'):
        return JsonResponse({
            'success': False,
            'message': 'Aucun fichier fourni'
        }, status=400)
    
    fichier = request.FILES['fichier']
    
    # Validation de l'extension
    if not fichier.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({
            'success': False,
            'message': 'Seuls les fichiers Excel (.xlsx, .xls) sont autorisés'
        }, status=400)
    
    try:
        # Créer l'enregistrement
        fichier_import = FichierImport.objects.create(
            etablissement_cnef=request.user.etablissement,
            uploader_par=request.user,
            fichier=fichier,
            nom_fichier=fichier.name,
            statut='EN_COURS'
        )
        
        # ✅ CORRECTION : Notification avec le bon objet
        nb_notifications = envoyer_email_notification_acnef(fichier_import)
        
        # Log du résultat
        if nb_notifications > 0:
            logger.info(f"✅ {nb_notifications} administrateur(s) notifié(s) pour {fichier.name}")
        else:
            logger.warning(f"⚠️ Aucun administrateur notifié pour {fichier.name}")
        
        # Journaliser
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='UPLOAD_FICHIER',
            description=f"Upload du fichier {fichier.name}",
            etablissement=request.user.etablissement,
            request=request
        )
        
        # Message adapté selon le nombre de notifications
        if nb_notifications > 0:
            message = f'Fichier "{fichier.name}" soumis avec succès ! {nb_notifications} administrateur(s) notifié(s).'
        else:
            message = f'Fichier "{fichier.name}" soumis avec succès !'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'fichier_id': fichier_import.id
        })
    
    except Exception as e:
        logger.error(f"Erreur upload UEF: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de l\'upload : {str(e)}'
        }, status=500)

@login_required
def api_historique_etablissement(request):    
    """
    API commune pour UEF et AEF : retourne l'historique de l'établissement
    """
    if not request.user.etablissement:
        return JsonResponse({'success': False, 'message': 'Aucun établissement associé'})

    fichiers = FichierImport.objects.filter(
        etablissement_cnef=request.user.etablissement,
        uploader_par=request.user  # 🔥 DIFFÉRENCE CLÉ : Uniquement ses propres fichier s
    ).select_related('uploader_par', 'valide_par').order_by('-date_import')

    data = []
    for f in fichiers:
        # Calculer la taille du fichier de manière sécurisée
        try:
            file_size = f.fichier.size if f.fichier else 0
        except Exception:
            file_size = 0
        
        # Extraire la raison du rejet
        raison_rejet = None
        if f.statut == 'REJETE':
            raison_rejet = f.commentaire_validation or f.erreurs
        
        data.append({
            'id': f.id,
            'name': f.nom_fichier,  
            'size': file_size,  
            'date': f.date_import.isoformat(),  
            'status': f.get_statut_display(),  
            'statut_code': f.statut,
            'lignes_importees': f.total_lignes_importees,  
            'uploader': f.uploader_par.get_full_name() if f.uploader_par else 'Inconnu',
            'raison_rejet': raison_rejet,  
        })

    return JsonResponse({
        'success': True,
        'historique': data
    })
    
from django.db import transaction

@user_passes_test(is_cnef_user)
@login_required
def valider_soumission(request, fichier_id):
    """Valider une soumission"""
    if request.method != 'POST':
        logger.warning(f"Tentative d'accès non-POST à valider_soumission par {request.user}")
        return HttpResponse(status=405)
    
    fichier = get_object_or_404(FichierImport, id=fichier_id)
    
    if fichier.statut != 'EN_COURS':
        logger.warning(f"Tentative de validation d'un fichier déjà traité {fichier.nom_fichier} par {request.user}")
        return JsonResponse({
            'success': False,
            'message': 'Ce fichier a déjà été traité'
        }, status=400)
    
    try:
        with transaction.atomic():
            logger.debug(f"Début du traitement de {fichier.nom_fichier}")
            resultat = traiter_fichier_excel(fichier)
            logger.debug(f"Résultat du traitement: {resultat}")
            
            if resultat['success']:
                #Pour le message de validation
                envoyer_email_validation(fichier)
                
                ActionUtilisateur.enregistrer_action(
                    utilisateur=request.user,
                    type_action='VALIDATION_FICHIER',
                    description=f"Validation réussie du fichier {fichier.nom_fichier}",
                    etablissement=fichier.etablissement_cnef,
                    request=request,
                    donnees_supplementaires={
                        'fichier_id': fichier.id,
                        'nom_fichier': fichier.nom_fichier,
                        'total_lignes': resultat['total_lignes'],
                        'details': {
                            'credits': resultat['credits'],
                            'decouverts': resultat['decouverts'],
                            'affacturages': resultat['affacturages'],
                            'cautions': resultat['cautions'],
                            'effets': resultat['effets'],
                            'spots': resultat.get('spots', 0)
                        }
                    }
                )
                

                statut_display = fichier.get_statut_display()
                logger.info(f"Fichier {fichier.nom_fichier} validé par {request.user}: {resultat['total_lignes']} lignes importées")
                
               
                
                return JsonResponse({
                    'success': True,
                    'message': f"Fichier validé avec succès! {resultat['total_lignes']} lignes importées.",
                    'statut': statut_display,
                    'total_lignes': resultat['total_lignes'],
                    'details': {
                        'credits': resultat['credits'],
                        'decouverts': resultat['decouverts'],
                        'affacturages': resultat['affacturages'],
                        'cautions': resultat['cautions'],
                        'effets': resultat['effets'],
                    }
                })
                
                
                
            else:
                logger.error(f"Erreur lors de la validation de {fichier.nom_fichier}: {resultat['message']}")
                return JsonResponse({
                    'success': False,
                    'message': f"Erreur lors de l'import: {resultat['message']}"
                }, status=400)
    except Exception as e:
        logger.error(f"Erreur lors du traitement de {fichier.nom_fichier} par {request.user}: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f"Erreur lors du DDDDtraitement: {str(e)}"
        }, status=500)


@login_required
def rejeter_soumission(request, fichier_id):
    """
    Rejeter une soumission de fichier
    
    Cette fonction :
    1. Vérifie que le fichier existe et est EN_COURS
    2. Récupère le motif du rejet
    3. Change le statut à REJETE
    4. Enregistre qui a rejeté et quand
    5. Envoie des emails à l'AEF et l'UEF
    6. Journalise l'action
    """
    
    # ══════════════════════════════════════════════════════════════════════
    # 1. VÉRIFICATIONS DE SÉCURITÉ
    # ══════════════════════════════════════════════════════════════════════
    
    if request.method != 'POST':
        logger.warning(f"Tentative d'accès non-POST à rejeter_soumission par {request.user}")
        return HttpResponse(status=405)
    
    # Récupérer le fichier
    fichier = get_object_or_404(FichierImport, id=fichier_id)
    
    # Vérifier que le fichier n'a pas déjà été traité
    if fichier.statut != 'EN_COURS':
        logger.warning(
            f"Tentative de rejet d'un fichier déjà traité : "
            f"{fichier.nom_fichier} (statut={fichier.statut}) par {request.user}"
        )
        return JsonResponse({
            'success': False,
            'message': f'Ce fichier a déjà été traité (statut: {fichier.get_statut_display()})'
        }, status=400)
    
    # ══════════════════════════════════════════════════════════════════════
    # 2. RÉCUPÉRATION DU MOTIF DU REJET
    # ══════════════════════════════════════════════════════════════════════
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        logger.error(f"Corps JSON invalide dans rejeter_soumission par {request.user}")
        return JsonResponse({
            'success': False,
            'message': 'Corps de la requête JSON invalide'
        }, status=400)
    
    motif_rejet = data.get('raison', '').strip()
    
    # Vérifier qu'un motif a été fourni
    if not motif_rejet:
        logger.warning(f"Tentative de rejet sans motif par {request.user} pour {fichier.nom_fichier}")
        return JsonResponse({
            'success': False,
            'message': 'Veuillez fournir un motif de rejet'
        }, status=400)
    
    # ══════════════════════════════════════════════════════════════════════
    # 3. REJET DU FICHIER (TRANSACTION ATOMIQUE)
    # ══════════════════════════════════════════════════════════════════════
    
    try:
        with transaction.atomic():
            # Mise à jour du fichier
            fichier.statut = 'REJETE'
            fichier.valide_par = request.user  # ✅ QUI a rejeté
            fichier.date_validation = timezone.now()  # ✅ QUAND rejeté
            fichier.commentaire_validation = f"[REJET] {motif_rejet}"  # ✅ Utilisation du bon champ
            
            # Optionnel : garder aussi dans erreurs pour compatibilité
            fichier.erreurs = f"Fichier rejeté par {request.user.get_full_name()}"
            
            fichier.save()
            
            logger.info(
                f"✅ Fichier {fichier.nom_fichier} (ID:{fichier.id}) rejeté par "
                f"{request.user.get_full_name()} ({request.user.get_role_display()})"
            )
            
            # ══════════════════════════════════════════════════════════════
            # 4. JOURNALISATION DE L'ACTION
            # ══════════════════════════════════════════════════════════════
            
            ActionUtilisateur.enregistrer_action(
                utilisateur=request.user,
                type_action='REJET_FICHIER',
                description=f"Rejet du fichier {fichier.nom_fichier}",
                etablissement=fichier.etablissement_cnef,
                request=request,
                donnees_supplementaires={
                    'fichier_id': fichier.id,
                    'nom_fichier': fichier.nom_fichier,
                    'motif_rejet': motif_rejet,
                    'etablissement': fichier.etablissement_cnef.Nom_etablissement if fichier.etablissement_cnef else None,
                    'uploader': fichier.uploader_par.get_full_name() if fichier.uploader_par else None,
                }
            )
            
            # ══════════════════════════════════════════════════════════════
            # 5. ENVOI DES EMAILS DE NOTIFICATION
            # ══════════════════════════════════════════════════════════════
            
            try:
                resultat_emails = envoyer_email_rejet(fichier, motif_rejet)
                
                if resultat_emails['aef']:
                    logger.info(f"📧 Email de rejet envoyé à l'AEF")
                else:
                    logger.warning(f"⚠️ Email de rejet NON envoyé à l'AEF")
                
                if resultat_emails['uef']:
                    logger.info(f"📧 Email de rejet envoyé à l'UEF")
                else:
                    logger.warning(f"⚠️ Email de rejet NON envoyé à l'UEF")
                    
            except Exception as e:
                # L'email a échoué mais le rejet est quand même enregistré
                logger.error(f"❌ Erreur lors de l'envoi des emails de rejet : {str(e)}")
                # On ne fait pas échouer la transaction pour autant
        
        # ══════════════════════════════════════════════════════════════════
        # 6. RETOUR SUCCÈS
        # ══════════════════════════════════════════════════════════════════
        
        return JsonResponse({
            'success': True,
            'message': 'Fichier rejeté avec succès. Les responsables ont été notifiés par email.',
            'statut': fichier.get_statut_display(),
            'motif_rejet': motif_rejet,
            'rejet_par': request.user.get_full_name(),
            'date_rejet': fichier.date_validation.strftime('%d/%m/%Y à %H:%M')
        })
        
    except Exception as e:
        logger.error(
            f"❌ ERREUR CRITIQUE lors du rejet de {fichier.nom_fichier} (ID:{fichier.id}) "
            f"par {request.user}: {str(e)}",
            exc_info=True
        )
        return JsonResponse({
            'success': False,
            'message': f"Erreur lors du rejet : {str(e)}"
        }, status=500)

@user_passes_test(is_cnef_user)
@login_required
def visualiser_base_donnees(request, model_type):
    """Vue pour visualiser les bases de données par modèle"""
    MODEL_MAPPING = {
        'Credit_Amortissables': Credit_Amortissables,
        'Decouverts': Decouverts,
        'Affacturage': Affacturage,
        'Cautions': Cautions,
        'Effets_commerces': Effets_commerces,
        'Spots' : Spot,
    }
    
    if model_type not in MODEL_MAPPING:
        logger.warning(f"Modèle non trouvé: {model_type} par {request.user}")
        return JsonResponse({'success': False, 'message': 'Modèle non trouvé'}, status=400)
    
    model_class = MODEL_MAPPING[model_type]
    
    # Récupérer TOUTES les données sans filtrage initial
    queryset = model_class.objects.all()
    
    # Filtres
    sigle_filter = request.GET.get('sigle')
    mois_filter = request.GET.get('mois')
    annee_filter = request.GET.get('annee')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 25))

    # Déterminer le champ de date à utiliser
    date_field = 'DATE_MISE_PLACE_I03' if hasattr(model_class, 'DATE_MISE_PLACE_I03') else 'DATE_MEP_I03'
    

    # IMPORTANT: Si format=xlsx, on exporte SANS pagination
    if request.GET.get('format') == 'xlsx':
        # Appliquer les filtres si présents
        if sigle_filter:
            sigle_fields = ['SIGLE_I01', 'ETABLISSEMENT_I01']
            for field in sigle_fields:
                if hasattr(model_class, field):
                    queryset = queryset.filter(**{f'{field}__icontains': sigle_filter})
                    break
        
        if mois_filter:
            if hasattr(model_class, date_field):
                queryset = queryset.filter(**{f'{date_field}__month': mois_filter})
        
        # NOUVEAU: Filtre par année pour export
        if annee_filter:
            if hasattr(model_class, date_field):
                queryset = queryset.filter(**{f'{date_field}__year': annee_filter})
        
        return exporter_excel(queryset, model_type)
    
    # Appliquer les filtres pour l'affichage JSON
    if sigle_filter:
        sigle_fields = ['SIGLE_I01', 'ETABLISSEMENT_I01']
        filtered = False
        for field in sigle_fields:
            if hasattr(model_class, field):
                queryset = queryset.filter(**{f'{field}__icontains': sigle_filter})
                filtered = True
                break
        if not filtered:
            logger.warning(f"Aucun champ de filtrage par sigle disponible pour {model_type}")
            return JsonResponse({'success': False, 'message': 'Aucun champ de filtrage par sigle disponible'}, status=400)
    
    if mois_filter:
        if hasattr(model_class, date_field):
            queryset = queryset.filter(**{f'{date_field}__month': mois_filter})
        else:
            logger.warning(f"Aucun champ de filtrage par date disponible pour {model_type}")
            return JsonResponse({'success': False, 'message': 'Aucun champ de filtrage par date disponible'}, status=400)
    
    # NOUVEAU: Filtre par année
    if annee_filter:
        if hasattr(model_class, date_field):
            queryset = queryset.filter(**{f'{date_field}__year': annee_filter})
        else:
            logger.warning(f"Aucun champ de filtrage par année disponible pour {model_type}")
            return JsonResponse({'success': False, 'message': 'Aucun champ de filtrage par année disponible'}, status=400)
    

    # Récupérer TOUS les sigles distincts pour le filtre
    sigle_field = 'SIGLE_I01' if hasattr(model_class, 'SIGLE_I01') else 'ETABLISSEMENT_I01'
    sigles_distincts = []
    if hasattr(model_class, sigle_field):
        sigles_distincts = list(
            model_class.objects.values_list(sigle_field, flat=True)
            .distinct()
            .order_by(sigle_field)
        )
    
    # NOUVEAU: Récupérer toutes les années distinctes
    annees_distinctes = []
    if hasattr(model_class, date_field):
        annees_distinctes = list(
            model_class.objects.dates(date_field, 'year', order='DESC')
            .values_list(date_field, flat=True)
        )
        # Extraire uniquement l'année
        annees_distinctes = sorted(list(set([date.year for date in annees_distinctes])), reverse=True)
  
    # Pagination pour l'affichage (seulement pour JSON)
    total = queryset.count()
    start = (page - 1) * page_size
    end = start + page_size
    queryset_pagine = queryset[start:end]
    
    # Restreindre les champs sensibles
    safe_fields = [
        'SIGLE_I01', 'ETABLISSEMENT_I01', 'CODE_ETAB_I02', 'DATE_MEP_I03', 'CHA_ORI_I04', 'NATURE_PRET_I05', 
        'BENEFICAIRE_I06', 'CATEGORIE_BENEF_I07', 'LIEU_RESIDENCE_I08','SECT_ACT_I09', 'MONTANT_CHAF_I10', 
        'EFFECTIF_I11', 'PROFESSION_I12', 'MONTANT_PRET_I13', 'DUREE_I14', 'DUREE_DIFFERRE_I15', 'FREQ_REMB_I16', 
        'TAUX_NOMINAL_I17', 'FRAIS_DOSSIER_I18', 'MODALITEPAIEMENT_ASS_I19', 'MONTANTASSURANCE_I20', 
        'FRAIS_ANNEXE_I21', 'MODEREMBOURSEMENT_I22', 'MONTANT_ECHEANCE_I23', 'MODE_DEBLOCAGE_I24', 
        'SITUATION_CREANCE_I25', 'TEG_I26', 'CODE_BANQUE_I02', 'DATE_MISE_PLACE_I03', 'BENEFICAIRE_I04', 
        'CATEGORIE_BENEF_I05', 'LIEU_RESIDENCE_I06', 'SECT_ACT_I07', 'MONTANT_DECOUVERT_I08', 
        'CUMUL_TIRAGES_DEC_I09', 'TAUX_NOMINAL_I10', 'FRAIS_DOSSIERS_ET_COMM_I11', 'COUTS_ASSURANCE_I12', 
        'FRAIS_ANNEXES_I13', 'AGIOS_I14', 'NOMBRE_DEBITEURS_I15', 'SITUATION_CREANCE_I16', 'TEG_I17', 
        'DATE_ECHEANCE_I04', 'DUREE_AFFACTURAGE_I05', 'MONTANT_AFFACTURAGE_I10', 'TAUX_AFFACTURAGE_I11', 
        'MONTANT_FRAIS_COMM_I12', 'MONTANT_FRAIS_ANNEXES_I13', 'TEG_I14', 'DUREE_CAUTION_I05', 
        'MONTANT_CAUTION_I10', 'TAUX_CAUTION_I11', 'DUREE_EFFET_I05', 'MONTANT_EFFET_I10', 'TAUX_EFFET_I11',
        'AUTRES_FRA_I14', 'TEG_I15'
    ]
    available_fields = [f.name for f in model_class._meta.fields if f.name in safe_fields]
    donnees = list(queryset_pagine.values(*available_fields))
    
    logger.info(f"Visualisation des données de {model_type} par {request.user}, page {page}")
    return JsonResponse({
        'success': True,
        'donnees': donnees,
        'total': total,
        'page': page,
        'page_size': page_size,
        'model_type': model_type,
        'sigles_distincts': sigles_distincts,  # NOUVEAU: Liste de tous les sigles
        'annees_distinctes': annees_distinctes
    })

def exporter_excel(queryset, model_type):
    """Exporter les données en format Excel"""
    from datetime import datetime
    from django.utils import timezone as django_timezone
    
    # Récupérer les données
    data = list(queryset.values())
    
    # Convertir les datetimes avec timezone en datetimes sans timezone
    for row in data:
        for key, value in row.items():
            # Si c'est un datetime avec timezone, le convertir
            if isinstance(value, datetime) and django_timezone.is_aware(value):
                row[key] = django_timezone.make_naive(value)
    
    # Créer le DataFrame
    df = pd.DataFrame(data)
    
    # Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{model_type}_{django_timezone.now().date()}.xlsx"'
    
    # Écrire dans Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=model_type, index=False)
    
    logger.info(f"Export Excel de {model_type} effectué")
    return response

@user_passes_test(is_cnef_user)
@login_required
def telecharger_fichier_original(request, fichier_id):
    """Télécharger le fichier Excel original soumis par l'utilisateur"""
    fichier = get_object_or_404(FichierImport, id=fichier_id)
    
    try:
        if fichier.fichier and fichier.fichier.name:
            response = HttpResponse(fichier.fichier.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{fichier.nom_fichier}"'
            logger.info(f"Téléchargement du fichier original {fichier.nom_fichier} par {request.user}")
            return response
        else:
            logger.error(f"Fichier non trouvé pour l'import {fichier_id}")
            return JsonResponse({
                'success': False,
                'message': 'Fichier non disponible'
            }, status=404)
            
    except Exception as e:
        logger.error(f"Erreur lors du téléchargement du fichier {fichier_id} par {request.user}: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors du téléchargement: {str(e)}'
        }, status=500)
        
@user_passes_test(is_cnef_user)
@login_required
def supprimer_fichier_api(request, fichier_id):
    """Supprimer un fichier via API"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'message': 'Méthode non autorisée'
        }, status=405)
    
    fichier = get_object_or_404(FichierImport, id=fichier_id)
    nom_fichier = fichier.nom_fichier
    
    try:
        # Supprimer le fichier physique si nécessaire
        if fichier.fichier:
            fichier.fichier.delete(save=False)
        
        # Supprimer l'objet de la base de données
        fichier.delete()
        
        logger.info(f"Fichier {nom_fichier} supprimé par {request.user}")
        return JsonResponse({
            'success': True,
            'message': f'Fichier "{nom_fichier}" supprimé avec succès'
        })
        
    except Exception as e:
        logger.error(f"Erreur lors de la suppression du fichier {fichier_id} par {request.user}: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la suppression: {str(e)}'
        }, status=500)

@login_required
@user_passes_test(is_chef)
def supprimer_etablissement_api(request, etablissement_id):
    """Supprimer un établissement via API"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'message': 'Méthode non autorisée'
        }, status=405)
    
    etablissement = get_object_or_404(Etablissement, id=etablissement_id)
    Nom_etablissement = etablissement.Nom_etablissement
    
    try:
        # Vérifier s'il y a des fichiers associés
        fichiers_associes = FichierImport.objects.filter(etablissement=etablissement).exists()
        
        if fichiers_associes:
            return JsonResponse({
                'success': False,
                'message': f'Impossible de supprimer l\'établissement "{Nom_etablissement}" car il a des fichiers associés'
            }, status=400)
        
        # Supprimer l'établissement
        etablissement.delete()
        
        logger.info(f"Établissement {Nom_etablissement} supprimé par {request.user}")
        return JsonResponse({
            'success': True,
            'message': f'Établissement "{Nom_etablissement}" supprimé avec succès'
        })
        
    except Exception as e:
        logger.error(f"Erreur lors de la suppression de l'établissement {etablissement_id} par {request.user}: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la suppression: {str(e)}'
        }, status=500)
   

def verifier_teg_unifie(fichier_import):
    """
    Fonction UNIFIÉE pour vérifier les TEG 
    Fonctionne pour TOUS les statuts de fichier
    """
    resultat = {
        'credits': {'conformes': 0, 'non_conformes': 0, 'records': []},
        'decouverts': {'conformes': 0, 'non_conformes': 0, 'records': []},
        'affacturages': {'conformes': 0, 'non_conformes': 0, 'records': []},
        'cautions': {'conformes': 0, 'non_conformes': 0, 'records': []},
        'effets': {'conformes': 0, 'non_conformes': 0, 'records': []},
        'spots': {'conformes': 0, 'non_conformes': 0, 'records': []},
        'erreurs': []
    }
    
    TOLERANCE = 0.001
    
    # Configuration complète
    model_config = {
        'credits': {
            'model': Credit_Amortissables,
            'relation': 'credits_amortissables_files',
            'teg_calcule': 'TEG_annualise',
            'teg_original': 'TEG_I26'
        },
        'decouverts': {
            'model': Decouverts,
            'relation': 'decouverts',
            'teg_calcule': 'TEG_decouvert',
            'teg_original': 'TEG_I17'
        },
        'affacturages': {
            'model': Affacturage,
            'relation': 'affacturages',
            'teg_calcule': 'TEG_affacturage',
            'teg_original': 'TEG_I14'
        },
        'cautions': {
            'model': Cautions,
            'relation': 'cautions',
            'teg_calcule': 'TEG_caution',
            'teg_original': 'TEG_I14'
        },
        'effets': {
            'model': Effets_commerces,
            'relation': 'effets',
            'teg_calcule': 'TEG_effet',
            'teg_original': 'TEG_I15'
        },
        'spots': {
            'model': Spot,
            'relation': 'spot_files',
            'teg_calcule': 'TEG_spot',
            'teg_original': 'TEG_I26'
        }
    }
    
    for key, config in model_config.items():
        try:
            logger.info(f"\n{'='*60}")
            logger.info(f"🔍 ANALYSE: {key.upper()}")
            logger.info(f"{'='*60}")
            
            # STRATÉGIE 1: Essayer avec fichier_import directement
            queryset = config['model'].objects.filter(fichier_import=fichier_import)
            count = queryset.count()
            logger.info(f"  Stratégie 1 (fichier_import): {count} enregistrements")
            
            # STRATÉGIE 2: Si vide, essayer avec la relation
            if count == 0:
                try:
                    related_manager = getattr(fichier_import, config['relation'], None)
                    if related_manager:
                        queryset = related_manager.all()
                        count = queryset.count()
                        logger.info(f"  Stratégie 2 (relation '{config['relation']}'): {count} enregistrements")
                except Exception as e:
                    logger.warning(f"  Stratégie 2 échouée: {e}")
            
            # STRATÉGIE 3: Si toujours vide, essayer avec etablissement
            if count == 0:
                queryset = config['model'].objects.filter(etablissement=fichier_import.etablissement)
                count = queryset.count()
                logger.info(f"  Stratégie 3 (etablissement): {count} enregistrements")
            
            if count == 0:
                logger.warning(f"  ⚠️ Aucune donnée trouvée pour {key}")
                continue
            
            # Analyse des données
            conformes = 0
            non_conformes = 0
            teg_zero = 0
            teg_null = 0
            
            for obj in queryset:
                try:
                    teg_calcule = getattr(obj, config['teg_calcule'], None)
                    teg_original = getattr(obj, config['teg_original'], None)
                    
                    # Compteurs de diagnostic
                    if teg_calcule is None:
                        teg_null += 1
                        continue
                    if teg_original is None:
                        teg_null += 1
                        continue
                    
                    # Conversion
                    try:
                        teg_calc_val = float(teg_calcule) / 100
                        teg_orig_val = float(teg_original)
                    except (ValueError, TypeError) as e:
                        logger.debug(f"    Erreur conversion: {e}")
                        continue
                    
                    # Ignorer les zéros
                    if teg_calc_val == 0 and teg_orig_val == 0:
                        teg_zero += 1
                        continue
                    
                    # Comparaison
                    difference = abs(teg_calc_val - teg_orig_val)
                    
                    if difference <= TOLERANCE:
                        conformes += 1
                    else:
                        non_conformes += 1
                
                except Exception as e:
                    logger.debug(f"    Erreur objet {getattr(obj, 'id', '?')}: {e}")
                    continue
            
            # Mise à jour des résultats
            resultat[key]['conformes'] = conformes
            resultat[key]['non_conformes'] = non_conformes
            
            # Résumé
            total_analyse = conformes + non_conformes

            
            if total_analyse == 0 and count > 0:
                resultat['erreurs'].append(
                    f" {key}: {count} enregistrements trouvés mais aucun TEG valide"
                )
        
        except Exception as e:
            error_msg = f"Erreur {key}: {str(e)}"
            resultat['erreurs'].append(error_msg)
            logger.error(f"   {error_msg}")
            import traceback
            logger.error(traceback.format_exc())
    
    # Statistiques globales finales
    total_global = sum(
        resultat[key]['conformes'] + resultat[key]['non_conformes']
        for key in model_config.keys()
    )

    if total_global == 0:
        msg = " ATTENTION: Aucune donnée TEG valide trouvée !"
        resultat['erreurs'].append(msg)
        logger.warning(msg)
    
    return resultat


def preparer_contexte_visualisation(fichier, resultat_teg):
    """Prépare le contexte pour les templates - VERSION ROBUSTE"""
    type_labels = {
        'credits': 'Crédits Amortissables',
        'decouverts': 'Découverts', 
        'affacturages': 'Affacturages',
        'cautions': 'Cautions',
        'effets': 'Effets de Commerce',
        'spots': 'Spots'
    }
    
    charts = []
    total_conformes = 0
    total_non_conformes = 0
    
    
    # Préparer les graphiques pour TOUS les types
    for key in type_labels.keys():
        conformes = resultat_teg[key]['conformes']
        non_conformes = resultat_teg[key]['non_conformes']
        total = conformes + non_conformes
 
        
        # CORRECTION CRITIQUE: Créer un graphique même si total = 0 (pour le debug)
        # Mais ne l'afficher que si total > 0
        if total > 0:
            total_conformes += conformes
            total_non_conformes += non_conformes
            
            percent_conformes = round((conformes / total * 100), 2)
            percent_non_conformes = round((non_conformes / total * 100), 2)
            
            chart_data = {
                "type": "doughnut",
                "data": {
                    "labels": ['Conformes', 'Non conformes'],
                    "datasets": [{
                        "data": [conformes, non_conformes],
                        "backgroundColor": ["#4CAF50", "#DA1D10"],
                        "borderColor": ["#ffffff", "#ffffff"],
                        "borderWidth": 2,
                        "hoverBackgroundColor": ["#45a049", "#c41a0f"],
                        "hoverBorderColor": ["#ffffff", "#ffffff"],
                        "hoverBorderWidth": 3
                    }]
                },
                "options": {
                    "responsive": True,
                    "maintainAspectRatio": False,
                    "interaction": {
                        "mode": "index",
                        "intersect": False
                    },
                    "plugins": {
                        "legend": {
                            "position": "top",
                            "labels": {
                                "font": {"size": 12},
                                "padding": 15,
                                "usePointStyle": True
                            }
                        },
                        "title": {
                            "display": True,
                            "text": f"{type_labels[key]} - {total} enregistrements",
                            "font": {"size": 14, "weight": "bold"}
                        },
                        "tooltip": {
                            "enabled": True,
                            "mode": "index",
                            "intersect": False,
                            "backgroundColor": "rgba(0, 0, 0, 0.8)",
                            "padding": 12,
                            "displayColors": True
                        }
                    }
                }
            }
            
            charts.append({
                "chartjs": json.dumps(chart_data),
                "conformes": conformes,
                "non_conformes": non_conformes,
                "total": total,
                "percent_conformes": percent_conformes,
                "percent_non_conformes": percent_non_conformes,
                "title": type_labels[key],
                "key": key
            })
            
            logger.info(f"  Graphique créé pour {key}")
        else:
            logger.warning(f"  Pas de graphique pour {key} (aucune donnée)")
    
    # Calculer les statistiques globales
    total_global = total_conformes + total_non_conformes
    percent_conformes_global = round((total_conformes / total_global * 100), 2) if total_global > 0 else 0
    percent_non_conformes_global = round((total_non_conformes / total_global * 100), 2) if total_global > 0 else 0
    
    stats_globales = {
        'total_conformes': total_conformes,
        'total_non_conformes': total_non_conformes,
        'total_global': total_global,
        'percent_conformes': percent_conformes_global,
        'percent_non_conformes': percent_non_conformes_global,
        'taux_conformite': percent_conformes_global
    }
    
    # Créer un graphique global si des données existent
    if total_global > 0:
        chart_global = {
            "type": "doughnut",
            "data": {
                "labels": ['Conformes', 'Non conformes'],
                "datasets": [{
                    "data": [total_conformes, total_non_conformes],
                    "backgroundColor": ["#4CAF50", "#DA1D10"],
                    "borderColor": ["#ffffff", "#ffffff"],
                    "borderWidth": 3,
                    "hoverBackgroundColor": ["#45a049", "#c41a0f"],
                    "hoverBorderColor": ["#ffffff", "#ffffff"],
                    "hoverBorderWidth": 4
                }]
            },
            "options": {
                "responsive": True,
                "maintainAspectRatio": False,
                "interaction": {
                    "mode": "index",
                    "intersect": False
                },
                "plugins": {
                    "legend": {
                        "position": "top",
                        "labels": {
                            "font": {"size": 14},
                            "padding": 20,
                            "usePointStyle": True
                        }
                    },
                    "title": {
                        "display": True,
                        "text": f"Vue d'ensemble - {total_global} enregistrements",
                        "font": {"size": 16, "weight": "bold"}
                    },
                    "tooltip": {
                        "enabled": True,
                        "mode": "index",
                        "intersect": False,
                        "backgroundColor": "rgba(0, 0, 0, 0.85)",
                        "padding": 15,
                        "displayColors": True
                    }
                }
            }
        }
        stats_globales['chart_global'] = json.dumps(chart_global)
        logger.info(f"  Graphique global créé")
    else:
        logger.warning(f"  Pas de graphique global (aucune donnée)")
    
    logger.info(f"\n  TOTAL: {len(charts)} graphiques générés")
    logger.info(f"{'='*60}\n")
    
    return {
        'fichier': fichier,
        'charts': charts,
        'stats_globales': stats_globales,
        'erreurs': resultat_teg.get('erreurs', []),
        'has_data': len(charts) > 0,
        'title': f'Rapport TEG - {fichier.nom_fichier}',
    }

def preparer_contexte_visualisation(fichier, resultat_teg):
    """Prépare le contexte optimisé pour les templates avec tooltips dynamiques"""
    type_labels = {
        'credits': 'Crédits Amortissables',
        'decouverts': 'Découverts', 
        'affacturages': 'Affacturages',
        'cautions': 'Cautions',
        'effets': 'Effets de Commerce',
        'spots': 'Spots'
    }
    
    charts = []
    total_conformes = 0
    total_non_conformes = 0
    
    # Préparer les graphiques de manière optimisée
    for key in type_labels.keys():
        conformes = resultat_teg[key]['conformes']
        non_conformes = resultat_teg[key]['non_conformes']
        total = conformes + non_conformes
        
        if total > 0:
            total_conformes += conformes
            total_non_conformes += non_conformes
            
            percent_conformes = round((conformes / total * 100), 2) if total > 0 else 0
            percent_non_conformes = round((non_conformes / total * 100), 2) if total > 0 else 0
            
            # CORRECTION: Configuration Chart.js avec tooltips interactifs
            chart_data = {
                "type": "doughnut",
                "data": {
                    "labels": ['Conformes', 'Non conformes'],
                    "datasets": [{
                        "data": [conformes, non_conformes],
                        "backgroundColor": ["#4CAF50", "#DA1D10"],
                        "borderColor": ["#ffffff", "#ffffff"],
                        "borderWidth": 2,
                        "hoverBackgroundColor": ["#45a049", "#c41a0f"],  # Couleurs au survol
                        "hoverBorderColor": ["#ffffff", "#ffffff"],
                        "hoverBorderWidth": 3
                    }]
                },
                "options": {
                    "responsive": True,
                    "maintainAspectRatio": False,
                    "interaction": {
                        "mode": "nearest",
                        "intersect": True
                    },
                    "plugins": {
                        "legend": {
                            "position": "top",
                            "labels": {
                                "font": {
                                    "size": 12,
                                    "family": "'Segoe UI', Tahoma, Geneva, Verdana, sans-serif"
                                },
                                "padding": 15,
                                "usePointStyle": True
                            }
                        },
                        "title": {
                            "display": True,
                            "text": f"{type_labels[key]} - {total} enregistrements",
                            "font": {
                                "size": 14,
                                "weight": "bold",
                                "family": "'Segoe UI', Tahoma, Geneva, Verdana, sans-serif"
                            },
                            "padding": {
                                "top": 10,
                                "bottom": 10
                            }
                        },
                        "tooltip": {
                            "enabled": True,
                            "backgroundColor": "rgba(0, 0, 0, 0.8)",
                            "titleColor": "#ffffff",
                            "bodyColor": "#ffffff",
                            "borderColor": "#ffffff",
                            "borderWidth": 1,
                            "padding": 12,
                            "displayColors": True,
                            "titleFont": {
                                "size": 14,
                                "weight": "bold"
                            },
                            "bodyFont": {
                                "size": 13
                            },
                            "callbacks": {
                                "label": f"""function(context) {{
                                    const label = context.label || '';
                                    const value = context.parsed || 0;
                                    const total = {total};
                                    const percentage = ((value / total) * 100).toFixed(2);
                                    return label + ': ' + value + ' (' + percentage + '%)';
                                }}"""
                            }
                        }
                    },
                    "animation": {
                        "animateRotate": True,
                        "animateScale": True,
                        "duration": 1000
                    }
                }
            }
            
            charts.append({
                "chartjs": json.dumps(chart_data),
                "conformes": conformes,
                "non_conformes": non_conformes,
                "total": total,
                "percent_conformes": percent_conformes,
                "percent_non_conformes": percent_non_conformes,
                "title": type_labels[key],
                "key": key,
                "records": resultat_teg[key].get('records', [])
            })
    
    # Calculer les statistiques globales
    total_global = total_conformes + total_non_conformes
    percent_conformes_global = round((total_conformes / total_global * 100), 2) if total_global > 0 else 0
    percent_non_conformes_global = round((total_non_conformes / total_global * 100), 2) if total_global > 0 else 0
    
    stats_globales = {
        'total_conformes': total_conformes,
        'total_non_conformes': total_non_conformes,
        'total_global': total_global,
        'percent_conformes': percent_conformes_global,
        'percent_non_conformes': percent_non_conformes_global,
        'taux_conformite': percent_conformes_global
    }
    
    # Créer un graphique global récapitulatif avec tooltips
    if total_global > 0:
        chart_global = {
            "type": "doughnut",
            "data": {
                "labels": ['Conformes', 'Non conformes'],
                "datasets": [{
                    "data": [total_conformes, total_non_conformes],
                    "backgroundColor": ["#4CAF50", "#DA1D10"],
                    "borderColor": ["#ffffff", "#ffffff"],
                    "borderWidth": 3,
                    "hoverBackgroundColor": ["#45a049", "#c41a0f"],
                    "hoverBorderColor": ["#ffffff", "#ffffff"],
                    "hoverBorderWidth": 4
                }]
            },
            "options": {
                "responsive": True,
                "maintainAspectRatio": False,
                "interaction": {
                    "mode": "nearest",
                    "intersect": True
                },
                "plugins": {
                    "legend": {
                        "position": "top",
                        "labels": {
                            "font": {
                                "size": 14,
                                "family": "'Segoe UI', Tahoma, Geneva, Verdana, sans-serif"
                            },
                            "padding": 20,
                            "usePointStyle": True
                        }
                    },
                    "title": {
                        "display": True,
                        "text": f"Vue d'ensemble - {total_global} enregistrements",
                        "font": {
                            "size": 16,
                            "weight": "bold",
                            "family": "'Segoe UI', Tahoma, Geneva, Verdana, sans-serif"
                        },
                        "padding": {
                            "top": 15,
                            "bottom": 15
                        }
                    },
                    "tooltip": {
                        "enabled": True,
                        "backgroundColor": "rgba(0, 0, 0, 0.85)",
                        "titleColor": "#ffffff",
                        "bodyColor": "#ffffff",
                        "borderColor": "#ffffff",
                        "borderWidth": 2,
                        "padding": 15,
                        "displayColors": True,
                        "titleFont": {
                            "size": 15,
                            "weight": "bold"
                        },
                        "bodyFont": {
                            "size": 14
                        },
                        "callbacks": {
                            "label": f"""function(context) {{
                                const label = context.label || '';
                                const value = context.parsed || 0;
                                const total = {total_global};
                                const percentage = ((value / total) * 100).toFixed(2);
                                return label + ': ' + value + ' enregistrements (' + percentage + '%)';
                            }}""",
                            "afterLabel": """function(context) {
                                return '━━━━━━━━━━━━━━━';
                            }""",
                            "footer": f"""function(tooltipItems) {{
                                return 'Total: {total_global} enregistrements';
                            }}"""
                        }
                    }
                },
                "animation": {
                    "animateRotate": True,
                    "animateScale": True,
                    "duration": 1200
                }
            }
        }
        stats_globales['chart_global'] = json.dumps(chart_global)
    
    return {
        'fichier': fichier,
        'charts': charts,
        'stats_globales': stats_globales,
        'erreurs': resultat_teg.get('erreurs', []),
        'has_data': len(charts) > 0,
        'title': f'Rapport TEG - {fichier.nom_fichier}',
    }

def custom_404_view(request, exception=None):
    """
    Vue personnalisée pour les erreurs 404 - Chemins incorrects
    """
    context = {
        'title': 'Chemin incorrect - CNEF',
        'error_message': 'La page que vous recherchez n\'existe pas ou a été déplacée.',
        'redirect_url': 'connexion',
        'redirect_label': 'Se connecter'
    }
    return render(request, 'errors/404.html', context, status=404)

def custom_500_view(request):
    """
    Vue personnalisée pour les erreurs 500 - Erreurs serveur
    """
    context = {
        'title': 'Erreur serveur - CNEF', 
        'error_message': 'Une erreur interne du serveur s\'est produite.',
        'redirect_url': 'connexion',
        'redirect_label': 'Se connecter'
    }
    return render(request, 'errors/500.html', context, status=500)

def custom_permission_denied_view(request, exception=None):
    """
    Vue personnalisée pour les accès refusés (403)
    """
    context = {
        'title': 'Accès refusé - CNEF',
        'error_message': 'Vous n\'avez pas les permissions nécessaires pour accéder à cette page.',
        'redirect_url': 'connexion',
        'redirect_label': 'Se connecter'
    }
    return render(request, 'errors/403.html', context, status=403)

def custom_bad_request_view(request, exception=None):
    """
    Vue personnalisée pour les mauvaises requêtes (400)
    """
    context = {
        'title': 'Requête invalide - CNEF',
        'error_message': 'La requête envoyée au serveur est invalide.',
        'redirect_url': 'connexion', 
        'redirect_label': 'Se connecter'
    }
    return render(request, 'errors/400.html', context, status=400)


class ErrorHandlerMiddleware:
    """
    Middleware personnalisé pour une meilleure gestion des erreurs
    """
    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        response = self.get_response(request)
        return response

    def process_exception(self, request, exception):
        """
        Intercepte toutes les exceptions non gérées
        """
        logger.error(f"Exception non gérée: {str(exception)}")
        return custom_500_view(request)
    

import json
import logging
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import FichierImport
from .utils import extraire_et_calculer_teg, generer_statistiques_teg

logger = logging.getLogger(__name__)


@login_required
@user_passes_test(is_cnef_user)
def visualisation_teg(request, fichier_id):
    """
    Vue simplifiée pour l'analyse TEG
    Extraction + Calcul + Comparaison en UNE SEULE FOIS
    """
    fichier = get_object_or_404(FichierImport, id=fichier_id)
    
    try:
        # 1. EXTRACTION + CALCUL DES TEG
        logger.info(f"🔍 Début analyse TEG pour {fichier.nom_fichier}")
        
        donnees_extraites = extraire_et_calculer_teg(
            fichier.fichier.path,
            fichier.etablissement_cnef
        )
        
        # 2. GÉNÉRATION DES STATISTIQUES
        stats = generer_statistiques_teg(donnees_extraites)
        
        # 3. PRÉPARATION DES GRAPHIQUES
        context = preparer_graphiques_teg(fichier, donnees_extraites, stats)
        
        logger.info(f" Analyse terminée: {stats['global']['total']} enregistrements")
        
        return render(request, 'admin/verification_teg.html', context)
        
    except Exception as e:
        logger.error(f"❌ Erreur analyse TEG: {e}")
        
        context = {
            'fichier': fichier,
            'charts': [],
            'stats_globales': {},
            'erreurs': [f"Erreur lors de l'analyse: {str(e)}"],
            'has_data': False,
            'title': f'Rapport TEG - {fichier.nom_fichier}'
        }
        
        return render(request, 'admin/verification_teg.html', context)

def preparer_graphiques_teg(fichier, donnees_extraites, stats):
    """
    Prépare les graphiques pour l'affichage
    """
    type_labels = {
        'credits': 'Crédits Amortissables',
        'decouverts': 'Découverts',
        'affacturages': 'Affacturages',
        'cautions': 'Cautions',
        'effets': 'Effets de Commerce',
        'spots': 'Spots'
    }
    
    charts = []
    
    # Créer un graphique pour chaque type de produit
    for key, label in type_labels.items():
        stat = stats.get(key, {})
        conformes = stat.get('conformes', 0)
        non_conformes = stat.get('non_conformes', 0)
        total = stat.get('total', 0)
        
        if total == 0:
            continue
        
        percent_conformes = round((conformes / total * 100), 2)
        percent_non_conformes = round((non_conformes / total * 100), 2)
        
        chart_data = {
            "type": "doughnut",
            "data": {
                "labels": ['Conformes', 'Non conformes'],
                "datasets": [{
                    "data": [conformes, non_conformes],
                    "backgroundColor": ["#4CAF50", "#DA1D10"],
                    "borderColor": ["#ffffff", "#ffffff"],
                    "borderWidth": 2
                }]
            },
            "options": {
                "responsive": True,
                "maintainAspectRatio": False,
                "plugins": {
                    "legend": {
                        "position": "top",
                        "labels": {
                            "font": {"size": 12},
                            "padding": 15
                        }
                    },
                    "title": {
                        "display": True,
                        "text": f"{label} - {total} enregistrements",
                        "font": {"size": 14, "weight": "bold"}
                    },
                    "tooltip": {
                        "enabled": True,
                        "backgroundColor": "rgba(0, 0, 0, 0.8)",
                        "callbacks": {
                            "label": f"function(context) {{ const label = context.label || ''; const value = context.parsed || 0; const percentage = ((value / {total}) * 100).toFixed(2); return label + ': ' + value + ' (' + percentage + '%)'; }}"
                        }
                    }
                }
            }
        }
        
        charts.append({
            "chartjs": json.dumps(chart_data),
            "conformes": conformes,
            "non_conformes": non_conformes,
            "total": total,
            "percent_conformes": percent_conformes,
            "percent_non_conformes": percent_non_conformes,
            "title": label,
            "key": key
        })
    
    # Statistiques globales
    stats_global = stats.get('global', {})
    total_global = stats_global.get('total', 0)
    conformes_global = stats_global.get('conformes', 0)
    non_conformes_global = stats_global.get('non_conformes', 0)
    
    stats_globales = {
        'total_conformes': conformes_global,
        'total_non_conformes': non_conformes_global,
        'total_global': total_global,
        'percent_conformes': round((conformes_global / total_global * 100), 2) if total_global > 0 else 0,
        'percent_non_conformes': round((non_conformes_global / total_global * 100), 2) if total_global > 0 else 0,
        'taux_conformite': stats_global.get('taux_conformite', 0)
    }
    
    # Graphique global
    if total_global > 0:
        chart_global = {
            "type": "doughnut",
            "data": {
                "labels": ['Conformes', 'Non conformes'],
                "datasets": [{
                    "data": [conformes_global, non_conformes_global],
                    "backgroundColor": ["#4CAF50", "#DA1D10"],
                    "borderColor": ["#ffffff", "#ffffff"],
                    "borderWidth": 3
                }]
            },
            "options": {
                "responsive": True,
                "maintainAspectRatio": False,
                "plugins": {
                    "legend": {
                        "position": "top",
                        "labels": {
                            "font": {"size": 14},
                            "padding": 20
                        }
                    },
                    "title": {
                        "display": True,
                        "text": f"Vue d'ensemble - {total_global} enregistrements",
                        "font": {"size": 16, "weight": "bold"}
                    },
                    "tooltip": {
                        "enabled": True,
                        "backgroundColor": "rgba(0, 0, 0, 0.85)",
                        "callbacks": {
                            "label": f"function(context) {{ const label = context.label || ''; const value = context.parsed || 0; const percentage = ((value / {total_global}) * 100).toFixed(2); return label + ': ' + value + ' (' + percentage + '%)'; }}"
                        }
                    }
                }
            }
        }
        stats_globales['chart_global'] = json.dumps(chart_global)
    
    return {
        'fichier': fichier,
        'charts': charts,
        'stats_globales': stats_globales,
        'erreurs': donnees_extraites.get('erreurs', []),
        'has_data': len(charts) > 0,
        'title': f'Rapport TEG - {fichier.nom_fichier}'
    }

@login_required
@user_passes_test(is_aef)
def visualisation_teg_aef(request, fichier_id):
    """
    Vue pour AEF : Vérification TEG d'un fichier spécifique de SON établissement
    Réutilise exactement la même logique que visualisation_teg mais avec contrôle d'accès AEF
    """
    fichier = get_object_or_404(
        FichierImport,
        id=fichier_id,
        etablissement_cnef=request.user.etablissement,
    )

    # Réutilisation directe de la fonction existante
    from .utils import extraire_et_calculer_teg, generer_statistiques_teg

    try:
        # 1. EXTRACTION + CALCUL DES TEG
        logger.info(f"🔍 Début analyse TEG pour {fichier.nom_fichier} (AEF)")
        
        donnees_extraites = extraire_et_calculer_teg(
            fichier.fichier.path,
            fichier.etablissement_cnef
        )
        
        # 2. GÉNÉRATION DES STATISTIQUES
        stats = generer_statistiques_teg(donnees_extraites)
        
        # 3. PRÉPARATION DES GRAPHIQUES
        context = preparer_graphiques_teg(fichier, donnees_extraites, stats)
        
        # Ajouter le flag is_aef
        context['is_aef'] = True
        
        logger.info(f" Analyse TEG AEF terminée: {stats['global']['total']} enregistrements")

        # Journalisation
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='AUTRE',
            description=f"Consultation rapport TEG : {fichier.nom_fichier}",
            etablissement=request.user.etablissement,
            request=request
        )

        return render(request, 'EF/verification_teg_ef.html', context)

    except Exception as e:
        logger.error(f"Erreur TEG AEF (fichier {fichier_id}): {str(e)}")
        messages.error(request, "Erreur lors du calcul des TEG. Contactez l'administrateur.")
        return redirect('aef_detail_soumission', fichier_id=fichier_id)   

@login_required
@user_passes_test(is_cnef_user)
def generer_communique_presse(request):
    """
    Vue CORRIGÉE pour générer le communiqué de presse avec statistiques
    """
    # Récupérer les paramètres de filtrage
    trimestre = request.GET.get('trimestre', 'T2')
    annee = request.GET.get('annee', str(timezone.now().year))
    type_etablissement = request.GET.get('type_etablissement', 'EMF Deuxième catégorie')

    # Calculer les données
    context = calculer_donnees_communique(trimestre, annee, type_etablissement)
    
    return render(request, 'admin/communique_presse.html', context)

@login_required
@user_passes_test(is_cnef_user)
def details_supplementaires(request):
    """
    Vue pour afficher la page des détails supplémentaires avec les mêmes calculs
    """
    # Récupérer les paramètres de filtrage
    trimestre = request.GET.get('trimestre', 'T2')
    annee = request.GET.get('annee', str(timezone.now().year))
    type_etablissement = request.GET.get('type_etablissement', 'EMF Deuxième catégorie')

    # Calculer les données (même logique que generer_communique_presse)
    context = calculer_donnees_communique(trimestre, annee, type_etablissement)
    
    return render(request, 'admin/details_supplementaires.html', context)

def calculer_donnees_communique(trimestre, annee, type_etablissement):
    """
    Fonction partagée pour calculer les données du communiqué et des détails
    """
    import logging
    from django.db.models import Q, Avg, Sum, Count
    from datetime import datetime
    
    logger = logging.getLogger(__name__)
    
    # Options de filtrage
    trimestres = ['T1', 'T2', 'T3', 'T4']
    types_etablissement = ['Banques', 'EMF Première catégorie', 'EMF Deuxième catégorie']

    # Extraire les années disponibles
    annees = set()
    for model in [Credit_Amortissables, Spot, Decouverts, Affacturage, Cautions, Effets_commerces]:
        date_field = 'DATE_MEP_I03' if hasattr(model, 'DATE_MEP_I03') else 'DATE_MISE_PLACE_I03'
        years = model.objects.dates(date_field, 'year')
        annees.update([y.year for y in years])
    
    annees = sorted(list(annees), reverse=True)

    # Mapper le type d'établissement
    type_mapping = {
        'Banques': 'BANQUE',
        'EMF Première catégorie': 'EMF',
        'EMF Deuxième catégorie': 'EMF',
        'EMF Troisième catégorie': 'EMF'
    }
    
    # Déterminer la catégorie EMF si applicable
    categorie_emf = None
    if type_etablissement == 'EMF Première catégorie':
        categorie_emf = 'PREMIERE_CATEGORIE'
    elif type_etablissement == 'EMF Deuxième catégorie':
        categorie_emf = 'DEUXIEME_CATEGORIE'
    elif type_etablissement == 'EMF Troisième catégorie':
        categorie_emf = 'TROISIEME_CATEGORIE'


    # CORRECTION MAJEURE : Codes de catégories normalisés
    categories_beneficiaires = {
        'Particuliers': ['6', '06'],
        'Petites et Moyennes Entreprises': ['3-2', '3_2', '3 2', '32'],
        'Grandes Entreprises': ['3-1', '3_1', '3 1', '31'],
        'Administrations publiques et collectivités locales': ['1', '01'],
        'Autres personnes morales': ['2', '3', '4', '5', '7', '02', '03', '04', '05', '07']
    }

    # Calculer le trimestre
    trimestre_ranges = {
        'T1': (1, 3),
        'T2': (4, 6),
        'T3': (7, 9),
        'T4': (10, 12)
    }
    start_month, end_month = trimestre_ranges.get(trimestre, (4, 6))
    start_date = datetime(int(annee), start_month, 1)
    end_date = datetime(int(annee), end_month + 1, 1) if end_month < 12 else datetime(int(annee) + 1, 1, 1)

    # FONCTION AMÉLIORÉE pour calculer les statistiques
    def calculer_stats_amelioree(model, date_field, categorie_field, teg_field, 
                                  montant_field, taux_nominal_field, 
                                  category_codes, additional_filters=Q()):
        """
        Fonction améliorée avec meilleure gestion des jointures et des filtres
        """
        try:
            # 1. FILTRE DE BASE : Type d'établissement
            base_filter = Q(
                **{
                    f'{date_field}__gte': start_date,
                    f'{date_field}__lt': end_date,
                    'etablissement__type_etablissement': type_mapping.get(type_etablissement, 'EMF')
                }
            )
            
            # 2. FILTRE CATÉGORIE EMF (si applicable)
            if categorie_emf:
                base_filter &= Q(etablissement__categorie_emf=categorie_emf)
            
            # 3. FILTRE CATÉGORIE BÉNÉFICIAIRE (CORRIGÉ)
            category_filter = Q()
            for code in category_codes:
                # Recherche exacte ET recherche partielle
                category_filter |= Q(**{f'{categorie_field}': code})
                category_filter |= Q(**{f'{categorie_field}__icontains': code})
            
            # 4. COMBINAISON DE TOUS LES FILTRES
            queryset = model.objects.filter(
                base_filter & category_filter & additional_filters
            ).exclude(
                **{f'{teg_field}__isnull': True}
            ).exclude(
                **{f'{teg_field}': 0}
            )
            
            # 5. DEBUG : Compter les résultats
            count = queryset.count()

            if count == 0:
                return None
            
            # 6. AGRÉGATIONS
            stats = queryset.aggregate(
                teg_moyen=Avg(teg_field),
                montant_total=Sum(montant_field),
                count=Count('id')
            )
            
            result = {
                'teg_moyen': float(stats['teg_moyen'] or 0),
                'montant_total': float(stats['montant_total'] or 0),
                'count': stats['count']
            }
            
            # 7. TAUX NOMINAL (si disponible)
            if taux_nominal_field:
                taux_stats = queryset.aggregate(avg=Avg(taux_nominal_field))
                result['taux_nominal_moyen'] = float(taux_stats['avg'] or 0)
            else:
                result['taux_nominal_moyen'] = 0
            
            logger.info(f"    TEG moyen: {result['teg_moyen']:.2f}%, Montant: {result['montant_total']:,.0f}")
            
            return result
            
        except Exception as e:
            logger.error(f"❌ Erreur dans calculer_stats_amelioree pour {model.__name__}: {e}")
            return None

    # Structure des crédits avec configuration détaillée
    structure_credits = {
        'Particuliers': {
            'Crédits à la consommation, autre que découvert': {
                'model': Credit_Amortissables,
                'date_field': 'DATE_MEP_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_annualise',
                'montant_field': 'MONTANT_PRET_I13',
                'taux_nominal_field': 'TAUX_NOMINAL_I17',
                'filters': Q(NATURE_PRET_I05__in=['2', '02', 'Consommation'])
            },
            'Découverts': {
                'model': Decouverts,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I05',
                'teg_field': 'TEG_decouvert',
                'montant_field': 'MONTANT_DECOUVERT_I08',
                'taux_nominal_field': 'TAUX_NOMINAL_I10',
                'filters': Q()
            },
            'Crédits à moyens terme': {
                'model': Credit_Amortissables,
                'date_field': 'DATE_MEP_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_annualise',
                'montant_field': 'MONTANT_PRET_I13',
                'taux_nominal_field': 'TAUX_NOMINAL_I17',
                'filters': Q(MATURITE='2-MT')
            },
            'Crédits à long terme': {
                'model': Credit_Amortissables,
                'date_field': 'DATE_MEP_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_annualise',
                'montant_field': 'MONTANT_PRET_I13',
                'taux_nominal_field': 'TAUX_NOMINAL_I17',
                'filters': Q(MATURITE='3-LT')
            },
            'Crédits immobilier': {
                'model': Credit_Amortissables,
                'date_field': 'DATE_MEP_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_annualise',
                'montant_field': 'MONTANT_PRET_I13',
                'taux_nominal_field': 'TAUX_NOMINAL_I17',
                'filters': Q(NATURE_PRET_I05__in=['3', '03', 'Immobilier'])
            },
            'Cautions': {
                'model': Cautions,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_caution',
                'montant_field': 'MONTANT_CAUTION_I10',
                'taux_nominal_field': 'TAUX_CAUTION_I11',
                'filters': Q()
            },
            'Effets commerciaux': {
                'model': Effets_commerces,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_effet',
                'montant_field': 'MONTANT_EFFET_I11',
                'taux_nominal_field': 'TAUX_NOMINAL_I10',
                'filters': Q()
            },
            'Affacturage': {
                'model': Affacturage,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_affacturage',
                'montant_field': 'MONTANT_CREANCE_I10',
                'taux_nominal_field': None,
                'filters': Q()
            }
        },
        'Petites et Moyennes Entreprises': {
            'Crédits de trésorerie, autre que découvert': {
                'models': [
                    {
                        'model': Credit_Amortissables,
                        'date_field': 'DATE_MEP_I03',
                        'categorie_field': 'CATEGORIE_BENEF_I07',
                        'teg_field': 'TEG_annualise',
                        'montant_field': 'MONTANT_PRET_I13',
                        'taux_nominal_field': 'TAUX_NOMINAL_I17',
                        'filters': Q(MATURITE='1-CT')
                    },
                    {
                        'model': Spot,
                        'date_field': 'DATE_MEP_I03',
                        'categorie_field': 'CATEGORIE_BENEF_I07',
                        'teg_field': 'TEG_spot',
                        'montant_field': 'MONTANT_PRET_I13',
                        'taux_nominal_field': 'TAUX_NOMINAL_I17',
                        'filters': Q()
                    }
                ]
            },
            'Découverts': {
                'model': Decouverts,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I05',
                'teg_field': 'TEG_decouvert',
                'montant_field': 'MONTANT_DECOUVERT_I08',
                'taux_nominal_field': 'TAUX_NOMINAL_I10',
                'filters': Q()
            },
            'Crédits à moyens terme': {
                'model': Credit_Amortissables,
                'date_field': 'DATE_MEP_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_annualise',
                'montant_field': 'MONTANT_PRET_I13',
                'taux_nominal_field': 'TAUX_NOMINAL_I17',
                'filters': Q(MATURITE='2-MT')
            },
            'Crédits à long terme': {
                'model': Credit_Amortissables,
                'date_field': 'DATE_MEP_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_annualise',
                'montant_field': 'MONTANT_PRET_I13',
                'taux_nominal_field': 'TAUX_NOMINAL_I17',
                'filters': Q(MATURITE='3-LT')
            },
            'Cautions': {
                'model': Cautions,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_caution',
                'montant_field': 'MONTANT_CAUTION_I10',
                'taux_nominal_field': 'TAUX_CAUTION_I11',
                'filters': Q()
            },
            'Effets commerciaux': {
                'model': Effets_commerces,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_effet',
                'montant_field': 'MONTANT_EFFET_I11',
                'taux_nominal_field': 'TAUX_NOMINAL_I10',
                'filters': Q()
            },
            'Affacturage': {
                'model': Affacturage,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_affacturage',
                'montant_field': 'MONTANT_CREANCE_I10',
                'taux_nominal_field': None,
                'filters': Q()
            }
        },
        'Grandes Entreprises': {
            'Crédits de trésorerie, autre que découvert': {
                'models': [
                    {
                        'model': Credit_Amortissables,
                        'date_field': 'DATE_MEP_I03',
                        'categorie_field': 'CATEGORIE_BENEF_I07',
                        'teg_field': 'TEG_annualise',
                        'montant_field': 'MONTANT_PRET_I13',
                        'taux_nominal_field': 'TAUX_NOMINAL_I17',
                        'filters': Q(MATURITE='1-CT')
                    },
                    {
                        'model': Spot,
                        'date_field': 'DATE_MEP_I03',
                        'categorie_field': 'CATEGORIE_BENEF_I07',
                        'teg_field': 'TEG_spot',
                        'montant_field': 'MONTANT_PRET_I13',
                        'taux_nominal_field': 'TAUX_NOMINAL_I17',
                        'filters': Q()
                    }
                ]
            },
            'Découverts': {
                'model': Decouverts,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I05',
                'teg_field': 'TEG_decouvert',
                'montant_field': 'MONTANT_DECOUVERT_I08',
                'taux_nominal_field': 'TAUX_NOMINAL_I10',
                'filters': Q()
            },
            'Crédits à moyens terme': {
                'model': Credit_Amortissables,
                'date_field': 'DATE_MEP_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_annualise',
                'montant_field': 'MONTANT_PRET_I13',
                'taux_nominal_field': 'TAUX_NOMINAL_I17',
                'filters': Q(MATURITE='2-MT')
            },
            'Crédits à long terme': {
                'model': Credit_Amortissables,
                'date_field': 'DATE_MEP_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_annualise',
                'montant_field': 'MONTANT_PRET_I13',
                'taux_nominal_field': 'TAUX_NOMINAL_I17',
                'filters': Q(MATURITE='3-LT')
            },
            'Cautions': {
                'model': Cautions,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_caution',
                'montant_field': 'MONTANT_CAUTION_I10',
                'taux_nominal_field': 'TAUX_CAUTION_I11',
                'filters': Q()
            },
            'Effets commerciaux': {
                'model': Effets_commerces,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_effet',
                'montant_field': 'MONTANT_EFFET_I11',
                'taux_nominal_field': 'TAUX_NOMINAL_I10',
                'filters': Q()
            },
            'Affacturage': {
                'model': Affacturage,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_affacturage',
                'montant_field': 'MONTANT_CREANCE_I10',
                'taux_nominal_field': None,
                'filters': Q()
            }
        },
        'Administrations publiques et collectivités locales': {
            'Crédits de trésorerie, autre que découvert': {
                'models': [
                    {
                        'model': Credit_Amortissables,
                        'date_field': 'DATE_MEP_I03',
                        'categorie_field': 'CATEGORIE_BENEF_I07',
                        'teg_field': 'TEG_annualise',
                        'montant_field': 'MONTANT_PRET_I13',
                        'taux_nominal_field': 'TAUX_NOMINAL_I17',
                        'filters': Q(MATURITE='1-CT')
                    },
                    {
                        'model': Spot,
                        'date_field': 'DATE_MEP_I03',
                        'categorie_field': 'CATEGORIE_BENEF_I07',
                        'teg_field': 'TEG_spot',
                        'montant_field': 'MONTANT_PRET_I13',
                        'taux_nominal_field': 'TAUX_NOMINAL_I17',
                        'filters': Q()
                    }
                ]
            },
            'Découverts': {
                'model': Decouverts,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I05',
                'teg_field': 'TEG_decouvert',
                'montant_field': 'MONTANT_DECOUVERT_I08',
                'taux_nominal_field': 'TAUX_NOMINAL_I10',
                'filters': Q()
            },
            'Crédits à moyens terme': {
                'model': Credit_Amortissables,
                'date_field': 'DATE_MEP_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_annualise',
                'montant_field': 'MONTANT_PRET_I13',
                'taux_nominal_field': 'TAUX_NOMINAL_I17',
                'filters': Q(MATURITE='2-MT')
            },
            'Crédits à long terme': {
                'model': Credit_Amortissables,
                'date_field': 'DATE_MEP_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_annualise',
                'montant_field': 'MONTANT_PRET_I13',
                'taux_nominal_field': 'TAUX_NOMINAL_I17',
                'filters': Q(MATURITE='3-LT')
            },
            'Cautions': {
                'model': Cautions,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_caution',
                'montant_field': 'MONTANT_CAUTION_I10',
                'taux_nominal_field': 'TAUX_CAUTION_I11',
                'filters': Q()
            },
            'Effets commerciaux': {
                'model': Effets_commerces,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_effet',
                'montant_field': 'MONTANT_EFFET_I11',
                'taux_nominal_field': 'TAUX_NOMINAL_I10',
                'filters': Q()
            },
            'Affacturage': {
                'model': Affacturage,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_affacturage',
                'montant_field': 'MONTANT_CREANCE_I10',
                'taux_nominal_field': None,
                'filters': Q()
            }
        },
        'Autres personnes morales': {
            'Crédits de trésorerie, autre que découvert': {
                'models': [
                    {
                        'model': Credit_Amortissables,
                        'date_field': 'DATE_MEP_I03',
                        'categorie_field': 'CATEGORIE_BENEF_I07',
                        'teg_field': 'TEG_annualise',
                        'montant_field': 'MONTANT_PRET_I13',
                        'taux_nominal_field': 'TAUX_NOMINAL_I17',
                        'filters': Q(MATURITE='1-CT')
                    },
                    {
                        'model': Spot,
                        'date_field': 'DATE_MEP_I03',
                        'categorie_field': 'CATEGORIE_BENEF_I07',
                        'teg_field': 'TEG_spot',
                        'montant_field': 'MONTANT_PRET_I13',
                        'taux_nominal_field': 'TAUX_NOMINAL_I17',
                        'filters': Q()
                    }
                ]
            },
            'Découverts': {
                'model': Decouverts,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I05',
                'teg_field': 'TEG_decouvert',
                'montant_field': 'MONTANT_DECOUVERT_I08',
                'taux_nominal_field': 'TAUX_NOMINAL_I10',
                'filters': Q()
            },
            'Crédits à moyens terme': {
                'model': Credit_Amortissables,
                'date_field': 'DATE_MEP_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_annualise',
                'montant_field': 'MONTANT_PRET_I13',
                'taux_nominal_field': 'TAUX_NOMINAL_I17',
                'filters': Q(MATURITE='2-MT')
            },
            'Crédits à long terme': {
                'model': Credit_Amortissables,
                'date_field': 'DATE_MEP_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_annualise',
                'montant_field': 'MONTANT_PRET_I13',
                'taux_nominal_field': 'TAUX_NOMINAL_I17',
                'filters': Q(MATURITE='3-LT')
            },
            'Cautions': {
                'model': Cautions,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_caution',
                'montant_field': 'MONTANT_CAUTION_I10',
                'taux_nominal_field': 'TAUX_CAUTION_I11',
                'filters': Q()
            },
            'Effets commerciaux': {
                'model': Effets_commerces,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_effet',
                'montant_field': 'MONTANT_EFFET_I11',
                'taux_nominal_field': 'TAUX_NOMINAL_I10',
                'filters': Q()
            },
            'Affacturage': {
                'model': Affacturage,
                'date_field': 'DATE_MISE_PLACE_I03',
                'categorie_field': 'CATEGORIE_BENEF_I07',
                'teg_field': 'TEG_affacturage',
                'montant_field': 'MONTANT_CREANCE_I10',
                'taux_nominal_field': None,
                'filters': Q()
            }
        }
    }

    # AGRÉGATION DES DONNÉES
    data = {}
    total_records_found = 0
    
    logger.info(f"\n{'='*80}")
    logger.info(f"🔍 GÉNÉRATION DONNÉES : {trimestre} {annee} - {type_etablissement}")
    logger.info(f"{'='*80}\n")
    
    for category_name, category_codes in categories_beneficiaires.items():
        data[category_name] = {}
        
        if category_name not in structure_credits:
            # Répliquer la structure pour toutes les catégories
            structure_credits[category_name] = structure_credits['Petites et Moyennes Entreprises']
        
        credit_types = structure_credits[category_name]
        
        for credit_type, config in credit_types.items():
            logger.info(f"\n📋 {category_name} - {credit_type}")
            
            # CAS SPÉCIAL : Multi-modèles (crédits de trésorerie)
            if 'models' in config:
                combined_stats = {
                    'teg_total': 0,
                    'montant_total': 0,
                    'taux_nominal_total': 0,
                    'count': 0
                }
                
                for model_config in config['models']:
                    stats = calculer_stats_amelioree(
                        model_config['model'],
                        model_config['date_field'],
                        model_config['categorie_field'],
                        model_config['teg_field'],
                        model_config['montant_field'],
                        model_config['taux_nominal_field'],
                        category_codes,
                        model_config['filters']
                    )
                    
                    if stats:
                        combined_stats['teg_total'] += stats['teg_moyen'] * stats['count']
                        combined_stats['montant_total'] += stats['montant_total']
                        combined_stats['taux_nominal_total'] += stats['taux_nominal_moyen'] * stats['count']
                        combined_stats['count'] += stats['count']
                
                if combined_stats['count'] > 0:
                    teg_moyen = combined_stats['teg_total'] / combined_stats['count']
                    taux_nominal_moyen = combined_stats['taux_nominal_total'] / combined_stats['count']
                    
                    data[category_name][credit_type] = {
                        'teg_moyen': round(teg_moyen, 2),
                        'seuil_usure': round((4 * teg_moyen) / 3, 2),
                        'montant_total': combined_stats['montant_total'],
                        'taux_nominal_moyen': round(taux_nominal_moyen, 2)
                    }
                    total_records_found += combined_stats['count']
                else:
                    data[category_name][credit_type] = {
                        'teg_moyen': 0,
                        'seuil_usure': 0,
                        'montant_total': 0,
                        'taux_nominal_moyen': 0
                    }
            
            # CAS NORMAL : Un seul modèle
            else:
                stats = calculer_stats_amelioree(
                    config['model'],
                    config['date_field'],
                    config['categorie_field'],
                    config['teg_field'],
                    config['montant_field'],
                    config['taux_nominal_field'],
                    category_codes,
                    config['filters']
                )
                
                if stats and stats['teg_moyen'] > 0:
                    data[category_name][credit_type] = {
                        'teg_moyen': round(stats['teg_moyen'], 2),
                        'seuil_usure': round((4 * stats['teg_moyen']) / 3, 2),
                        'montant_total': stats['montant_total'],
                        'taux_nominal_moyen': round(stats['taux_nominal_moyen'], 2)
                    }
                    total_records_found += stats['count']
                else:
                    data[category_name][credit_type] = {
                        'teg_moyen': 0,
                        'seuil_usure': 0,
                        'montant_total': 0,
                        'taux_nominal_moyen': 0
                    }

    # Date et heure actuelles
    current_date = timezone.now().strftime('%d/%m/%Y')
    current_time = timezone.now().strftime('%H:%M')
    default_ministre = 'Christian YOKA'

    context = {
        'data': data,
        'trimestre': trimestre,
        'annee': annee,
        'type_etablissement': type_etablissement,
        'current_date': current_date,
        'current_time': current_time,
        'trimestres': trimestres,
        'annees': annees,
        'types_etablissement': types_etablissement,
        'default_ministre': default_ministre,
        'total_records': total_records_found
    }

    return context


@login_required
@user_passes_test(is_cnef_user)
def calculer_statistiques_model(config, category_codes, start_date, end_date, type_mapping, type_etablissement, categorie_emf):
    """
    Fonction utilitaire pour calculer les statistiques d'un modèle donné
    """
    model = config['model']
    date_field = config['date_field']
    categorie_field = config['categorie_field']
    teg_field = config['teg_field']
    montant_field = config['montant_field']
    taux_nominal_field = config['taux_nominal_field']
    filters = config['filters']

    # Filtrer par type d'établissement et période
    date_filter = {
        f'{date_field}__gte': start_date,
        f'{date_field}__lt': end_date
    }
    
    queryset = model.objects.select_related('etablissement').filter(
        **date_filter,
        etablissement__type_etablissement=type_mapping.get(type_etablissement, 'EMF')
    )
    
    if categorie_emf:
        queryset = queryset.filter(etablissement__categorie_emf=categorie_emf)
    
    # Filtrer par catégorie de bénéficiaire
    category_filter = Q()
    for code in category_codes:
        category_filter |= Q(**{f'{categorie_field}__icontains': code})
    
    queryset = queryset.filter(category_filter & filters)

    # Agrégations avec gestion des valeurs None
    stats = queryset.aggregate(
        teg_moyen=Avg(teg_field),
        montant_total=Sum(montant_field),
    )
    
    # Gestion des valeurs None
    teg_moyen = stats['teg_moyen'] or 0
    montant_total = stats['montant_total'] or 0
    
    # Ajouter le taux nominal moyen si le champ existe
    if taux_nominal_field:
        taux_nominal_stats = queryset.aggregate(avg=Avg(taux_nominal_field))
        taux_nominal_moyen = taux_nominal_stats['avg'] or 0
    else:
        taux_nominal_moyen = 0

    return {
        'teg_moyen': teg_moyen,
        'montant_total': montant_total,
        'taux_nominal_moyen': taux_nominal_moyen
    }

# ==========================================
# VUES POUR GESTION DES ÉTABLISSEMENTS
# ==========================================

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.core.exceptions import ValidationError
from .models import Etablissement, User, ActionUtilisateur
import json

# ==========================================
# 1. CRÉER UN NOUVEL ÉTABLISSEMENT
# ==========================================

@login_required
@user_passes_test(is_cnef_user)
@require_http_methods(["POST"])
def creer_etablissement(request):
    """
    Crée un nouvel établissement financier
    """
    try:
        data = json.loads(request.body)
        
        # Validation des données
        Nom_etablissement = data.get('Nom_etablissement', '').strip()
        code_etablissement = data.get('code_etablissement', '').strip()
        type_etablissement = data.get('type_etablissement', '')
        categorie_emf = data.get('categorie_emf', None)
        
        # Vérifications
        if not Nom_etablissement or not code_etablissement or not type_etablissement:
            return JsonResponse({
                'success': False,
                'message': 'Tous les champs obligatoires doivent être remplis'
            }, status=400)
        
        # Vérifier si le code existe déjà
        if Etablissement.objects.filter(code_etablissement=code_etablissement).exists():
            return JsonResponse({
                'success': False,
                'message': f'Le code "{code_etablissement}" existe déjà'
            }, status=400)
        
        # Vérifier si le nom existe déjà
        if Etablissement.objects.filter(Nom_etablissement=Nom_etablissement).exists():
            return JsonResponse({
                'success': False,
                'message': f'L\'établissement "{Nom_etablissement}" existe déjà'
            }, status=400)
        
        # Validation pour les EMF
        if type_etablissement == 'EMF' and not categorie_emf:
            return JsonResponse({
                'success': False,
                'message': 'La catégorie est obligatoire pour les EMF'
            }, status=400)
        
        # Si c'est une banque, catégorie doit être None
        if type_etablissement == 'BANQUE':
            categorie_emf = None
        
        # Créer l'établissement
        with transaction.atomic():
            etablissement = Etablissement.objects.create(
                Nom_etablissement=Nom_etablissement,
                code_etablissement=code_etablissement,
                type_etablissement=type_etablissement,
                categorie_emf=categorie_emf,
                is_active=True
            )
            
            # Journaliser l'action
            ActionUtilisateur.enregistrer_action(
                utilisateur=request.user,
                type_action='CREATION_ETABLISSEMENT',
                description=f"Création de l'établissement {Nom_etablissement} ({code_etablissement})",
                etablissement=etablissement,
                request=request,
                donnees_supplementaires={
                    'type': type_etablissement,
                    'categorie': categorie_emf
                }
            )
        
        return JsonResponse({
            'success': True,
            'message': f'Établissement "{Nom_etablissement}" créé avec succès',
            'etablissement': {
                'id': etablissement.id,
                'nom': etablissement.Nom_etablissement,
                'code': etablissement.code_etablissement,
                'type': etablissement.get_type_etablissement_display(),
                'categorie': etablissement.get_categorie_emf_display() if etablissement.categorie_emf else None,
                'date_creation': etablissement.date_creation.strftime('%d/%m/%Y %H:%M'),
                'is_active': etablissement.is_active
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Données JSON invalides'
        }, status=400)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la création : {str(e)}'
        }, status=500)


# ==========================================
# 2. LISTER TOUS LES ÉTABLISSEMENTS
# ==========================================

@login_required
@user_passes_test(is_acnef)
def lister_etablissements(request):
    """
    Retourne la liste de tous les établissements avec statistiques et filtres
    """
    try:
        # Récupérer tous les établissements (sans select_related car pas de relation cree_par)
        etablissements = Etablissement.objects.all()
        
        # ========================================
        # APPLIQUER LES FILTRES
        # ========================================
        type_filter = request.GET.get('type_etablissement', '')
        categorie_filter = request.GET.get('categorie_emf', '')
        statut_filter = request.GET.get('is_active', '')
        
        if type_filter:
            etablissements = etablissements.filter(type_etablissement=type_filter)
        
        if categorie_filter:
            etablissements = etablissements.filter(categorie_emf=categorie_filter)
        
        if statut_filter:
            is_active = statut_filter.lower() == 'true'
            etablissements = etablissements.filter(is_active=is_active)
        
        # ========================================
        # CALCULER LES STATISTIQUES GLOBALES
        # (sans filtres pour avoir le total complet)
        # ========================================
        total_etablissements = Etablissement.objects.count()
        total_banques = Etablissement.objects.filter(type_etablissement='BANQUE').count()
        total_emf = Etablissement.objects.filter(type_etablissement='EMF').count()
        etablissements_actifs = Etablissement.objects.filter(is_active=True).count()
        
        stats = {
            'total': total_etablissements,
            'banques': total_banques,
            'emf': total_emf,
            'actifs': etablissements_actifs
        }
        
        # ========================================
        # FORMATER LES DONNÉES
        # ========================================
        liste_etablissements = []
        for etab in etablissements.order_by('Nom_etablissement'):
            # Compter les utilisateurs par rôle
            nb_aef = User.objects.filter(etablissement=etab, role='AEF', is_active=True).count()
            nb_uef = User.objects.filter(etablissement=etab, role='UEF', is_active=True).count()
            nb_total = nb_aef + nb_uef
            
            liste_etablissements.append({
                'id': etab.id,
                'Nom_etablissement': etab.Nom_etablissement,  # Format attendu par JS
                'nom': etab.Nom_etablissement,
                'code_etablissement': etab.code_etablissement,  # Format attendu par JS
                'code': etab.code_etablissement,
                'type_etablissement': etab.type_etablissement,  # Format attendu par JS
                'type': etab.get_type_etablissement_display(),
                'categorie_emf': etab.categorie_emf,  # Format attendu par JS
                'categorie': etab.get_categorie_emf_display() if etab.categorie_emf else '-',
                'is_active': etab.is_active,
                'date_creation': etab.date_creation.isoformat() if etab.date_creation else None,
                'nombre_utilisateurs': nb_total,  # Format attendu par JS
                'nb_aef': nb_aef,
                'nb_uef': nb_uef,
                'nb_total_utilisateurs': nb_total
            })
        
        logger.info(f"Liste des établissements chargée par {request.user} - {len(liste_etablissements)} établissements")
        
        return JsonResponse({
            'success': True,
            'etablissements': liste_etablissements,
            'stats': stats,
            'total': len(liste_etablissements),
            'message': f'{len(liste_etablissements)} établissement(s) trouvé(s)'
        })
        
    except Exception as e:
        logger.error(f"Erreur lors du listage des établissements par {request.user}: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors du chargement : {str(e)}'
        }, status=500)       

# ==========================================
# 3. DÉTAILS D'UN ÉTABLISSEMENT (avec utilisateurs)
# ==========================================

@login_required
@user_passes_test(is_cnef_user)
@require_http_methods(["GET"])
def details_etablissement(request, etablissement_id):
    """
    Retourne les détails complets d'un établissement avec tous ses utilisateurs
    """
    try:
        etablissement = get_object_or_404(Etablissement, id=etablissement_id)
        
        # Récupérer tous les utilisateurs de cet établissement
        utilisateurs_list = []
        utilisateurs = etablissement.utilisateurs.all().order_by('role', 'nom', 'prenom')
        
        for user in utilisateurs:
            utilisateurs_list.append({
                'id': user.id,
                'nom': user.nom,
                'prenom': user.prenom,
                'email': user.email,
                'telephone': user.telephone or '-',
                'role': user.get_role_display(),
                'role_code': user.role,
                'is_active': user.is_active,
                'date_joined': user.date_joined.strftime('%d/%m/%Y %H:%M'),
                'derniere_connexion': user.derniere_connexion.strftime('%d/%m/%Y %H:%M') if user.derniere_connexion else 'Jamais connecté'
            })
        
        # Statistiques
        nb_aef = etablissement.utilisateurs.filter(role='AEF').count()
        nb_uef = etablissement.utilisateurs.filter(role='UEF').count()
        nb_actifs = etablissement.utilisateurs.filter(is_active=True).count()
        nb_inactifs = etablissement.utilisateurs.filter(is_active=False).count()
        
        return JsonResponse({
            'success': True,
            'etablissement': {
                'id': etablissement.id,
                'nom': etablissement.Nom_etablissement,
                'code': etablissement.code_etablissement,
                'type': etablissement.get_type_etablissement_display(),
                'type_code': etablissement.type_etablissement,
                'categorie': etablissement.get_categorie_emf_display() if etablissement.categorie_emf else '-',
                'categorie_code': etablissement.categorie_emf,
                'is_active': etablissement.is_active,
                'date_creation': etablissement.date_creation.strftime('%d/%m/%Y %H:%M'),
                'date_modification': etablissement.date_modification.strftime('%d/%m/%Y %H:%M'),
            },
            'statistiques': {
                'nb_aef': nb_aef,
                'nb_uef': nb_uef,
                'nb_total': nb_aef + nb_uef,
                'nb_actifs': nb_actifs,
                'nb_inactifs': nb_inactifs
            },
            'utilisateurs': utilisateurs_list
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors du chargement : {str(e)}'
        }, status=500)


# ==========================================
# 4. MODIFIER UN ÉTABLISSEMENT
# ==========================================

@login_required
@user_passes_test(is_acnef)
@require_http_methods(["PUT", "PATCH"])
def modifier_etablissement(request, etablissement_id):
    """
    Modifie les informations d'un établissement
    """
    try:
        etablissement = get_object_or_404(Etablissement, id=etablissement_id)
        data = json.loads(request.body)
        
        ancien_nom = etablissement.Nom_etablissement
        modifications = []
        
        # Mise à jour des champs
        if 'Nom_etablissement' in data:
            nouveau_nom = data['Nom_etablissement'].strip()
            if nouveau_nom != etablissement.Nom_etablissement:
                # Vérifier l'unicité
                if Etablissement.objects.filter(Nom_etablissement=nouveau_nom).exclude(id=etablissement_id).exists():
                    return JsonResponse({
                        'success': False,
                        'message': f'Le nom "{nouveau_nom}" existe déjà'
                    }, status=400)
                etablissement.Nom_etablissement = nouveau_nom
                modifications.append(f'Nom modifié: {ancien_nom} → {nouveau_nom}')
        
        if 'type_etablissement' in data:
            if data['type_etablissement'] != etablissement.type_etablissement:
                etablissement.type_etablissement = data['type_etablissement']
                modifications.append(f'Type modifié: {etablissement.get_type_etablissement_display()}')
        
        if 'categorie_emf' in data:
            if etablissement.type_etablissement == 'EMF':
                etablissement.categorie_emf = data['categorie_emf']
                modifications.append(f'Catégorie modifiée: {etablissement.get_categorie_emf_display()}')
            else:
                etablissement.categorie_emf = None
        
        if 'is_active' in data:
            if data['is_active'] != etablissement.is_active:
                etablissement.is_active = data['is_active']
                statut = 'activé' if etablissement.is_active else 'désactivé'
                modifications.append(f'Établissement {statut}')
        
        # Sauvegarder
        etablissement.save()
        
        # Journaliser
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='MODIFICATION_ETABLISSEMENT',
            description=f"Modification de l'établissement {etablissement.Nom_etablissement}: {', '.join(modifications)}",
            etablissement=etablissement,
            request=request
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Établissement modifié avec succès',
            'modifications': modifications
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la modification : {str(e)}'
        }, status=500)


# ==========================================
# 5. SUPPRIMER UN ÉTABLISSEMENT
# ==========================================

@login_required
@user_passes_test(is_acnef)
@require_http_methods(["DELETE"])
def supprimer_etablissement(request, etablissement_id):
    """
    Supprime un établissement et tous ses utilisateurs associés
    ATTENTION : Action irréversible !
    """
    try:
        etablissement = get_object_or_404(Etablissement, id=etablissement_id)
        
        # Compter les utilisateurs qui seront supprimés
        nb_utilisateurs = etablissement.utilisateurs.count()
        Nom_etablissement = etablissement.Nom_etablissement
        code_etablissement = etablissement.code_etablissement
        
        # Journaliser AVANT la suppression
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='SUPPRESSION_ETABLISSEMENT',
            description=f"Suppression de l'établissement {Nom_etablissement} ({code_etablissement}) avec {nb_utilisateurs} utilisateur(s)",
            etablissement=None,  # L'établissement va être supprimé
            request=request,
            donnees_supplementaires={
                'etablissement_id': etablissement_id,
                'nom': Nom_etablissement,
                'code': code_etablissement,
                'nb_utilisateurs_supprimes': nb_utilisateurs
            }
        )
        
        # Supprimer (CASCADE supprimera automatiquement les utilisateurs)
        with transaction.atomic():
            etablissement.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Établissement "{Nom_etablissement}" supprimé avec succès',
            'details': {
                'nom': Nom_etablissement,
                'code': code_etablissement,
                'nb_utilisateurs_supprimes': nb_utilisateurs
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la suppression : {str(e)}'
        }, status=500)


# ==========================================
# 6. ACTIVER/DÉSACTIVER UN ÉTABLISSEMENT
# ==========================================

@login_required
@user_passes_test(is_acnef)
@require_http_methods(["POST"])
def toggle_etablissement_status(request, etablissement_id):
    """
    Active ou désactive un établissement
    """
    try:
        etablissement = get_object_or_404(Etablissement, id=etablissement_id)
        
        # Inverser le statut
        etablissement.is_active = not etablissement.is_active
        etablissement.save()
        
        statut = 'activé' if etablissement.is_active else 'désactivé'
        
        # Journaliser
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='MODIFICATION_ETABLISSEMENT',
            description=f"Établissement {etablissement.Nom_etablissement} {statut}",
            etablissement=etablissement,
            request=request
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Établissement {statut} avec succès',
            'is_active': etablissement.is_active
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur : {str(e)}'
        }, status=500)
        
# ==========================================
# GESTION DES UTILISATEURS - VUES COMPLÈTES
# ==========================================


from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from cnef.models import User, Etablissement, TokenInscription, ActionUtilisateur

logger = logging.getLogger(__name__)


# ==========================================
# API GESTION DES UTILISATEURS
# ==========================================

@login_required
@user_passes_test(is_chef)
@require_http_methods(["GET", "POST"])
def api_utilisateurs(request):
    """
    GET: Liste tous les utilisateurs avec filtres et pagination
    POST: Créer un nouvel utilisateur
    """
    
    if request.method == "GET":
        try:
            # Paramètres de pagination
            page = int(request.GET.get('page', 1))
            per_page = int(request.GET.get('per_page', 20))
            
            # Paramètres de filtrage
            role_filter = request.GET.get('role', '')
            etablissement_filter = request.GET.get('etablissement', '')
            search_query = request.GET.get('search', '')
            actif_filter = request.GET.get('actif', '')
            
            # Construction de la requête
            utilisateurs = User.objects.all().select_related('etablissement', 'cree_par')
            
            # Application des filtres
            if role_filter:
                utilisateurs = utilisateurs.filter(role=role_filter)
            
            if etablissement_filter:
                utilisateurs = utilisateurs.filter(etablissement_id=etablissement_filter)
            
            if search_query:
                utilisateurs = utilisateurs.filter(
                    Q(nom__icontains=search_query) |
                    Q(prenom__icontains=search_query) |
                    Q(email__icontains=search_query)
                )
            
            if actif_filter:
                utilisateurs = utilisateurs.filter(is_active=(actif_filter.lower() == 'true'))
            
            # Tri
            sort_by = request.GET.get('sort', '-date_joined')
            utilisateurs = utilisateurs.order_by(sort_by)
            
            # Pagination
            paginator = Paginator(utilisateurs, per_page)
            page_obj = paginator.get_page(page)
            
            # Formatage des données
            utilisateurs_data = []
            for user in page_obj:
                utilisateurs_data.append({
                    'id': user.id,
                    'nom': user.nom,
                    'prenom': user.prenom,
                    'email': user.email,
                    'telephone': user.telephone or '',
                    'role': user.role,
                    'role_display': user.get_role_display(),
                    'etablissement': {
                        'id': user.etablissement.id,
                        'nom': user.etablissement.Nom_etablissement,
                        'code': user.etablissement.code_etablissement
                    } if user.etablissement else None,
                    'is_active': user.is_active,
                    'date_joined': user.date_joined.isoformat(),
                    'derniere_connexion': user.derniere_connexion.isoformat() if user.derniere_connexion else None,
                    'cree_par': {
                        'id': user.cree_par.id,
                        'nom': user.cree_par.get_full_name()
                    } if user.cree_par else None,
                })
            
            # Statistiques
            total = utilisateurs.count()
            stats = {
                'total': total,
                'actifs': utilisateurs.filter(is_active=True).count(),
                'inactifs': utilisateurs.filter(is_active=False).count(),
                'par_role': {}
            }
            
            for role, _ in User.ROLE_CHOICES:
                stats['par_role'][role] = utilisateurs.filter(role=role).count()
            
            return JsonResponse({
                'success': True,
                'utilisateurs': utilisateurs_data,
                'pagination': {
                    'page': page_obj.number,
                    'per_page': per_page,
                    'total_pages': paginator.num_pages,
                    'total': total,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                },
                'stats': stats
            })
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des utilisateurs: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f"Erreur: {str(e)}"
            }, status=500)
    
    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            
            # Validation des champs obligatoires
            required_fields = ['nom', 'prenom', 'email', 'role']
            for field in required_fields:
                if not data.get(field):
                    return JsonResponse({
                        'success': False,
                        'message': f"Le champ '{field}' est obligatoire"
                    }, status=400)
            
            # Vérifier si l'email existe déjà
            if User.objects.filter(email=data['email']).exists():
                return JsonResponse({
                    'success': False,
                    'message': f"Un utilisateur avec l'email '{data['email']}' existe déjà"
                }, status=400)
            
            # Validation du rôle et de l'établissement
            role = data['role']
            if role in ['AEF', 'UEF']:
                if not data.get('etablissement_id'):
                    return JsonResponse({
                        'success': False,
                        'message': "Un établissement est requis pour les rôles AEF et UEF"
                    }, status=400)
                
                try:
                    etablissement = Etablissement.objects.get(id=data['etablissement_id'])
                except Etablissement.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'message': "Établissement introuvable"
                    }, status=404)
            else:
                etablissement = None
            
            # Créer l'utilisateur avec transaction
            with transaction.atomic():
                # Générer un mot de passe temporaire
                temp_password = User.objects.make_random_password(length=12)
                
                # Créer l'utilisateur
                user = User.objects.create_user(
                    email=data['email'],
                    nom=data['nom'],
                    prenom=data['prenom'],
                    password=temp_password,
                    telephone=data.get('telephone', ''),
                    role=role,
                    etablissement=etablissement,
                    cree_par=request.user
                )
                
                # Générer un token d'inscription si c'est AEF ou UEF
                token_obj = None
                if role in ['AEF', 'UEF']:
                    token_obj = TokenInscription.objects.create(
                        etablissement=etablissement,
                        nom_user=f"{user.prenom} {user.nom}" if role == 'UEF' else None,
                        role=role,
                        cree_par=request.user
                    )
                    user.token_inscription = token_obj.token
                    user.save()
                
                # Enregistrer l'action dans les logs
                ActionUtilisateur.enregistrer_action(
                    utilisateur=request.user,
                    type_action='CREATION_UTILISATEUR',
                    description=f"Création de l'utilisateur {user.get_full_name()} ({user.get_role_display()})",
                    etablissement=etablissement,
                    request=request,
                    donnees_supplementaires={
                        'utilisateur_cree_id': user.id,
                        'email': user.email,
                        'role': user.role
                    }
                )
                
                logger.info(f"Utilisateur créé: {user.email} par {request.user.email}")
            
            # Préparer la réponse
            response_data = {
                'success': True,
                'message': f"Utilisateur '{user.get_full_name()}' créé avec succès",
                'utilisateur': {
                    'id': user.id,
                    'nom': user.nom,
                    'prenom': user.prenom,
                    'email': user.email,
                    'role': user.role,
                    'role_display': user.get_role_display(),
                    'etablissement': {
                        'id': etablissement.id,
                        'nom': etablissement.Nom_etablissement
                    } if etablissement else None,
                },
                'mot_de_passe_temporaire': temp_password if role in ['ACNEF', 'UCNEF'] else None,
            }
            
            # Ajouter le lien d'inscription si applicable
            if token_obj:
                response_data['lien_inscription'] = token_obj.generer_lien_inscription()
                response_data['token_expiration'] = token_obj.temps_restant_minutes()
            
            return JsonResponse(response_data, status=201)
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Données JSON invalides'
            }, status=400)
        except Exception as e:
            logger.error(f"Erreur lors de la création de l'utilisateur: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f"Erreur: {str(e)}"
            }, status=500)


@login_required
@user_passes_test(is_chef)
@require_http_methods(["GET", "PUT", "DELETE"])
def api_utilisateur_detail(request, utilisateur_id):
    """
    GET: Récupérer les détails d'un utilisateur
    PUT: Modifier un utilisateur
    DELETE: Supprimer un utilisateur
    """
    
    try:
        utilisateur = User.objects.select_related('etablissement', 'cree_par').get(id=utilisateur_id)
    except User.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Utilisateur introuvable'
        }, status=404)
    
    if request.method == "GET":
        try:
            # Récupérer les statistiques de l'utilisateur
            nb_actions = ActionUtilisateur.objects.filter(utilisateur=utilisateur).count()
            
            utilisateur_data = {
                'id': utilisateur.id,
                'nom': utilisateur.nom,
                'prenom': utilisateur.prenom,
                'email': utilisateur.email,
                'telephone': utilisateur.telephone or '',
                'role': utilisateur.role,
                'role_display': utilisateur.get_role_display(),
                'etablissement': {
                    'id': utilisateur.etablissement.id,
                    'nom': utilisateur.etablissement.Nom_etablissement,
                    'code': utilisateur.etablissement.code_etablissement
                } if utilisateur.etablissement else None,
                'is_active': utilisateur.is_active,
                'date_joined': utilisateur.date_joined.isoformat(),
                'derniere_connexion': utilisateur.derniere_connexion.isoformat() if utilisateur.derniere_connexion else None,
                'cree_par': {
                    'id': utilisateur.cree_par.id,
                    'nom': utilisateur.cree_par.get_full_name()
                } if utilisateur.cree_par else None,
                'nb_actions': nb_actions,
                'token_inscription': utilisateur.token_inscription,
                'lien_inscription_utilise': utilisateur.lien_inscription_utilise,
            }
            
            return JsonResponse({
                'success': True,
                'utilisateur': utilisateur_data
            })
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des détails: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f"Erreur: {str(e)}"
            }, status=500)
    
    elif request.method == "PUT":
        try:
            data = json.loads(request.body)
            
            # Mise à jour des champs autorisés
            with transaction.atomic():
                if 'nom' in data:
                    utilisateur.nom = data['nom']
                if 'prenom' in data:
                    utilisateur.prenom = data['prenom']
                if 'telephone' in data:
                    utilisateur.telephone = data['telephone']
                if 'is_active' in data:
                    utilisateur.is_active = data['is_active']
                
                utilisateur.save()
                
                # Enregistrer l'action
                ActionUtilisateur.enregistrer_action(
                    utilisateur=request.user,
                    type_action='MODIFICATION_UTILISATEUR',
                    description=f"Modification de l'utilisateur {utilisateur.get_full_name()}",
                    etablissement=utilisateur.etablissement,
                    request=request,
                    donnees_supplementaires={
                        'utilisateur_modifie_id': utilisateur.id,
                        'champs_modifies': list(data.keys())
                    }
                )
                
                logger.info(f"Utilisateur modifié: {utilisateur.email} par {request.user.email}")
            
            return JsonResponse({
                'success': True,
                'message': f"Utilisateur '{utilisateur.get_full_name()}' modifié avec succès"
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Données JSON invalides'
            }, status=400)
        except Exception as e:
            logger.error(f"Erreur lors de la modification: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f"Erreur: {str(e)}"
            }, status=500)
    
    elif request.method == "DELETE":
        try:
            # Vérifier que l'utilisateur ne se supprime pas lui-même
            if utilisateur.id == request.user.id:
                return JsonResponse({
                    'success': False,
                    'message': "Vous ne pouvez pas supprimer votre propre compte"
                }, status=400)
            
            with transaction.atomic():
                nom_complet = utilisateur.get_full_name()
                email = utilisateur.email
                
                # Enregistrer l'action avant la suppression
                ActionUtilisateur.enregistrer_action(
                    utilisateur=request.user,
                    type_action='CREATION_UTILISATEUR',  # Utiliser un type existant
                    description=f"Suppression de l'utilisateur {nom_complet} ({email})",
                    etablissement=utilisateur.etablissement,
                    request=request,
                    donnees_supplementaires={
                        'utilisateur_supprime_id': utilisateur.id,
                        'email': email,
                        'role': utilisateur.role
                    }
                )
                
                utilisateur.delete()
                
                logger.warning(f"Utilisateur supprimé: {email} par {request.user.email}")
            
            return JsonResponse({
                'success': True,
                'message': f"Utilisateur '{nom_complet}' supprimé avec succès"
            })
            
        except Exception as e:
            logger.error(f"Erreur lors de la suppression: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f"Erreur: {str(e)}"
            }, status=500)


# ==========================================
# GÉNÉRATION DE LIENS D'INSCRIPTION
# ==========================================

@login_required
@user_passes_test(is_chef)
@require_http_methods(["POST"])
def generer_lien_inscription(request):
    """Génère un lien d'inscription pour un établissement"""
    
    try:
        data = json.loads(request.body)
        
        # Validation
        etablissement_id = data.get('etablissement_id')
        role = data.get('role')
        nom_user = data.get('nom_user', '')
        duree_heures = int(data.get('duree_heures', 48))
        
        if not etablissement_id or not role:
            return JsonResponse({
                'success': False,
                'message': 'Établissement et rôle requis'
            }, status=400)
        
        if role not in ['AEF', 'UEF']:
            return JsonResponse({
                'success': False,
                'message': 'Le rôle doit être AEF ou UEF'
            }, status=400)
        
        if role == 'UEF' and not nom_user:
            return JsonResponse({
                'success': False,
                'message': 'Le nom de l\'utilisateur est requis pour le rôle UEF'
            }, status=400)
        
        # Récupérer l'établissement
        try:
            etablissement = Etablissement.objects.get(id=etablissement_id)
        except Etablissement.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Établissement introuvable'
            }, status=404)
        
        # Créer le token
        with transaction.atomic():
            token_obj = TokenInscription.objects.create(
                etablissement=etablissement,
                nom_user=nom_user if role == 'UEF' else None,
                role=role,
                cree_par=request.user,
                date_expiration=timezone.now() + timedelta(hours=duree_heures)
            )
            
            # Enregistrer l'action
            ActionUtilisateur.enregistrer_action(
                utilisateur=request.user,
                type_action='GENERATION_LIEN',
                description=f"Génération d'un lien d'inscription {role} pour {etablissement.Nom_etablissement}",
                etablissement=etablissement,
                request=request,
                donnees_supplementaires={
                    'token_id': token_obj.id,
                    'role': role,
                    'duree_heures': duree_heures
                }
            )
            
            envoyer_email_invitation(token_obj, request.user)
            
            logger.info(f"Lien d'inscription généré pour {etablissement.Nom_etablissement} par {request.user.email}")
        
        return JsonResponse({
            'success': True,
            'message': 'Lien d\'inscription généré avec succès',
            'token': {
                'id': token_obj.id,
                'lien': token_obj.generer_lien_inscription(),
                'expiration': token_obj.date_expiration.isoformat(),
                'temps_restant_minutes': token_obj.temps_restant_minutes()
            }
        }, status=201)
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Données JSON invalides'
        }, status=400)
    except Exception as e:
        logger.error(f"Erreur lors de la génération du lien: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f"Erreur: {str(e)}"
        }, status=500)


# ==========================================
# JOURNALISATION / LOGS
# ==========================================

@login_required
@user_passes_test(is_chef)
@require_http_methods(["GET"])
def api_journalisation(request):
    """Récupère les logs avec filtres et pagination"""
    
    try:
        # Paramètres de pagination
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 50))
        
        # Paramètres de filtrage
        type_action = request.GET.get('type_action', '')
        utilisateur_id = request.GET.get('utilisateur_id', '')
        etablissement_id = request.GET.get('etablissement_id', '')
        date_debut = request.GET.get('date_debut', '')
        date_fin = request.GET.get('date_fin', '')
        search_query = request.GET.get('search', '')
        
        # Construction de la requête
        actions = ActionUtilisateur.objects.all().select_related('utilisateur', 'etablissement')
        
        # Application des filtres
        if type_action:
            actions = actions.filter(type_action=type_action)
        
        if utilisateur_id:
            actions = actions.filter(utilisateur_id=utilisateur_id)
        
        if etablissement_id:
            actions = actions.filter(etablissement_id=etablissement_id)
        
        if date_debut:
            try:
                date_debut_obj = datetime.fromisoformat(date_debut.replace('Z', '+00:00'))
                actions = actions.filter(date_action__gte=date_debut_obj)
            except ValueError:
                pass
        
        if date_fin:
            try:
                date_fin_obj = datetime.fromisoformat(date_fin.replace('Z', '+00:00'))
                actions = actions.filter(date_action__lte=date_fin_obj)
            except ValueError:
                pass
        
        if search_query:
            actions = actions.filter(
                Q(description__icontains=search_query) |
                Q(utilisateur__nom__icontains=search_query) |
                Q(utilisateur__prenom__icontains=search_query)
            )
        
        # Tri
        actions = actions.order_by('-date_action')
        
        # Pagination
        paginator = Paginator(actions, per_page)
        page_obj = paginator.get_page(page)
        
        # Formatage des données
        actions_data = []
        for action in page_obj:
            actions_data.append({
                'id': action.id,
                'type_action': action.type_action,
                'type_action_display': action.get_type_action_display(),
                'description': action.description,
                'utilisateur': {
                    'id': action.utilisateur.id,
                    'nom': action.utilisateur.get_full_name(),
                    'email': action.utilisateur.email
                } if action.utilisateur else None,
                'etablissement': {
                    'id': action.etablissement.id,
                    'nom': action.etablissement.Nom_etablissement
                } if action.etablissement else None,
                'adresse_ip': action.adresse_ip,
                'date_action': action.date_action.isoformat(),
            })
        
        # Statistiques
        total = actions.count()
        stats = {
            'total': total,
            'par_type': {}
        }
        
        for type_action, _ in ActionUtilisateur.TYPE_ACTION_CHOICES:
            stats['par_type'][type_action] = actions.filter(type_action=type_action).count()
        
        return JsonResponse({
            'success': True,
            'actions': actions_data,
            'pagination': {
                'page': page_obj.number,
                'per_page': per_page,
                'total_pages': paginator.num_pages,
                'total': total,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
            },
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des logs: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f"Erreur: {str(e)}"
        }, status=500)


@login_required
@user_passes_test(is_chef)
@require_http_methods(["GET"])
def exporter_journalisation_csv(request):
    """Exporte les logs en format CSV"""
    
    try:
        # Récupérer les mêmes filtres que l'API
        type_action = request.GET.get('type_action', '')
        utilisateur_id = request.GET.get('utilisateur_id', '')
        etablissement_id = request.GET.get('etablissement_id', '')
        date_debut = request.GET.get('date_debut', '')
        date_fin = request.GET.get('date_fin', '')
        search_query = request.GET.get('search', '')
        
        # Construction de la requête
        actions = ActionUtilisateur.objects.all().select_related('utilisateur', 'etablissement')
        
        # Application des filtres (même logique que l'API)
        if type_action:
            actions = actions.filter(type_action=type_action)
        if utilisateur_id:
            actions = actions.filter(utilisateur_id=utilisateur_id)
        if etablissement_id:
            actions = actions.filter(etablissement_id=etablissement_id)
        if date_debut:
            try:
                date_debut_obj = datetime.fromisoformat(date_debut.replace('Z', '+00:00'))
                actions = actions.filter(date_action__gte=date_debut_obj)
            except ValueError:
                pass
        if date_fin:
            try:
                date_fin_obj = datetime.fromisoformat(date_fin.replace('Z', '+00:00'))
                actions = actions.filter(date_action__lte=date_fin_obj)
            except ValueError:
                pass
        if search_query:
            actions = actions.filter(
                Q(description__icontains=search_query) |
                Q(utilisateur__nom__icontains=search_query) |
                Q(utilisateur__prenom__icontains=search_query)
            )
        
        actions = actions.order_by('-date_action')
        
        # Créer la réponse CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="journalisation_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # Ajouter le BOM UTF-8 pour Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # En-têtes
        writer.writerow([
            'ID',
            'Date/Heure',
            'Type d\'action',
            'Utilisateur',
            'Email',
            'Établissement',
            'Description',
            'Adresse IP'
        ])
        
        # Données
        for action in actions:
            writer.writerow([
                action.id,
                action.date_action.strftime('%d/%m/%Y %H:%M:%S'),
                action.get_type_action_display(),
                action.utilisateur.get_full_name() if action.utilisateur else 'N/A',
                action.utilisateur.email if action.utilisateur else 'N/A',
                action.etablissement.Nom_etablissement if action.etablissement else 'N/A',
                action.description,
                action.adresse_ip or 'N/A'
            ])
        
        # Enregistrer l'export dans les logs
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='AUTRE',
            description=f"Export CSV de {actions.count()} actions de journalisation",
            request=request
        )
        
        logger.info(f"Export CSV de {actions.count()} actions par {request.user.email}")
        
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'export CSV: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f"Erreur lors de l'export: {str(e)}"
        }, status=500)


# ==========================================
# API SUPPRESSION DE JOURNAUX
# ==========================================

@login_required
@user_passes_test(is_acnef)
@require_http_methods(["POST"])
def api_compter_journalisation(request):
    """
    Compte le nombre d'entrées de journal à supprimer selon les critères
    """
    try:
        data = json.loads(request.body)
        mode = data.get('mode')
        date_debut = data.get('date_debut')
        date_fin = data.get('date_fin')
        
        if not mode or not date_debut:
            return JsonResponse({
                'success': False,
                'message': 'Paramètres manquants'
            }, status=400)
        
        # Parser les dates
        try:
            date_debut_obj = timezone.datetime.strptime(date_debut, '%Y-%m-%dT%H:%M')
            date_debut_obj = timezone.make_aware(date_debut_obj)
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Format de date de début invalide'
            }, status=400)
        
        # Construire la requête
        if mode == 'avant':
            # Supprimer toutes les entrées AVANT la date
            queryset = ActionUtilisateur.objects.filter(date_action__lt=date_debut_obj)
        
        elif mode == 'intervalle':
            if not date_fin:
                return JsonResponse({
                    'success': False,
                    'message': 'Date de fin requise pour le mode intervalle'
                }, status=400)
            
            try:
                date_fin_obj = timezone.datetime.strptime(date_fin, '%Y-%m-%dT%H:%M')
                date_fin_obj = timezone.make_aware(date_fin_obj)
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Format de date de fin invalide'
                }, status=400)
            
            # Vérifier que date_debut < date_fin
            if date_debut_obj >= date_fin_obj:
                return JsonResponse({
                    'success': False,
                    'message': 'La date de début doit être antérieure à la date de fin'
                }, status=400)
            
            # Supprimer les entrées ENTRE les deux dates
            queryset = ActionUtilisateur.objects.filter(
                date_action__gte=date_debut_obj,
                date_action__lte=date_fin_obj
            )
        else:
            return JsonResponse({
                'success': False,
                'message': 'Mode invalide'
            }, status=400)
        
        nombre = queryset.count()
        
        return JsonResponse({
            'success': True,
            'nombre': nombre
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'JSON invalide'
        }, status=400)
    
    except Exception as e:
        logger.error(f"Erreur lors du comptage des journaux: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur serveur: {str(e)}'
        }, status=500)


@login_required
@user_passes_test(is_acnef)
@require_http_methods(["POST"])
def api_supprimer_journalisation(request):
    """
    Supprime les entrées de journal selon les critères
    """
    try:
        data = json.loads(request.body)
        mode = data.get('mode')
        date_debut = data.get('date_debut')
        date_fin = data.get('date_fin')
        
        if not mode or not date_debut:
            return JsonResponse({
                'success': False,
                'message': 'Paramètres manquants'
            }, status=400)
        
        # Parser les dates
        try:
            date_debut_obj = timezone.datetime.strptime(date_debut, '%Y-%m-%dT%H:%M')
            date_debut_obj = timezone.make_aware(date_debut_obj)
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Format de date de début invalide'
            }, status=400)
        
        # Construire la requête
        if mode == 'avant':
            queryset = ActionUtilisateur.objects.filter(date_action__lt=date_debut_obj)
        
        elif mode == 'intervalle':
            if not date_fin:
                return JsonResponse({
                    'success': False,
                    'message': 'Date de fin requise pour le mode intervalle'
                }, status=400)
            
            try:
                date_fin_obj = timezone.datetime.strptime(date_fin, '%Y-%m-%dT%H:%M')
                date_fin_obj = timezone.make_aware(date_fin_obj)
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Format de date de fin invalide'
                }, status=400)
            
            if date_debut_obj >= date_fin_obj:
                return JsonResponse({
                    'success': False,
                    'message': 'La date de début doit être antérieure à la date de fin'
                }, status=400)
            
            queryset = ActionUtilisateur.objects.filter(
                date_action__gte=date_debut_obj,
                date_action__lte=date_fin_obj
            )
        else:
            return JsonResponse({
                'success': False,
                'message': 'Mode invalide'
            }, status=400)
        
        # Compter avant suppression
        nombre = queryset.count()
        
        # Enregistrer l'action avant de supprimer
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='SUPPRESSION',
            description=f"Suppression de {nombre} entrées de journal ({mode})",
            request=request,
            donnees_supplementaires={
                'mode': mode,
                'date_debut': date_debut,
                'date_fin': date_fin,
                'nombre_supprime': nombre
            }
        )
        
        # Supprimer
        queryset.delete()
        
        return JsonResponse({
            'success': True,
            'nombre_lignes': nombre,
            'message': f'{nombre} entrée(s) supprimée(s) avec succès'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'JSON invalide'
        }, status=400)
    
    except Exception as e:
        logger.error(f"Erreur lors de la suppression des journaux: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur serveur: {str(e)}'
        }, status=500)


# ==========================================
# VUE POUR GÉNÉRER LES INVITATIONS
# ==========================================

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import get_object_or_404
from .models import Etablissement, TokenInscription, ActionUtilisateur
import json


def is_acnef(user):
    """Vérifie si l'utilisateur est ACNEF"""
    return user.is_authenticated and user.role == 'ACNEF'


def is_aef(user):
    """Vérifie si l'utilisateur est AEF"""
    return user.is_authenticated and user.role == 'AEF'


# ==========================================
# 1. GÉNÉRER INVITATION (ACNEF)
# ==========================================

@login_required
@user_passes_test(is_acnef)
@require_http_methods(["POST"])
def generer_invitation_acnef(request):
    """
    Génère un lien d'invitation pour ACNEF, UCNEF ou AEF
    Utilisé par l'ACNEF uniquement
    """
    try:
        data = json.loads(request.body)
        
        email_destinataire = data.get('email', '').strip()
        role = data.get('role', '')
        etablissement_id = data.get('etablissement_id', None)
        
        # Validation de base
        if not email_destinataire or not role:
            return JsonResponse({
                'success': False,
                'message': 'Email et rôle sont obligatoires'
            }, status=400)
        
        # Vérifier que le rôle est autorisé
        if role not in ['ACNEF', 'UCNEF', 'AEF']:
            return JsonResponse({
                'success': False,
                'message': 'Rôle non autorisé. Choisissez ACNEF, UCNEF ou AEF'
            }, status=400)
        
        # Si AEF, l'établissement est obligatoire
        etablissement = None
        if role == 'AEF':
            if not etablissement_id:
                return JsonResponse({
                    'success': False,
                    'message': 'L\'établissement est obligatoire pour le rôle AEF'
                }, status=400)
            
            try:
                etablissement = Etablissement.objects.get(id=etablissement_id)
            except Etablissement.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Établissement introuvable'
                }, status=404)
            
            # ⭐ Vérification : Limiter à 2 AEF par établissement
            nb_aef_existants = User.objects.filter(
                etablissement=etablissement, 
                role='AEF', 
                is_active=True
            ).count()
            
            # Compter aussi les invitations AEF en attente
            nb_invitations_aef_attente = TokenInscription.objects.filter(
                etablissement=etablissement,
                role='AEF',
                utilise=False,
                date_expiration__gt=timezone.now()
            ).count()
            
            total_aef = nb_aef_existants + nb_invitations_aef_attente
            
            if total_aef >= 2:
                return JsonResponse({
                    'success': False,
                    'message': f'Limite atteinte : Cet établissement a déjà {nb_aef_existants} AEF actif(s) et {nb_invitations_aef_attente} invitation(s) en attente. Maximum autorisé : 2 AEF par établissement.'
                }, status=400)
        
        # Vérifier si l'email existe déjà (User est déjà importé en haut du fichier)
        if User.objects.filter(email=email_destinataire).exists():
            return JsonResponse({
                'success': False,
                'message': f'Un utilisateur avec l\'email {email_destinataire} existe déjà'
            }, status=400)
        # Créer le token d'inscription
        token = TokenInscription.objects.create(
            role=role,
            etablissement=etablissement,
            email_destinataire=email_destinataire,
            cree_par=request.user
        )
        
        # Générer le lien
        lien = token.generer_lien_inscription()
        envoyer_email_invitation(token, request.user)
        
        # Journaliser
        description = f"Génération d'invitation {role} pour {email_destinataire}"
        if etablissement:
            description += f" - Établissement: {etablissement.Nom_etablissement}"
        
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='GENERATION_LIEN',
            description=description,
            etablissement=etablissement,
            request=request,
            donnees_supplementaires={
                'email_destinataire': email_destinataire,
                'role': role,
                'token_id': token.id
            }
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Lien d\'invitation généré avec succès',
            'invitation': {
                'id': token.id,
                'email': email_destinataire,
                'role': token.get_role_display(),
                'role_code': role,
                'etablissement': etablissement.Nom_etablissement if etablissement else None,
                'lien': lien,
                'expiration': token.date_expiration.strftime('%d/%m/%Y à %H:%M'),
                'temps_restant_minutes': token.temps_restant_minutes()
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Données JSON invalides'
        }, status=400)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur : {str(e)}'
        }, status=500)


# ==========================================
# 2. GÉNÉRER INVITATION UEF (AEF)
# ==========================================

@login_required
@user_passes_test(is_aef)
@require_http_methods(["POST"])
def generer_invitation_uef(request):
    """
    Génère un lien d'invitation pour UEF
    Utilisé par l'AEF pour inviter des utilisateurs de SON établissement
    """
    try:
        data = json.loads(request.body)
        
        email_destinataire = data.get('email', '').strip()
        
        # Validation
        if not email_destinataire:
            return JsonResponse({
                'success': False,
                'message': 'L\'email est obligatoire'
            }, status=400)
        
        # Vérifier que l'AEF a bien un établissement
        if not request.user.etablissement:
            return JsonResponse({
                'success': False,
                'message': 'Votre compte n\'est pas associé à un établissement'
            }, status=400)
        
        # Vérifier si l'email existe déjà
        from .models import User
        if User.objects.filter(email=email_destinataire).exists():
            return JsonResponse({
                'success': False,
                'message': f'Un utilisateur avec l\'email {email_destinataire} existe déjà'
            }, status=400)
        
        # Créer le token d'inscription (automatiquement rattaché à l'établissement de l'AEF)
        token = TokenInscription.objects.create(
            role='UEF',
            etablissement=request.user.etablissement,
            email_destinataire=email_destinataire,
            cree_par=request.user
        )
        
        # Générer le lien
        lien = token.generer_lien_inscription()
        # Envoi automatique de l'email d'invitation
        envoyer_email_invitation(token, request.user)
        
        # Journaliser
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='GENERATION_LIEN',
            description=f"Génération d'invitation UEF pour {email_destinataire} - Établissement: {request.user.etablissement.Nom_etablissement}",
            etablissement=request.user.etablissement,
            request=request,
            donnees_supplementaires={
                'email_destinataire': email_destinataire,
                'role': 'UEF',
                'token_id': token.id
            }
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Lien d\'invitation généré avec succès',
            'invitation': {
                'id': token.id,
                'email': email_destinataire,
                'role': 'Utilisateur Établissement (UEF)',
                'role_code': 'UEF',
                'etablissement': request.user.etablissement.Nom_etablissement,
                'lien': lien,
                'expiration': token.date_expiration.strftime('%d/%m/%Y à %H:%M'),
                'temps_restant_minutes': token.temps_restant_minutes()
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Données JSON invalides'
        }, status=400)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur : {str(e)}'
        }, status=500)


# ==========================================
# 3. LISTER LES INVITATIONS EN ATTENTE
# ==========================================

@login_required
@require_http_methods(["GET"])
def lister_invitations_en_attente(request):
    """
    Liste toutes les invitations non utilisées et non expirées
    - ACNEF voit toutes les invitations
    - AEF voit uniquement les invitations pour son établissement
    """
    try:
        from django.utils import timezone
        
        # Filtrer selon le rôle
        if request.user.role == 'ACNEF':
            # ACNEF voit toutes les invitations
            invitations = TokenInscription.objects.filter(
                utilise=False,
                date_expiration__gt=timezone.now()
            ).select_related('etablissement', 'cree_par').order_by('-date_creation')
        
        elif request.user.role == 'AEF':
            # AEF voit uniquement les invitations pour son établissement
            if not request.user.etablissement:
                return JsonResponse({
                    'success': False,
                    'message': 'Votre compte n\'est pas associé à un établissement'
                }, status=400)
            
            invitations = TokenInscription.objects.filter(
                utilise=False,
                date_expiration__gt=timezone.now(),
                etablissement=request.user.etablissement
            ).select_related('etablissement', 'cree_par').order_by('-date_creation')
        
        else:
            # UCNEF et UEF ne peuvent pas voir les invitations
            return JsonResponse({
                'success': False,
                'message': 'Vous n\'avez pas les permissions pour voir les invitations'
            }, status=403)
        
        # Formater les données
        liste_invitations = []
        for inv in invitations:
            liste_invitations.append({
                'id': inv.id,
                'email': inv.email_destinataire,
                'role': inv.get_role_display(),
                'role_code': inv.role,
                'etablissement': inv.etablissement.Nom_etablissement if inv.etablissement else '-',
                'cree_par': inv.cree_par.get_full_name() if inv.cree_par else 'Système',
                'date_creation': inv.date_creation.strftime('%d/%m/%Y %H:%M'),
                'expiration': inv.date_expiration.strftime('%d/%m/%Y %H:%M'),
                'temps_restant_minutes': inv.temps_restant_minutes(),
                'lien': inv.generer_lien_inscription()
            })
        
        return JsonResponse({
            'success': True,
            'invitations': liste_invitations,
            'total': len(liste_invitations)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur : {str(e)}'
        }, status=500)


# ==========================================
# 4. RÉVOQUER UNE INVITATION
# ==========================================

@login_required
@require_http_methods(["DELETE"])
def revoquer_invitation(request, token_id):
    """
    Révoque (supprime) une invitation non utilisée
    """
    try:
        token = get_object_or_404(TokenInscription, id=token_id)
        
        # Vérifier les permissions
        if request.user.role == 'ACNEF':
            # ACNEF peut révoquer toutes les invitations
            pass
        elif request.user.role == 'AEF':
            # AEF peut uniquement révoquer les invitations de son établissement
            if token.etablissement != request.user.etablissement:
                return JsonResponse({
                    'success': False,
                    'message': 'Vous ne pouvez révoquer que les invitations de votre établissement'
                }, status=403)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Vous n\'avez pas les permissions pour révoquer des invitations'
            }, status=403)
        
        # Vérifier que l'invitation n'est pas déjà utilisée
        if token.utilise:
            return JsonResponse({
                'success': False,
                'message': 'Cette invitation a déjà été utilisée et ne peut pas être révoquée'
            }, status=400)
        
        # Sauvegarder les infos avant suppression
        email = token.email_destinataire
        role = token.get_role_display()
        
        # Journaliser
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='AUTRE',
            description=f"Révocation de l'invitation {role} pour {email}",
            etablissement=token.etablissement,
            request=request,
            donnees_supplementaires={
                'token_id': token_id,
                'email': email,
                'role': token.role
            }
        )
        
        # Supprimer le token
        token.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Invitation pour {email} révoquée avec succès'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur : {str(e)}'
        }, status=500)
        

# ==========================================
# VUE POUR L'INSCRIPTION AVEC TOKEN
# À ajouter dans views.py
# ==========================================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login
from django.utils import timezone
from django.db import transaction
from .models import TokenInscription, User, ActionUtilisateur


def inscription_avec_token(request, token_str):
    """
    Vue pour gérer l'inscription d'un utilisateur via un token
    Gère tous les rôles : ACNEF, UCNEF, AEF, UEF
    """
    
    # Récupérer le token
    try:
        token = TokenInscription.objects.get(token=token_str)
    except TokenInscription.DoesNotExist:
        return render(request, 'inscription/inscription_token.html', {
            'token_valide': False,
            'error': 'Ce lien d\'inscription est invalide.'
        })
    
    # Vérifier si le token est expiré
    if token.est_expire():
        return render(request, 'inscription/inscription_token.html', {
            'token_valide': False,
            'error': f'Ce lien a expiré le {token.date_expiration.strftime("%d/%m/%Y à %H:%M")}. Demandez un nouveau lien à votre administrateur.'
        })
    
    # Vérifier si le token a déjà été utilisé
    if token.utilise:
        return render(request, 'inscription/inscription_token.html', {
            'token_valide': False,
            'error': 'Ce lien a déjà été utilisé. Vous ne pouvez pas l\'utiliser à nouveau.'
        })
    
    # Si méthode POST : traiter l'inscription
    if request.method == 'POST':
        return traiter_inscription(request, token)
    
    # Si méthode GET : afficher le formulaire
    return render(request, 'inscription/inscription_token.html', {
        'token_valide': True,
        'token': token
    })


def traiter_inscription(request, token):
    """
    Traite le formulaire d'inscription et crée l'utilisateur
    CORRIGÉ : Gestion de la session avant connexion
    """
    
    # Récupérer les données du formulaire
    nom = request.POST.get('nom', '').strip()
    prenom = request.POST.get('prenom', '').strip()
    telephone = request.POST.get('telephone', '').strip()
    password = request.POST.get('password', '')
    password_confirm = request.POST.get('password_confirm', '')
    
    # Validation de base
    errors = []
    
    if not nom or not prenom:
        errors.append('Le nom et le prénom sont obligatoires')
    
    if not password or not password_confirm:
        errors.append('Le mot de passe est obligatoire')
    
    if password != password_confirm:
        errors.append('Les mots de passe ne correspondent pas')
    
    if len(password) < 8:
        errors.append('Le mot de passe doit contenir au moins 8 caractères')
    
    # Vérifier que l'email n'existe pas déjà
    if User.objects.filter(email=token.email_destinataire).exists():
        errors.append('Un compte avec cet email existe déjà')
    
    # S'il y a des erreurs, réafficher le formulaire
    if errors:
        for error in errors:
            messages.error(request, error)
        return render(request, 'inscription/inscription_token.html', {
            'token_valide': True,
            'token': token,
            'form': request.POST
        })
    
    # Créer l'utilisateur
    try:
        with transaction.atomic():
            # CORRECTION CRITIQUE : Vider la session AVANT de créer l'utilisateur
            # Cela évite les conflits de session lors de l'auto-login
            request.session.flush()
            
            # Créer le compte utilisateur
            user = User.objects.create_user(
                email=token.email_destinataire,
                nom=nom,
                prenom=prenom,
                password=password,
                role=token.role,
                etablissement=token.etablissement,
                telephone=telephone if telephone else None
            )
            
            # Marquer le token comme utilisé
            token.utilise = True
            token.date_utilisation = timezone.now()
            token.save()
            
            # CORRECTION : Créer une NOUVELLE session propre avant le login
            request.session.create()
            
            # Connecter automatiquement l'utilisateur
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            
            # Journaliser l'action (APRÈS le login pour avoir une session valide)
            ActionUtilisateur.enregistrer_action(
                utilisateur=user,
                type_action='INSCRIPTION',
                description=f"Inscription réussie via token - Rôle: {token.get_role_display()}",
                etablissement=token.etablissement,
                request=request,
                donnees_supplementaires={
                    'token_id': token.id,
                    'cree_par': token.cree_par.email if token.cree_par else 'Système'
                }
            )
            
            # Message de succès
            messages.success(request, f'Bienvenue {prenom} ! Votre compte a été créé avec succès.')
            
            # Redirection selon le rôle
            if user.role in ['ACNEF', 'UCNEF']:
                return redirect('interface_chef')
            elif user.role == 'AEF':
                return redirect('interface_aef')
            else:  # UEF
                return redirect('tableau_de_bord')
    
    except Exception as e:
        logger.error(f"Erreur lors de la création du compte: {str(e)}")
        messages.error(request, f'Erreur lors de la création du compte : {str(e)}')
        return render(request, 'inscription/inscription_token.html', {
            'token_valide': True,
            'token': token,
            'form': request.POST
        })
     
# ==========================================
# VERSION COMPATIBLE AVEC ?token=xxx
# Pour les liens du type: /inscription/acnef/?token=abc123
# ==========================================


def inscription_acnef_get(request):
    """
    Inscription ACNEF avec token en GET parameter
    URL: /inscription/acnef/?token=xxx
    """
    token_str = request.GET.get('token')
    
    if not token_str:
        return render(request, 'inscription/inscription_token.html', {
            'token_valide': False,
            'error': 'Token manquant dans l\'URL'
        })
    
    return inscription_avec_token(request, token_str)


def inscription_ucnef_get(request):
    """
    Inscription UCNEF avec token en GET parameter
    URL: /inscription/ucnef/?token=xxx
    """
    token_str = request.GET.get('token')
    
    if not token_str:
        return render(request, 'inscription/inscription_token.html', {
            'token_valide': False,
            'error': 'Token manquant dans l\'URL'
        })
    
    return inscription_avec_token(request, token_str)


def inscription_aef_get(request, code_etablissement):
    """
    Inscription AEF avec token en GET parameter
    URL: /inscription/aef/BCC001/?token=xxx
    """
    token_str = request.GET.get('token')
    
    if not token_str:
        return render(request, 'inscription/inscription_token.html', {
            'token_valide': False,
            'error': 'Token manquant dans l\'URL'
        })
    
    return inscription_avec_token(request, token_str)


def inscription_uef_get(request, code_etablissement, nom_user=None):
    """
    Inscription UEF avec token en GET parameter
    URL: /inscription/uef/BCC001/?token=xxx
    ou   /inscription/uef/BCC001/jean.dupont/?token=xxx
    """
    token_str = request.GET.get('token')
    
    if not token_str:
        return render(request, 'inscription/inscription_token.html', {
            'token_valide': False,
            'error': 'Token manquant dans l\'URL'
        })
    
    return inscription_avec_token(request, token_str)


@login_required
@user_passes_test(is_acnef)
def charger_etablissements_select(request):
    etabs = Etablissement.objects.filter(is_active=True).order_by('Nom_etablissement')
    data = [
        {
            'id': etab.id,
            'nom': etab.Nom_etablissement,
            'code': etab.code_etablissement
        }
        for etab in etabs
    ]
    return JsonResponse({'etablissements': data})

# ==========================================
# VUES POUR L'INTERFACE AEF
# ==========================================

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.db.models import Count, Q
from django.views.decorators.http import require_http_methods
from django.db import transaction
import json

@login_required
@user_passes_test(is_aef)
def interface_aef(request):
    """
    Page principale pour l'administrateur établissement (AEF)
    """
    return render(request, 'EF/interface_aef.html', {
        'title': f'Interface AEF - {request.user.etablissement.Nom_etablissement}'
    })

@login_required
@user_passes_test(is_uef)
def interface_uef(request):
    """
    Page principale pour l'administrateur établissement (UEF)
    """
    return render(request, 'accounts/tableau_de_bord.html', {
        'title': f'Interface UEF - {request.user.etablissement.Nom_etablissement}'
    })


@login_required
@user_passes_test(is_aef)
@require_http_methods(["GET"])
def aef_api_dashboard(request):
    """
    API pour récupérer les statistiques du dashboard AEF
    """
    try:
        etablissement = request.user.etablissement
        
        # Statistiques des soumissions
        soumissions = FichierImport.objects.filter(etablissement_cnef=etablissement)
        
        stats = {
            'total': soumissions.count(),
            'en_attente': soumissions.filter(statut='EN_COURS').count(),
            'validees': soumissions.filter(statut='REUSSI').count(),
            'rejetees': soumissions.filter(statut='REJETE').count(),
            'utilisateurs_actifs': User.objects.filter(
                etablissement=etablissement,
                role='UEF',
                is_active=True
            ).count()
        }
        
        # Dernières soumissions (5 plus récentes)
        dernieres_soumissions = []
        for fichier in soumissions.order_by('-date_import')[:5]:
            dernieres_soumissions.append({
                'id': fichier.id,
                'nom_fichier': fichier.nom_fichier,
                'date_import': fichier.date_import.isoformat(),
                'statut': fichier.statut,
                'statut_display': fichier.get_statut_display(),
                'statut_class': get_statut_class(fichier.statut),
                'total_lignes': fichier.total_lignes_importees,
            })
        
        return JsonResponse({
            'success': True,
            'stats': stats,
            'dernieres_soumissions': dernieres_soumissions
        })
        
    except Exception as e:
        logger.error(f"Erreur dans aef_api_dashboard: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@user_passes_test(is_aef)
@require_http_methods(["GET"])
def aef_api_historique(request):
    """
    API pour récupérer l'historique des soumissions
    """
    try:
        etablissement = request.user.etablissement
        statut_filter = request.GET.get('statut', '')
        
        # Filtrer les soumissions
        soumissions = FichierImport.objects.filter(
            etablissement_cnef=etablissement
        ).order_by('-date_import')
        
        if statut_filter:
            soumissions = soumissions.filter(statut=statut_filter)
        
        # Formater les données
        data = []
        for fichier in soumissions:
            data.append({
                'id': fichier.id,
                'nom_fichier': fichier.nom_fichier,
                'date_import': fichier.date_import.isoformat(),
                'statut': fichier.statut,
                'statut_display': fichier.get_statut_display(),
                'statut_class': get_statut_class(fichier.statut),
                'total_lignes': fichier.total_lignes_importees,
                'commentaire': fichier.erreurs if fichier.statut == 'REJETE' else None,
            })
        
        return JsonResponse({
            'success': True,
            'soumissions': data
        })
        
    except Exception as e:
        logger.error(f"Erreur dans aef_api_historique: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)



@require_http_methods(["POST"])
@login_required
@user_passes_test(is_aef)
def aef_upload_fichier(request):
    """
    Upload d'un fichier Excel par l'AEF
    """
    if not request.FILES.get('fichier'):
        return JsonResponse({
            'success': False,
            'message': 'Aucun fichier fourni'
        }, status=400)
    
    fichier = request.FILES['fichier']
    
    # Validation de l'extension
    if not fichier.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({
            'success': False,
            'message': 'Seuls les fichiers Excel (.xlsx, .xls) sont autorisés'
        }, status=400)
    
    try:
        # Créer l'enregistrement
        fichier_import = FichierImport.objects.create(
            etablissement_cnef=request.user.etablissement,
            uploader_par=request.user,
            fichier=fichier,
            nom_fichier=fichier.name,
            statut='EN_COURS'
        )
        
        # ✅ CORRECTION : Notification avec le bon objet
        nb_notifications = envoyer_email_notification_acnef(fichier_import)
        
        # Log du résultat
        if nb_notifications > 0:
            logger.info(f"✅ {nb_notifications} administrateur(s) CNEF notifié(s) pour {fichier.name}")
        else:
            logger.warning(f"⚠️ Aucun administrateur CNEF notifié pour {fichier.name}")
        
        # Journaliser
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='UPLOAD_FICHIER',
            description=f"Upload du fichier {fichier.name}",
            etablissement=request.user.etablissement,
            request=request
        )
        
        # Message adapté
        if nb_notifications > 0:
            message = f'Fichier "{fichier.name}" soumis avec succès ! {nb_notifications} administrateur(s) CNEF notifié(s).'
        else:
            message = f'Fichier "{fichier.name}" soumis avec succès !'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'fichier_id': fichier_import.id
        })
        
    except Exception as e:
        logger.error(f"Erreur upload AEF: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de l\'upload : {str(e)}'
        }, status=500)

@login_required
@user_passes_test(is_aef)
@require_http_methods(["GET"])
def aef_api_utilisateurs_uef(request):
    """
    API pour récupérer les utilisateurs UEF de l'établissement
    """
    try:
        etablissement = request.user.etablissement
        
        utilisateurs = User.objects.filter(
            etablissement=etablissement,
            role='UEF'
        ).order_by('nom', 'prenom')
        
        data = []
        for user in utilisateurs:
            data.append({
                'id': user.id,
                'nom': user.nom,
                'prenom': user.prenom,
                'email': user.email,
                'telephone': user.telephone,
                'is_active': user.is_active,
                'date_joined': user.date_joined.isoformat(),
                'derniere_connexion': user.derniere_connexion.isoformat() if user.derniere_connexion else None,
            })
        
        return JsonResponse({
            'success': True,
            'utilisateurs': data
        })
        
    except Exception as e:
        logger.error(f"Erreur dans aef_api_utilisateurs_uef: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@user_passes_test(is_aef)
@require_http_methods(["GET"])
def aef_api_invitations_attente(request):
    """
    API pour récupérer les invitations en attente de l'établissement
    """
    try:
        etablissement = request.user.etablissement
        
        invitations = TokenInscription.objects.filter(
            etablissement=etablissement,
            role='UEF',
            utilise=False,
            date_expiration__gt=timezone.now()
        ).order_by('-date_creation')
        
        data = []
        for inv in invitations:
            data.append({
                'id': inv.id,
                'email': inv.email_destinataire,
                'date_creation': inv.date_creation.isoformat(),
                'expiration': inv.date_expiration.isoformat(),
                'temps_restant_minutes': inv.temps_restant_minutes(),
                'lien': inv.generer_lien_inscription(),
            })
        
        return JsonResponse({
            'success': True,
            'invitations': data
        })
        
    except Exception as e:
        logger.error(f"Erreur dans aef_api_invitations_attente: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@user_passes_test(is_aef)
@require_http_methods(["POST"])
def aef_api_generer_invitation_uef(request):
    """
    Génère une invitation pour un utilisateur UEF
    RÉUTILISE LA VUE EXISTANTE generer_invitation_uef
    """
    return generer_invitation_uef(request)


@login_required
@user_passes_test(is_aef)
@require_http_methods(["DELETE"])
def aef_api_revoquer_invitation(request, invitation_id):
    """
    Révoque une invitation
    RÉUTILISE LA VUE EXISTANTE revoquer_invitation
    """
    return revoquer_invitation(request, invitation_id)


@login_required
@user_passes_test(is_aef)
@require_http_methods(["PUT"])
def aef_api_modifier_profil(request):
    """
    Modifie le profil de l'AEF
    """
    try:
        data = json.loads(request.body)
        
        user = request.user
        user.prenom = data.get('prenom', user.prenom)
        user.nom = data.get('nom', user.nom)
        user.telephone = data.get('telephone', user.telephone)
        user.save()
        
        
        
        # Journaliser
        ActionUtilisateur.enregistrer_action(
            utilisateur=user,
            type_action='MODIFICATION_UTILISATEUR',
            description=f"Modification du profil AEF",
            etablissement=user.etablissement,
            request=request
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Profil modifié avec succès'
        })
        
    except Exception as e:
        logger.error(f"Erreur dans aef_api_modifier_profil: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@user_passes_test(is_aef)
@require_http_methods(["GET"])
def aef_detail_soumission(request, fichier_id):
    """
    Page de détails d'une soumission (similaire à detail_soumission du CNEF mais en lecture seule)
    """
    fichier = get_object_or_404(
        FichierImport, 
        id=fichier_id,
        etablissement_cnef=request.user.etablissement
    )
    
    # Réutiliser la même logique que detail_soumission
    from .utils import previsualiser_fichier_excel
    
    preview_result = previsualiser_fichier_excel(fichier)
    
    donnees_importees = {
        'credits': fichier.nb_credits_importes if fichier.statut == 'REUSSI' else preview_result.get('credits', 0),
        'decouverts': fichier.nb_decouverts_importes if fichier.statut == 'REUSSI' else preview_result.get('decouverts', 0),
        'affacturages': fichier.nb_affacturages_importes if fichier.statut == 'REUSSI' else preview_result.get('affacturages', 0),
        'cautions': fichier.nb_cautions_importes if fichier.statut == 'REUSSI' else preview_result.get('cautions', 0),
        'effets': fichier.nb_effets_importes if fichier.statut == 'REUSSI' else preview_result.get('effets', 0),
        'spots': getattr(fichier, 'nb_spots_importes', 0) if fichier.statut == 'REUSSI' else preview_result.get('spot', 0),
    }
    
    context = {
        'fichier': fichier,
        'donnees_importees': donnees_importees,
        'preview_result': preview_result,
        'total_lignes': preview_result.get('total_lignes', 0),
        'is_aef': True,  # Pour adapter le template
        'title': f'Détail - {fichier.nom_fichier}'
    }
    
    return render(request, 'EF/detail_soumission_ef.html', context)


# ==========================================
# FONCTION UTILITAIRE
# ==========================================

def get_statut_class(statut):
    """Retourne la classe CSS selon le statut"""
    mapping = {
        'EN_COURS': 'warning',
        'REUSSI': 'success',
        'REJETE': 'danger',
        'ERREUR': 'danger',
    }
    return mapping.get(statut, 'secondary')

@login_required
@user_passes_test(is_aef)
@require_http_methods(["GET"])
def aef_api_journalisation(request):
    """
    API pour récupérer les logs de journalisation de l'établissement AEF
    """
    try:
        etablissement = request.user.etablissement
        
        # Paramètres de pagination
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 50))
        
        # Paramètres de filtrage
        type_action = request.GET.get('type_action', '')
        date_debut = request.GET.get('date_debut', '')
        date_fin = request.GET.get('date_fin', '')
        search_query = request.GET.get('search', '')
        
        # Construction de la requête - FILTRER PAR ÉTABLISSEMENT
        actions = ActionUtilisateur.objects.filter(
            etablissement=etablissement
        ).select_related('utilisateur', 'etablissement')
        
        # Application des filtres
        if type_action:
            actions = actions.filter(type_action=type_action)
        
        if date_debut:
            try:
                date_debut_obj = datetime.fromisoformat(date_debut.replace('Z', '+00:00'))
                actions = actions.filter(date_action__gte=date_debut_obj)
            except ValueError:
                pass
        
        if date_fin:
            try:
                date_fin_obj = datetime.fromisoformat(date_fin.replace('Z', '+00:00'))
                actions = actions.filter(date_action__lte=date_fin_obj)
            except ValueError:
                pass
        
        if search_query:
            actions = actions.filter(
                Q(description__icontains=search_query) |
                Q(utilisateur__nom__icontains=search_query) |
                Q(utilisateur__prenom__icontains=search_query)
            )
        
        # Tri
        actions = actions.order_by('-date_action')
        
        # Pagination
        paginator = Paginator(actions, per_page)
        page_obj = paginator.get_page(page)
        
        # Formatage des données
        actions_data = []
        for action in page_obj:
            actions_data.append({
                'id': action.id,
                'type_action': action.type_action,
                'type_action_display': action.get_type_action_display(),
                'description': action.description,
                'utilisateur': {
                    'id': action.utilisateur.id,
                    'nom': action.utilisateur.get_full_name(),
                    'email': action.utilisateur.email
                } if action.utilisateur else None,
                'adresse_ip': action.adresse_ip,
                'date_action': action.date_action.isoformat(),
            })
        
        # Statistiques
        total = actions.count()
        stats = {
            'total': total,
            'par_type': {}
        }
        
        for type_action, _ in ActionUtilisateur.TYPE_ACTION_CHOICES:
            stats['par_type'][type_action] = actions.filter(type_action=type_action).count()
        
        return JsonResponse({
            'success': True,
            'actions': actions_data,
            'pagination': {
                'page': page_obj.number,
                'per_page': per_page,
                'total_pages': paginator.num_pages,
                'total': total,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
            },
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f"Erreur dans aef_api_journalisation: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f"Erreur: {str(e)}"
        }, status=500)


@login_required
@user_passes_test(is_aef)
@require_http_methods(["GET"])
def aef_exporter_journalisation_csv(request):
    """
    Exporte les logs de journalisation de l'établissement en CSV
    """
    try:
        etablissement = request.user.etablissement
        
        # Récupérer les mêmes filtres que l'API
        type_action = request.GET.get('type_action', '')
        date_debut = request.GET.get('date_debut', '')
        date_fin = request.GET.get('date_fin', '')
        search_query = request.GET.get('search', '')
        
        # Construction de la requête
        actions = ActionUtilisateur.objects.filter(
            etablissement=etablissement
        ).select_related('utilisateur', 'etablissement')
        
        # Application des filtres
        if type_action:
            actions = actions.filter(type_action=type_action)
        if date_debut:
            try:
                date_debut_obj = datetime.fromisoformat(date_debut.replace('Z', '+00:00'))
                actions = actions.filter(date_action__gte=date_debut_obj)
            except ValueError:
                pass
        if date_fin:
            try:
                date_fin_obj = datetime.fromisoformat(date_fin.replace('Z', '+00:00'))
                actions = actions.filter(date_action__lte=date_fin_obj)
            except ValueError:
                pass
        if search_query:
            actions = actions.filter(
                Q(description__icontains=search_query) |
                Q(utilisateur__nom__icontains=search_query) |
                Q(utilisateur__prenom__icontains=search_query)
            )
        
        actions = actions.order_by('-date_action')
        
        # Créer la réponse CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="journalisation_{etablissement.code_etablissement}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # Ajouter le BOM UTF-8 pour Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # En-têtes
        writer.writerow([
            'ID',
            'Date/Heure',
            'Type d\'action',
            'Utilisateur',
            'Email',
            'Description',
            'Adresse IP'
        ])
        
        # Données
        for action in actions:
            writer.writerow([
                action.id,
                action.date_action.strftime('%d/%m/%Y %H:%M:%S'),
                action.get_type_action_display(),
                action.utilisateur.get_full_name() if action.utilisateur else 'N/A',
                action.utilisateur.email if action.utilisateur else 'N/A',
                action.description,
                action.adresse_ip or 'N/A'
            ])
        
        # Journaliser l'export
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='AUTRE',
            description=f"Export CSV de {actions.count()} actions de journalisation",
            etablissement=etablissement,
            request=request
        )
        
        logger.info(f"Export CSV journalisation par AEF: {request.user.email} - {actions.count()} actions")
        
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'export CSV AEF: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f"Erreur lors de l'export: {str(e)}"
        }, status=500)
        
        
@login_required
@user_passes_test(is_aef)
def telecharger_rapport_teg_aef(request, fichier_id):
    """
    Vue pour AEF : Télécharger un rapport Excel détaillé des TEG
    avec toutes les colonnes originales + TEG Original, TEG Calculé, Conformité
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    from django.http import HttpResponse
    import io
    
    # Vérifier que le fichier appartient bien à l'établissement de l'AEF
    fichier = get_object_or_404(
        FichierImport,
        id=fichier_id,
        etablissement_cnef=request.user.etablissement,
    )
    
    try:
        logger.info(f"🔍 Génération rapport TEG Excel pour {fichier.nom_fichier} (AEF)")
        
        # 1. EXTRAIRE ET CALCULER LES TEG
        donnees_extraites = extraire_et_calculer_teg(
            fichier.fichier.path,
            fichier.etablissement_cnef
        )
        
        # 2. CHARGER LE FICHIER EXCEL ORIGINAL
        workbook_original = openpyxl.load_workbook(fichier.fichier.path, data_only=True)
        
        # 3. CRÉER UN NOUVEAU WORKBOOK POUR LE RAPPORT
        workbook_rapport = openpyxl.Workbook()
        workbook_rapport.remove(workbook_rapport.active)  # Supprimer la feuille par défaut
        
        # Styles pour le formatage
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        conforme_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        non_conforme_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Mapping des types de produits
        sheet_mapping = {
            'credits': ['credits amortissables', 'credit amortissable', 'crédits amortissables'],
            'decouverts': ['découverts bancaires', 'decouvert', 'découverts'],
            'affacturages': ['affacturage', 'affacturages'],
            'cautions': ['cautions', 'caution'],
            'effets': ['effets de commerce', 'effet'],
            'spots': ['spot', 'spots', 'cours spot']
        }
        
        # Labels pour les feuilles
        type_labels = {
            'credits': 'Crédits Amortissables',
            'decouverts': 'Découverts',
            'affacturages': 'Affacturage',
            'cautions': 'Cautions',
            'effets': 'Effets de Commerce',
            'spots': 'Spots'
        }
        
        # 4. TRAITER CHAQUE TYPE DE PRODUIT
        for type_produit, variations in sheet_mapping.items():
            if type_produit not in donnees_extraites or not donnees_extraites[type_produit]:
                continue  # Passer si aucune donnée
            
            # Trouver la feuille originale correspondante
            sheet_original = None
            for sheet_name in workbook_original.sheetnames:
                if any(var in sheet_name.lower() for var in variations):
                    sheet_original = workbook_original[sheet_name]
                    break
            
            if not sheet_original:
                continue
            
            # Créer une nouvelle feuille dans le rapport
            ws = workbook_rapport.create_sheet(title=type_labels[type_produit])
            
            # 5. COPIER LES EN-TÊTES ORIGINAUX
            header_row = None
            for row_num, row in enumerate(sheet_original.iter_rows(min_row=1, max_row=10), start=1):
                if any(cell.value for cell in row):
                    header_row = row_num
                    break
            
            if not header_row:
                continue
            
            # Copier l'en-tête
            headers = [cell.value for cell in sheet_original[header_row]]
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Ajouter les 3 nouvelles colonnes d'en-tête
            nouvelle_col_start = len(headers) + 1
            ws.cell(row=1, column=nouvelle_col_start, value="TEG Original (%)").fill = header_fill
            ws.cell(row=1, column=nouvelle_col_start, value="TEG Original (%)").font = header_font
            ws.cell(row=1, column=nouvelle_col_start, value="TEG Original (%)").alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=nouvelle_col_start, value="TEG Original (%)").border = border
            
            ws.cell(row=1, column=nouvelle_col_start + 1, value="TEG Calculé (%)").fill = header_fill
            ws.cell(row=1, column=nouvelle_col_start + 1, value="TEG Calculé (%)").font = header_font
            ws.cell(row=1, column=nouvelle_col_start + 1, value="TEG Calculé (%)").alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=nouvelle_col_start + 1, value="TEG Calculé (%)").border = border
            
            ws.cell(row=1, column=nouvelle_col_start + 2, value="Conformité").fill = header_fill
            ws.cell(row=1, column=nouvelle_col_start + 2, value="Conformité").font = header_font
            ws.cell(row=1, column=nouvelle_col_start + 2, value="Conformité").alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=nouvelle_col_start + 2, value="Conformité").border = border
            
            # 6. CRÉER UN MAPPING DES DONNÉES CALCULÉES PAR LIGNE
            donnees_map = {item['ligne']: item for item in donnees_extraites[type_produit]}
            
            # 7. COPIER LES DONNÉES + AJOUTER LES CALCULS
            output_row = 2
            for row_num in range(header_row + 1, sheet_original.max_row + 1):
                row_data = [cell.value for cell in sheet_original[row_num]]
                
                # Vérifier si la ligne n'est pas vide
                if not any(row_data):
                    continue
                
                # Copier les données originales
                for col_num, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=output_row, column=col_num, value=value)
                    cell.border = border
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
                
                # Ajouter les données calculées si disponibles
                if row_num in donnees_map:
                    item = donnees_map[row_num]
                    
                    # TEG Original (converti en %)
                    teg_original = item.get('teg_original', 0)
                    if teg_original < 1:  # Si c'est en décimal
                        teg_original *= 100
                    teg_original = round(teg_original, 2)
                    
                    # TEG Calculé (déjà en %)
                    teg_calcule = round(item.get('teg_calcule', 0), 2)
                    
                    # Conformité
                    conforme = item.get('conforme', False)
                    conformite_text = "Conforme" if conforme else "Non conforme"
                    
                    # Écrire les valeurs
                    cell_teg_original = ws.cell(row=output_row, column=nouvelle_col_start, value=teg_original)
                    cell_teg_calcule = ws.cell(row=output_row, column=nouvelle_col_start + 1, value=teg_calcule)
                    cell_conformite = ws.cell(row=output_row, column=nouvelle_col_start + 2, value=conformite_text)
                    
                    # Appliquer le formatage
                    cell_teg_original.border = border
                    cell_teg_original.alignment = Alignment(horizontal='right')
                    cell_teg_original.number_format = '0.00'
                    
                    cell_teg_calcule.border = border
                    cell_teg_calcule.alignment = Alignment(horizontal='right')
                    cell_teg_calcule.number_format = '0.00'
                    
                    cell_conformite.border = border
                    cell_conformite.alignment = Alignment(horizontal='center')
                    
                    # Couleur selon conformité
                    if conforme:
                        cell_conformite.fill = conforme_fill
                        cell_conformite.font = Font(color="006100", bold=True)
                    else:
                        cell_conformite.fill = non_conforme_fill
                        cell_conformite.font = Font(color="9C0006", bold=True)
                
                output_row += 1
            
            # 8. AJUSTER LA LARGEUR DES COLONNES
            for col_num in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0
                for cell in ws[column_letter]:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Figer la première ligne
            ws.freeze_panes = 'A2'
        
        # 9. SAUVEGARDER ET RETOURNER LE FICHIER
        # Nom du fichier
        date_str = datetime.now().strftime("%d-%m-%Y")
        nom_etablissement = fichier.etablissement_cnef.Nom_etablissement.replace(' ', '_')
        nom_fichier = f"Rapport_TEG_{nom_etablissement}_{date_str}.xlsx"
        
        # Créer la réponse HTTP
        output = io.BytesIO()
        workbook_rapport.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
        
        # Journalisation
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user,
            type_action='EXPORT',
            description=f"Téléchargement rapport TEG Excel - Fichier: {nom_fichier}",
            etablissement=request.user.etablissement,
            request=request
        )
        
        logger.info(f" Rapport TEG Excel généré avec succès: {nom_fichier}")
        
        return response
        
    except Exception as e:
        logger.error(f"❌ Erreur génération rapport TEG Excel (fichier {fichier_id}): {str(e)}")
        messages.error(request, f"Erreur lors de la génération du rapport : {str(e)}")
        return redirect('aef_detail_soumission', fichier_id=fichier_id)
    
    

# ==========================================
# 1. BANNIR UN UTILISATEUR
# ==========================================
@login_required
@user_passes_test(is_chef)
@require_http_methods(["POST"])
def bannir_utilisateur(request, user_id):
    try:
        utilisateur = get_object_or_404(User, id=user_id)
        if utilisateur.id == request.user.id:
            return JsonResponse({'success': False, 'message': 'Vous ne pouvez pas vous bannir vous-même'}, status=400)
        if utilisateur.is_superuser:
            return JsonResponse({'success': False, 'message': 'Impossible de bannir un super administrateur'}, status=403)
        if not utilisateur.is_active:
            return JsonResponse({'success': False, 'message': 'Cet utilisateur est déjà banni'}, status=400)
        
        nom_complet = utilisateur.get_full_name()
        utilisateur.is_active = False
        utilisateur.save()
        
        ActionUtilisateur.enregistrer_action(
            utilisateur=request.user, type_action='MODIFICATION_UTILISATEUR',
            description=f"Bannissement de l'utilisateur {nom_complet}",
            etablissement=utilisateur.etablissement, request=request
        )
        return JsonResponse({'success': True, 'message': f'Utilisateur {nom_complet} banni'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur : {str(e)}'}, status=500)

@login_required
@user_passes_test(is_chef)
@require_http_methods(["POST"])
def supprimer_journaux(request):
    try:
        data = json.loads(request.body)
        mode = data.get('mode')
        date_debut = data.get('date_debut')
        date_fin = data.get('date_fin')
        
        print(f"DEBUG - Mode: {mode}, Date début: {date_debut}, Date fin: {date_fin}")  # LOG
        
        # Validation
        if not mode or not date_debut:
            return JsonResponse({
                'success': False,
                'message': 'Paramètres manquants'
            }, status=400)
        
        # Convertir les dates
        date_debut_obj = None
        date_fin_obj = None
        
        try:
            # Essayer le format datetime-local (avec T)
            date_debut_obj = timezone.datetime.strptime(date_debut, '%Y-%m-%dT%H:%M')
            if mode == 'intervalle' and date_fin:
                date_fin_obj = timezone.datetime.strptime(date_fin, '%Y-%m-%dT%H:%M')
        except ValueError:
            try:
                # Essayer le format date simple
                date_debut_obj = timezone.datetime.strptime(date_debut, '%Y-%m-%d')
                if mode == 'intervalle' and date_fin:
                    date_fin_obj = timezone.datetime.strptime(date_fin, '%Y-%m-%d')
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': f'Format de date invalide. Reçu: {date_debut}'
                }, status=400)
        
        print(f"DEBUG - Date début parsée: {date_debut_obj}, Date fin parsée: {date_fin_obj}")  # LOG
        
        # Construire la requête de suppression
        if mode == 'avant':
            queryset = ActionUtilisateur.objects.filter(
                date_action__lt=date_debut_obj
            )
        elif mode == 'intervalle':
            if not date_fin:
                return JsonResponse({
                    'success': False,
                    'message': 'Date de fin manquante pour le mode intervalle'
                }, status=400)
            
            queryset = ActionUtilisateur.objects.filter(
                date_action__gte=date_debut_obj,
                date_action__lte=date_fin_obj
            )
        else:
            return JsonResponse({
                'success': False,
                'message': 'Mode invalide'
            }, status=400)
        
        # Compter avant suppression
        nombre = queryset.count()
        print(f"DEBUG - Nombre d'entrées à supprimer: {nombre}")  # LOG
        
        # Si c'est une requête de calcul seulement
        if request.GET.get('calcul_only') == 'true':
            return JsonResponse({
                'success': True,
                'nombre': nombre,
                'message': f'{nombre} entrée(s) seront supprimées'
            })
        
        # Effectuer la suppression réelle
        queryset.delete()
        
        # Enregistrer l'action
        ActionUtilisateur.objects.create(
            utilisateur=request.user,
            action='SUPPRESSION_JOURNAL',
            description=f"Suppression de {nombre} entrée(s) du journal (mode: {mode})",
            adresse_ip=request.META.get('REMOTE_ADDR', ''),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
        )
        
        logger.info(f"{nombre} entrées du journal supprimées par {request.user} (mode: {mode})")
        
        return JsonResponse({
            'success': True,
            'nombre_lignes': nombre,
            'message': f'{nombre} entrée(s) supprimée(s) avec succès'
        })
        
    except Exception as e:
        logger.error(f"Erreur lors de la suppression du journal par {request.user}: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur: {str(e)}'
        }, status=500)

@login_required
@user_passes_test(is_chef)
@require_http_methods(["POST"])
def compter_journaux_a_supprimer(request):
    try:
        import json
        from datetime import datetime
        from django.utils import timezone
        
        data = json.loads(request.body)
        mode = data.get('mode')
        date_debut_str = data.get('date_debut')
        date_fin_str = data.get('date_fin')
        
        date_debut = timezone.make_aware(datetime.fromisoformat(date_debut_str))
        
        if mode == 'intervalle':
            date_fin = timezone.make_aware(datetime.fromisoformat(date_fin_str))
            queryset = ActionUtilisateur.objects.filter(date_action__gte=date_debut, date_action__lte=date_fin)
        else:
            queryset = ActionUtilisateur.objects.filter(date_action__lt=date_debut)
        
        return JsonResponse({'success': True, 'nombre': queryset.count()})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    
# ==========================================
# VUES POUR L'HISTORIQUE DES EMAILS
# ==========================================

@login_required
@user_passes_test(is_chef)
@require_http_methods(["GET"])
def historique_emails(request):
    """
    Affiche l'historique de tous les emails envoyés
    Accessible uniquement aux ACNEF/Chef
    """
    
    # Si c'est une requête AJAX (appelée depuis JavaScript)
    if request.GET.get('ajax') == '1':
        try:
            # Récupérer les filtres depuis l'URL
            type_filter = request.GET.get('type', '')
            objet_filter = request.GET.get('objet', '')
            
            # Query de base : tous les emails, triés par date décroissante
            emails = HistoriqueEmail.objects.all().select_related(
                'utilisateur_envoyeur',
                'etablissement',
                'token_lie',
                'fichier_lie'
            ).order_by('-date_envoi')
            
            # Appliquer le filtre par type si présent
            if type_filter:
                emails = emails.filter(type_email=type_filter)
            
            # Appliquer le filtre par objet si présent
            if objet_filter:
                emails = emails.filter(objet__icontains=objet_filter)
            
            # Limiter aux 100 derniers emails pour la performance
            emails = emails[:100]
            
            # Sérialiser les données en JSON
            data = []
            for email in emails:
                data.append({
                    'id': email.id,
                    'date_envoi': email.date_envoi.strftime('%d/%m/%Y %H:%M'),
                    'type_email': email.type_email,
                    'type_display': email.get_type_email_display(),
                    'destinataire_email': email.destinataire_email,
                    'destinataire_nom': email.destinataire_nom or '',
                    'objet': email.objet,
                    'statut': email.statut,
                    'etablissement': email.etablissement.Nom_etablissement if email.etablissement else '',
                })
            
            # ========== NOUVEAU: Calculer les statistiques ==========
            total_emails = HistoriqueEmail.objects.count()
            emails_envoyes = HistoriqueEmail.objects.filter(statut='ENVOYE').count()
            emails_echec = HistoriqueEmail.objects.filter(statut='ECHEC').count()
            emails_aujourdhui = HistoriqueEmail.objects.filter(
                date_envoi__date=timezone.now().date()
            ).count()
            # ========================================================
            
            # Retourner les données en JSON avec les stats
            return JsonResponse({
                'emails': data,
                'stats': {
                    'total': total_emails,
                    'envoyes': emails_envoyes,
                    'echec': emails_echec,
                    'aujourdhui': emails_aujourdhui
                }
            })
        
        except Exception as e:
            logger.error(f"Erreur lors du chargement de l'historique des emails: {str(e)}")
            return JsonResponse({
                'emails': [],
                'error': str(e)
            }, status=500)
    
    # Si ce n'est pas une requête AJAX, afficher la page HTML
    # (mais normalement on n'utilisera que l'AJAX depuis l'interface chef)
    context = {
        'types_email': HistoriqueEmail.TYPE_EMAIL_CHOICES,
    }
    return render(request, 'historique_emails.html', context)

@login_required
@user_passes_test(is_chef)
@require_http_methods(["POST"])
def renvoyer_email_api(request, email_id):
    """
    Renvoie un email depuis l'historique
    Accessible uniquement aux ACNEF/Chef
    """
    try:
        # Récupérer l'email depuis l'historique
        email_historique = HistoriqueEmail.objects.get(id=email_id)
        
        # Renvoyer l'email en utilisant la fonction d'email_utils
        succes = renvoyer_email(email_historique)
        
        if succes:
            # Journaliser l'action dans ActionUtilisateur
            ActionUtilisateur.enregistrer_action(
                utilisateur=request.user,
                type_action='AUTRE',
                description=f"Renvoi d'email à {email_historique.destinataire_email} (Type: {email_historique.get_type_email_display()})",
                request=request
            )
            
            logger.info(f"Email renvoyé avec succès à {email_historique.destinataire_email} par {request.user.email}")
            
            return JsonResponse({
                'success': True,
                'message': f'Email renvoyé avec succès à {email_historique.destinataire_email}'
            })
        else:
            logger.warning(f"Échec du renvoi d'email à {email_historique.destinataire_email}")
            return JsonResponse({
                'success': False,
                'message': 'Échec du renvoi de l\'email. Vérifiez les logs.'
            }, status=500)
            
    except HistoriqueEmail.DoesNotExist:
        logger.error(f"Tentative de renvoi d'un email inexistant (ID: {email_id})")
        return JsonResponse({
            'success': False,
            'message': 'Email introuvable dans l\'historique'
        }, status=404)
        
    except Exception as e:
        logger.error(f"Erreur lors du renvoi d'email (ID: {email_id}): {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur: {str(e)}'
        }, status=500)


# ==========================================
# FIN DES VUES POUR L'HISTORIQUE DES EMAILS
# ==========================================

# ==========================================
# RÉINITIALISATION DU MOT DE PASSE
# ==========================================

from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password
from .email_utils import generate_reset_code, send_reset_code_email, is_code_expired
import datetime
import logging

logger = logging.getLogger(__name__)
User = get_user_model()


def saisie_mail_view(request):
    """
    Étape 1 : Saisie de l'adresse email pour réinitialisation
    """
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        
        if not email:
            messages.error(request, "Veuillez saisir une adresse email.")
            return render(request, 'réintialisation/saisie_mail.html')
        
        # Vérifier si l'email existe dans la base de données
        try:
            user = User.objects.get(email=email)
            
            # Générer le code de réinitialisation
            reset_code = generate_reset_code()
            
            # Sauvegarder dans la session
            request.session['reset_code'] = reset_code
            request.session['reset_email'] = email
            request.session['reset_code_sent_at'] = timezone.now().isoformat()
            request.session['reset_attempts'] = 0  # Compteur de tentatives
            
            # Envoyer le code par email
            if send_reset_code_email(email, reset_code):
                messages.success(request, f"Un code de confirmation a été envoyé à {email}")
                logger.info(f"Code de réinitialisation envoyé à {email}")
                return redirect('code_validation')
            else:
                messages.error(request, "Erreur lors de l'envoi du code. Veuillez réessayer.")
                logger.error(f"Échec envoi code reset à {email}")
            
        except User.DoesNotExist:
            # Pour la sécurité, on ne dit pas si l'email existe ou non
            # Mais on peut afficher un message générique
            messages.error(request, "Aucun compte trouvé avec cette adresse email.")
            logger.warning(f"Tentative reset pour email inexistant: {email}")
    
    return render(request, 'réintialisation/saisie_mail.html')


def code_validation_view(request):
    """
    Étape 2 : Validation du code de 6 caractères
    Avec validation automatique dès que 6 caractères corrects sont saisis
    """
    # Vérifier que l'utilisateur vient de la page saisie_mail
    if 'reset_code' not in request.session:
        messages.error(request, "Veuillez d'abord saisir votre adresse email.")
        return redirect('saisie_mail')
    
    # Vérifier si le code a expiré
    try:
        sent_at = datetime.datetime.fromisoformat(request.session['reset_code_sent_at'])
        sent_at = timezone.make_aware(sent_at) if timezone.is_naive(sent_at) else sent_at
        
        if is_code_expired(sent_at):
            messages.error(request, "⏰ Le code a expiré (15 minutes). Veuillez recommencer.")
            # Nettoyer la session
            _clear_reset_session(request)
            return redirect('saisie_mail')
            
    except (ValueError, KeyError) as e:
        logger.error(f"Erreur validation date session: {e}")
        messages.error(request, "Session invalide. Veuillez recommencer.")
        _clear_reset_session(request)
        return redirect('saisie_mail')
    
    if request.method == 'POST':
        # Gérer le renvoi de code
        if 'resend' in request.POST:
            # Régénérer un nouveau code
            reset_code = generate_reset_code()
            request.session['reset_code'] = reset_code
            request.session['reset_code_sent_at'] = timezone.now().isoformat()
            request.session['reset_attempts'] = 0  # Réinitialiser le compteur
            
            if send_reset_code_email(request.session['reset_email'], reset_code):
                messages.success(request, "✅ Un nouveau code a été envoyé à votre adresse email.")
                logger.info(f"Code renvoyé à {request.session['reset_email']}")
            else:
                messages.error(request, "❌ Erreur lors de l'envoi du code. Veuillez réessayer.")
            
            return redirect('code_validation')
        
        # Validation du code
        entered_code = request.POST.get('code', '').strip().upper()
        stored_code = request.session.get('reset_code', '').upper()
        
        # Récupérer le nombre de tentatives
        attempts = request.session.get('reset_attempts', 0)
        
        # Vérifier la limite de tentatives (5 max)
        if attempts >= 5:
            messages.error(request, "❌ Trop de tentatives incorrectes. Veuillez recommencer.")
            logger.warning(f"Trop de tentatives pour {request.session['reset_email']}")
            _clear_reset_session(request)
            return redirect('saisie_mail')
        
        if entered_code == stored_code:
            # Code correct, réinitialiser les tentatives et rediriger
            request.session['reset_attempts'] = 0
            request.session['code_validated'] = True  # Marquer que le code est validé
            logger.info(f"Code validé pour {request.session['reset_email']}")
            return redirect('confirmation')
        else:
            # Code incorrect, incrémenter les tentatives
            attempts += 1
            request.session['reset_attempts'] = attempts
            remaining = 5 - attempts
            
            if remaining > 0:
                messages.error(request, f"❌ Code incorrect. Il vous reste {remaining} tentative(s).")
            else:
                messages.error(request, "❌ Code incorrect. Limite de tentatives atteinte.")
                _clear_reset_session(request)
                return redirect('saisie_mail')
    
    return render(request, 'réintialisation/code_validation.html')


def code_validation_ajax(request):
    """
    Vue AJAX pour valider le code en temps réel (validation automatique)
    Appelée par JavaScript quand 6 caractères sont saisis
    """
    from django.http import JsonResponse
    
    if request.method == 'POST':
        import json
        try:
            data = json.loads(request.body)
            entered_code = data.get('code', '').strip().upper()
            
            # Vérifier que la session est valide
            if 'reset_code' not in request.session:
                return JsonResponse({
                    'valid': False,
                    'error': 'Session expirée',
                    'redirect': '/saisie-mail/'
                })
            
            # Vérifier expiration
            try:
                sent_at = datetime.datetime.fromisoformat(request.session['reset_code_sent_at'])
                sent_at = timezone.make_aware(sent_at) if timezone.is_naive(sent_at) else sent_at
                
                if is_code_expired(sent_at):
                    _clear_reset_session(request)
                    return JsonResponse({
                        'valid': False,
                        'error': 'Code expiré',
                        'redirect': '/saisie-mail/'
                    })
            except (ValueError, KeyError):
                return JsonResponse({
                    'valid': False,
                    'error': 'Session invalide',
                    'redirect': '/saisie-mail/'
                })
            
            stored_code = request.session.get('reset_code', '').upper()
            attempts = request.session.get('reset_attempts', 0)
            
            # Vérifier limite tentatives
            if attempts >= 5:
                _clear_reset_session(request)
                return JsonResponse({
                    'valid': False,
                    'error': 'Trop de tentatives',
                    'redirect': '/saisie-mail/'
                })
            
            # Vérifier le code
            if entered_code == stored_code:
                request.session['reset_attempts'] = 0
                request.session['code_validated'] = True
                return JsonResponse({
                    'valid': True,
                    'redirect': '/confirmation/'
                })
            else:
                # Incrémenter tentatives
                attempts += 1
                request.session['reset_attempts'] = attempts
                remaining = 5 - attempts
                
                return JsonResponse({
                    'valid': False,
                    'remaining': remaining,
                    'error': f'Code incorrect. {remaining} tentative(s) restante(s)'
                })
                
        except Exception as e:
            logger.error(f"Erreur validation AJAX: {e}")
            return JsonResponse({
                'valid': False,
                'error': 'Erreur serveur'
            })
    
    return JsonResponse({'valid': False, 'error': 'Méthode non autorisée'})


def confirmation_view(request):
    """
    Étape 3 : Création du nouveau mot de passe
    """
    # Vérifier que l'utilisateur a validé le code
    if not request.session.get('code_validated'):
        messages.error(request, "Veuillez d'abord valider votre code de confirmation.")
        return redirect('saisie_mail')
    
    # Vérifier si le code a expiré
    try:
        sent_at = datetime.datetime.fromisoformat(request.session['reset_code_sent_at'])
        sent_at = timezone.make_aware(sent_at) if timezone.is_naive(sent_at) else sent_at
        
        if is_code_expired(sent_at):
            messages.error(request, "⏰ La session a expiré. Veuillez recommencer.")
            _clear_reset_session(request)
            return redirect('saisie_mail')
            
    except (ValueError, KeyError):
        messages.error(request, "Session invalide. Veuillez recommencer.")
        _clear_reset_session(request)
        return redirect('saisie_mail')
    
    if request.method == 'POST':
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        
        # Validations
        if not password or not password_confirm:
            messages.error(request, "⚠️ Veuillez remplir tous les champs.")
        elif password != password_confirm:
            messages.error(request, "❌ Les mots de passe ne correspondent pas.")
        elif len(password) < 8:
            messages.error(request, "❌ Le mot de passe doit contenir au moins 8 caractères.")
        else:
            # Mettre à jour le mot de passe
            try:
                user = User.objects.get(email=request.session['reset_email'])
                user.set_password(password)  # Utiliser set_password pour le hachage
                user.save()
                
                logger.info(f"Mot de passe réinitialisé pour {user.email}")
                
                # Nettoyer la session
                _clear_reset_session(request)
                
                messages.success(request, "✅ Votre mot de passe a été réinitialisé avec succès. Vous pouvez maintenant vous connecter.")
                return redirect('connexion')
                
            except User.DoesNotExist:
                logger.error(f"Utilisateur non trouvé lors de la réinitialisation: {request.session.get('reset_email')}")
                messages.error(request, "❌ Erreur lors de la réinitialisation du mot de passe. Utilisateur non trouvé.")
                _clear_reset_session(request)
                return redirect('saisie_mail')
            except Exception as e:
                logger.error(f"Erreur lors de la réinitialisation: {e}")
                messages.error(request, "❌ Une erreur est survenue. Veuillez réessayer.")
    
    return render(request, 'réintialisation/confirmation.html')


def _clear_reset_session(request):
    """Fonction utilitaire pour nettoyer la session de réinitialisation"""
    request.session.pop('reset_code', None)
    request.session.pop('reset_email', None)
    request.session.pop('reset_code_sent_at', None)
    request.session.pop('reset_attempts', None)
    request.session.pop('code_validated', None)