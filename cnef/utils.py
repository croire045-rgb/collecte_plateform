import openpyxl
from datetime import datetime
from decimal import Decimal, InvalidOperation
import logging
from django.db import transaction, IntegrityError, DatabaseError
from .models import (
    FichierImport, Credit_Amortissables, Decouverts, 
    Affacturage, Cautions, Effets_commerces, Spot
)
import numpy_financial as npf

# Configuration du logger
logger = logging.getLogger(__name__)

def convertir_date(valeur):
    """Convertit une valeur en date"""
    if valeur is None or valeur == '':
        return None
    
    if isinstance(valeur, datetime):
        return valeur.date()
    
    if isinstance(valeur, str):
        formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y', '%Y/%m/%d']
        for fmt in formats:
            try:
                return datetime.strptime(valeur.strip(), fmt).date()
            except ValueError:
                continue
    
    return None

def nettoyer_valeur(valeur):
    """Nettoie et retourne la valeur brute"""
    if valeur is None or valeur == '':
        return None
    
    # Si c'est une chaîne, la nettoyer
    if isinstance(valeur, str):
        valeur = valeur.strip()
        if valeur in ('', '-', '—', 'N/A', 'n/a', 'NA'):
            return None
    
    return valeur

def convertir_decimal_safe(valeur, defaut=0):
    """Convertit une valeur en float de manière sécurisée"""
    if valeur is None or valeur == '':
        return float(defaut)
    
    # Si c'est déjà un nombre
    if isinstance(valeur, (int, float)):
        try:
            return float(valeur)
        except (ValueError, TypeError):
            return float(defaut)
    
    # Si c'est une chaîne
    if isinstance(valeur, str):
        valeur = valeur.strip()
        if valeur in ('', '-', '—', 'N/A', 'n/a', 'NA'):
            return float(defaut)
        
        # Nettoyer la chaîne - approche plus robuste
        import re
        # Garder seulement les chiffres, points, virgules et signes négatifs
        valeur = re.sub(r'[^\d,.\-]', '', valeur)
        
        # Remplacer la virgule par un point pour le format float
        valeur = valeur.replace(',', '.')
        
        # Vérifier s'il y a plusieurs points décimaux
        if valeur.count('.') > 1:
            # Garder seulement le premier point comme séparateur décimal
            parts = valeur.split('.')
            valeur = parts[0] + '.' + ''.join(parts[1:])
        
        # Vérifier si la chaîne résultante est vide ou seulement des signes
        if not valeur or valeur in ('-', '.'):
            return float(defaut)
        
        try:
            return float(valeur)
        except (ValueError, TypeError) as e:
            print(f"Warning: Impossible de convertir '{valeur}' en float: {e}")
            return float(defaut)
    
    try:
        return float(valeur)
    except (ValueError, TypeError) as e:
        print(f"Warning: Impossible de convertir {valeur} en float: {e}")
        return float(defaut)

def convertir_entier_safe(valeur, defaut=0):
    """Convertit une valeur en entier de manière sécurisée"""
    if valeur is None or valeur == '':
        return defaut
    
    # Si c'est déjà un entier
    if isinstance(valeur, int):
        return valeur
    
    # Si c'est un float
    if isinstance(valeur, float):
        try:
            return int(valeur)
        except (ValueError, TypeError):
            return defaut
    
    # Si c'est une chaîne
    if isinstance(valeur, str):
        valeur = valeur.strip()
        if valeur in ('', '-', '—', 'N/A', 'n/a', 'NA'):
            return defaut
        
        # Nettoyer la chaîne
        import re
        valeur = re.sub(r'[^\d,.\-]', '', valeur)
        valeur = valeur.replace(',', '.')
        
        try:
            # Essayer de convertir via Decimal d'abord pour gérer les nombres à virgule
            decimal_val = convertir_decimal_safe(valeur, defaut)
            return int(decimal_val)
        except (ValueError, TypeError, InvalidOperation):
            return defaut
    
    try:
        return int(float(valeur))
    except (ValueError, TypeError, InvalidOperation):
        return defaut

def creer_objet_avec_gestion_erreurs(constructeur, **kwargs):
    """
    Crée un objet en gérant les erreurs de conversion Decimal
    """
    try:
        return constructeur(**kwargs)
    except (InvalidOperation, ValueError) as e:
        print(f"Erreur lors de la création de l'objet: {e}")
        # Essayer de convertir tous les Decimal problématiques en 0
        for key, value in kwargs.items():
            if isinstance(value, Decimal):
                try:
                    # Vérifier si le Decimal est valide
                    value + Decimal('0')
                except (InvalidOperation, ValueError):
                    kwargs[key] = Decimal('0')
        
        # Réessayer avec les valeurs corrigées
        try:
            return constructeur(**kwargs)
        except Exception as e:
            print(f"Erreur persistante lors de la création de l'objet: {e}")
            raise

def extraire_credits_amortissables(worksheet, etablissement_cnef, fichier_import):
    """Extrait les crédits amortissables d'une feuille Excel"""
    credits = []
    erreurs = []
    
    # Trouver la ligne d'en-tête
    header_row = None
    for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10), start=1):
        if any(cell.value for cell in row):
            header_row = row_num
            break
    
    if not header_row:
        return [], ["Aucune ligne d'en-tête trouvée"]
    
    # Extraction des données
    for row_num, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values = [cell.value for cell in row]
        
        # Ignorer les lignes vides
        if not any(values):
            continue
        
        try:
            # Vérifier le nombre minimum de colonnes
            if len(values) < 26:
                erreurs.append(f"Ligne {row_num}: Nombre de colonnes insuffisant ({len(values)}/26)")
                continue
            
            # Extraction et nettoyage des valeurs
            date_mep = convertir_date(values[2])
            if not date_mep:
                erreurs.append(f"Ligne {row_num}: Date de mise en place invalide")
                continue
            
            # Préparer les données avec gestion robuste des Decimal
            taux_nominal = convertir_decimal_safe(values[16])
            if taux_nominal and taux_nominal > 1:
                taux_nominal /= 100
            
            teg = convertir_decimal_safe(values[25]) if len(values) > 25 else 0
            if teg and teg > 1:
                teg /= 100
            
            montant_pret = convertir_decimal_safe(values[12])
            duree = convertir_entier_safe(values[13])
            montant_echeance = convertir_decimal_safe(values[22])
            frais_dossier = convertir_decimal_safe(values[17])
            montant_assurance = convertir_decimal_safe(values[19])
            frais_annexe = convertir_decimal_safe(values[20])
            freq_remb = str(nettoyer_valeur(values[15]) or '')
            
            # ✅ CALCUL DE LA MATURITÉ SELON LA FORMULE
            if duree is not None:
                if duree <= 24:
                    maturite = "1-CT"  # Court terme
                elif duree > 60:
                    maturite = "3-LT"  # Long terme
                else:
                    maturite = "2-MT"  # Moyen terme
            else:
                maturite = "Non définie"
            
            
            try:
                if duree and duree > 0 and montant_echeance and montant_pret:
                    montant_net = (montant_pret - 
                                 (frais_dossier or 0) - 
                                 (montant_assurance or 0) - 
                                 (frais_annexe or 0))
                    teg_mensuel = npf.rate(
                        nper=duree,
                        pmt=-montant_echeance,
                        pv=montant_net,
                        fv=0
                    ) * 100
                else:
                    teg_mensuel = 0.0
            except Exception:
                teg_mensuel = 0.0
            
            # Calcul du TEG_annualisé
            try:
                if teg_mensuel and freq_remb:
                    freq = freq_remb.strip().lower()
                    if freq == '1' or 'mensuel' in freq or 'mois' in freq:
                        multiplicateur = 12
                    elif freq == '2' or 'trimestriel' in freq:
                        multiplicateur = 4
                    elif freq == '3' or 'semestriel' in freq:
                        multiplicateur = 2
                    elif freq == '4' or 'annuel' in freq:
                        multiplicateur = 1
                    elif freq == '5' or 'heb' in freq or 'hebdo' in freq:
                        multiplicateur = 52
                    else:
                        multiplicateur = 12
                    teg_annualise = teg_mensuel * multiplicateur
                else:
                    teg_annualise = 0.0
            except Exception:
                teg_annualise = 0.0
            
            donnees_credit = {
                'etablissement': etablissement_cnef,
                'fichier_import': fichier_import,
                'ETABLISSEMENT_I01': str(nettoyer_valeur(values[0]) or ''),
                'CODE_ETAB_I02': str(nettoyer_valeur(values[1]) or ''),
                'DATE_MEP_I03': date_mep,
                'CHA_ORI_I04': str(nettoyer_valeur(values[3]) or ''),
                'NATURE_PRET_I05': str(nettoyer_valeur(values[4]) or ''),
                'BENEFICIAIRE_I06': str(nettoyer_valeur(values[5]) or ''),
                'CATEGORIE_BENEF_I07': str(nettoyer_valeur(values[6]) or ''),
                'LIEU_RESIDENCE_I08': str(nettoyer_valeur(values[7]) or ''),
                'SECT_ACT_I09': str(nettoyer_valeur(values[8]) or ''),
                'MONTANT_CHAF_I10': convertir_decimal_safe(values[9]),
                'EFFECTIF_I11': convertir_entier_safe(values[10]),
                'PROFESSION_I12': str(nettoyer_valeur(values[11]) or ''),
                'MONTANT_PRET_I13': montant_pret,
                'DUREE_I14': duree,
                'DUREE_DIFFERRE_I15': convertir_entier_safe(values[14]),
                'FREQ_REMB_I16': freq_remb,
                'TAUX_NOMINAL_I17': taux_nominal,
                'FRAIS_DOSSIER_I18': frais_dossier,
                'MODALITEPAIEMENT_ASS_I19': str(nettoyer_valeur(values[18]) or ''),
                'MONTANTASSURANCE_I20': montant_assurance,
                'FRAIS_ANNEXE_I21': frais_annexe,
                'MODEREMBOURSEMENT_I22': str(nettoyer_valeur(values[21]) or ''),
                'MONTANT_ECHEANCE_I23': montant_echeance,
                'MODE_DEBLOCAGE_I24': str(nettoyer_valeur(values[23]) or ''),
                'SITUATION_CREANCE_I25': str(nettoyer_valeur(values[24]) or ''),
                'TEG_I26': teg,
                'TEG_mensuel': teg_mensuel,
                'TEG_annualise': teg_annualise,
                'MATURITE': maturite  
            }
            
            credit = creer_objet_avec_gestion_erreurs(Credit_Amortissables, **donnees_credit)
            credits.append(credit)
            
        except Exception as e:
            erreurs.append(f"Ligne {row_num}: {str(e)}")
            logger.error(f"Erreur ligne {row_num} lors de l'extraction crédits amortissables: {str(e)}")
    
    logger.info(f"Extraction crédits amortissables terminée: {len(credits)} crédits extraits, {len(erreurs)} erreurs")
    return credits, erreurs

def extraire_decouverts(worksheet, etablissement_cnef, fichier_import):
    """Extrait les découverts d'une feuille Excel"""
    decouverts = []
    erreurs = []
    
    # Trouver la ligne d'en-tête
    header_row = None
    for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10), start=1):
        if any(cell.value for cell in row):
            header_row = row_num
            break
    
    if not header_row:
        return [], ["Aucune ligne d'en-tête trouvée"]
    
    # Extraction des données
    for row_num, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values = [cell.value for cell in row]
        
        if not any(values):
            continue
        
        try:
            if len(values) < 17:
                erreurs.append(f"Ligne {row_num}: Nombre de colonnes insuffisant ({len(values)}/17)")
                continue
            
            date_mise_place = convertir_date(values[2])
            if not date_mise_place:
                erreurs.append(f"Ligne {row_num}: Date de mise en place invalide")
                continue
            
            taux_nominal = convertir_decimal_safe(values[9])
            if taux_nominal and taux_nominal > 1:
                taux_nominal /= 100
            
            teg = convertir_decimal_safe(values[16]) if len(values) > 16 else 0
            if teg and teg > 1:
                teg /= 100
            
            montant_decouvert = convertir_decimal_safe(values[7])
            frais_dossiers_comm = convertir_decimal_safe(values[10])
            couts_assurance = convertir_decimal_safe(values[11])
            frais_annexes = convertir_decimal_safe(values[12])
            
            # Calcul du TEG_decouvert
            try:
                if montant_decouvert and montant_decouvert != 0:
                    interets = montant_decouvert * taux_nominal
                    frais_totaux = ((frais_dossiers_comm or 0) + 
                                  (couts_assurance or 0) + 
                                  (frais_annexes or 0))
                    teg_decouvert = round(((interets + frais_totaux) / 
                                        montant_decouvert) * 100, 2)
                else:
                    teg_decouvert = 0.0
            except Exception:
                teg_decouvert = 0.0
            
            donnees_decouvert = {
                'etablissement': etablissement_cnef,
                'fichier_import': fichier_import,
                'SIGLE_I01': str(nettoyer_valeur(values[0]) or ''),
                'CODE_BANQUE_I02': str(nettoyer_valeur(values[1]) or ''),
                'DATE_MISE_PLACE_I03': date_mise_place,
                'BENEFICAIRE_I04': str(nettoyer_valeur(values[3]) or ''),
                'CATEGORIE_BENEF_I05': str(nettoyer_valeur(values[4]) or ''),
                'LIEU_RESIDENCE_I06': str(nettoyer_valeur(values[5]) or ''),
                'SECT_ACT_I07': str(nettoyer_valeur(values[6]) or ''),
                'MONTANT_DECOUVERT_I08': montant_decouvert,
                'CUMUL_TIRAGES_DEC_I09': convertir_decimal_safe(values[8]),
                'TAUX_NOMINAL_I10': taux_nominal,
                'FRAIS_DOSSIERS_ET_COMM_I11': frais_dossiers_comm,
                'COUTS_ASSURANCE_I12': couts_assurance,
                'FRAIS_ANNEXES_I13': frais_annexes,
                'AGIOS_I14': convertir_decimal_safe(values[13]),
                'NOMBRE_DEBITEURS_I15': convertir_entier_safe(values[14], 1),
                'SITUATION_CREANCE_I16': str(nettoyer_valeur(values[15]) or ''),
                'TEG_I17': teg,
                'TEG_decouvert': teg_decouvert
            }
            
            decouvert = creer_objet_avec_gestion_erreurs(Decouverts, **donnees_decouvert)
            decouverts.append(decouvert)
            
        except Exception as e:
            erreurs.append(f"Ligne {row_num}: {str(e)}")
    
    return decouverts, erreurs

def extraire_affacturages(worksheet, etablissement_cnef, fichier_import):
    """Extrait les affacturages d'une feuille Excel"""
    affacturages = []
    erreurs = []
    
    # Trouver la ligne d'en-tête
    header_row = None
    for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10), start=1):
        if any(cell.value for cell in row):
            header_row = row_num
            break
    
    if not header_row:
        return [], ["Aucune ligne d'en-tête trouvée"]
    
    # Extraction des données
    for row_num, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values = [cell.value for cell in row]
        
        if not any(values):
            continue
        
        try:
            if len(values) < 14:
                erreurs.append(f"Ligne {row_num}: Nombre de colonnes insuffisant ({len(values)}/14)")
                continue
            
            date_mise_place = convertir_date(values[2])
            date_echeance = convertir_date(values[3])
            
            if not date_mise_place:
                erreurs.append(f"Ligne {row_num}: Date de mise en place invalide")
                continue
            
            # Conversion et normalisation du TEG
            teg = convertir_decimal_safe(values[13]) if len(values) > 13 else 0
            if teg is not None and teg > 1:
                teg /= 100
            
            # CORRECTION 1: Retirer la virgule pour avoir une valeur simple
            duree_affacturage = convertir_entier_safe(values[4])
            if not duree_affacturage or duree_affacturage <= 0:
                erreurs.append(f"Ligne {row_num}: Durée d'affacturage invalide: {duree_affacturage}")
                continue
            
            # Conversion des montants avec valeurs par défaut
            montant_creance = convertir_entier_safe(values[9]) or 0
            montant_com_affacturage = convertir_decimal_safe(values[10]) or 0
            montant_comm_financement = convertir_entier_safe(values[11]) or 0
            montant_frais_annexes = convertir_entier_safe(values[12]) or 0
            
            # CORRECTION 2: Calcul correct du TEG_affacturage
            try:
                if montant_creance and montant_creance > 0:
                    # Calcul des frais totaux en pourcentage de la créance
                    frais_totaux = (montant_com_affacturage + 
                                  montant_comm_financement + 
                                  montant_frais_annexes) / montant_creance
                    
                    # CORRECTION 3: Formule correcte avec 360 jours/an
                    # (frais_totaux * 360 / durée) * 100 pour obtenir le pourcentage
                    teg_affacturage = round((frais_totaux * 360 / duree_affacturage) * 100, 2)
                    
                    
                else:
                    teg_affacturage = 0.0
                    print(f"Ligne {row_num}: Montant créance = {montant_creance}, TEG=0")
                    
            except ZeroDivisionError:
                teg_affacturage = 0.0
                erreurs.append(f"Ligne {row_num}: Division par zéro dans le calcul du TEG")
            except Exception as calc_error:
                teg_affacturage = 0.0
                erreurs.append(f"Ligne {row_num}: Erreur calcul TEG: {str(calc_error)}")
            
            donnees_affacturage = {
                'etablissement': etablissement_cnef,
                'fichier_import': fichier_import,
                'SIGLE_I01': str(nettoyer_valeur(values[0]) or ''),
                'CODE_BANQUE_I02': str(nettoyer_valeur(values[1]) or ''),
                'DATE_MISE_PLACE_I03': date_mise_place,
                'DATE_ECHEANCE_I04': date_echeance,
                'DUREE_AFFACTURAGE_I05': duree_affacturage, 
                'BENEFICAIRE_I06': str(nettoyer_valeur(values[5]) or ''),
                'CATEGORIE_BENEF_I07': str(nettoyer_valeur(values[6]) or ''),
                'LIEU_RESIDENCE_I08': str(nettoyer_valeur(values[7]) or ''),
                'SECT_ACT_I09': str(nettoyer_valeur(values[8]) or ''),
                'MONTANT_CREANCE_I10': montant_creance,
                'MONTANT_COM_AFFACTURAGE_I11': montant_com_affacturage,
                'MONTANT_COMM_FINANCEMENT_I12': montant_comm_financement,
                'MONTANT_FRAIS_ANNEXES_I13': montant_frais_annexes,
                'TEG_I14': teg,
                'TEG_affacturage': teg_affacturage
            }
            
            affacturage = creer_objet_avec_gestion_erreurs(Affacturage, **donnees_affacturage)
            if affacturage:
                affacturages.append(affacturage)
            else:
                erreurs.append(f"Ligne {row_num}: Erreur lors de la création de l'objet Affacturage")
            
        except Exception as e:
            erreurs.append(f"Ligne {row_num}: {str(e)}")
    
    return affacturages, erreurs

def extraire_cautions(worksheet, etablissement_cnef, fichier_import):
    """Extrait les cautions d'une feuille Excel"""
    cautions = []
    erreurs = []
    
    # Trouver la ligne d'en-tête
    header_row = None
    for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10), start=1):
        if any(cell.value for cell in row):
            header_row = row_num
            break
    
    if not header_row:
        return [], ["Aucune ligne d'en-tête trouvée"]
    
    # Extraction des données
    for row_num, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values = [cell.value for cell in row]
        
        if not any(values):
            continue
        
        try:
            if len(values) < 14:
                erreurs.append(f"Ligne {row_num}: Nombre de colonnes insuffisant ({len(values)}/14)")
                continue
            
            date_mise_place = convertir_date(values[2])
            date_echeance = convertir_date(values[3])
            
            if not date_mise_place:
                erreurs.append(f"Ligne {row_num}: Date de mise en place invalide")
                continue
            
            taux_caution = convertir_decimal_safe(values[10])
            if taux_caution and taux_caution > 1:
                taux_caution /= 100
            
            teg = convertir_decimal_safe(values[13]) if len(values) > 13 else 0
            if teg and teg > 1:
                teg /= 100
            
            montant_caution = convertir_entier_safe(values[9])
            montant_frais_comm = convertir_entier_safe(values[11])
            montant_frais_annexes = convertir_entier_safe(values[12])
            duree_caution = convertir_entier_safe(values[4])
            
            # Calcul du TEG_caution
            try:
                if montant_caution and montant_caution != 0 and duree_caution and duree_caution != 0:
                    montant_net = (montant_caution - 
                                 (montant_frais_comm or 0) - 
                                 (montant_frais_annexes or 0))
                    if montant_net != 0:
                        cout_annualise = ((montant_caution * taux_caution * duree_caution) / 360)
                        teg_caution = round((cout_annualise / montant_net) * (360 / duree_caution) * 100, 2)
                    else:
                        teg_caution = 0.0
                else:
                    teg_caution = 0.0
            except Exception:
                teg_caution = 0.0
            
            donnees_caution = {
                'etablissement': etablissement_cnef,
                'fichier_import': fichier_import,
                'SIGLE_I01': str(nettoyer_valeur(values[0]) or ''),
                'CODE_BANQUE_I02': str(nettoyer_valeur(values[1]) or ''),
                'DATE_MISE_PLACE_I03': date_mise_place,
                'DATE_ECHEANCE_I04': date_echeance,
                'DUREE_CAUTION_I05': duree_caution,
                'BENEFICAIRE_I06': str(nettoyer_valeur(values[5]) or ''),
                'CATEGORIE_BENEF_I07': str(nettoyer_valeur(values[6]) or ''),
                'LIEU_RESIDENCE_I08': str(nettoyer_valeur(values[7]) or ''),
                'SECT_ACT_I09': str(nettoyer_valeur(values[8]) or ''),
                'MONTANT_CAUTION_I10': montant_caution,
                'TAUX_CAUTION_I11': taux_caution,
                'MONTANT_FRAIS_COMM_I12': montant_frais_comm,
                'MONTANT_FRAIS_ANNEXES_I13': montant_frais_annexes,
                'TEG_I14': teg,
                'TEG_caution': teg_caution
            }
            
            caution = creer_objet_avec_gestion_erreurs(Cautions, **donnees_caution)
            cautions.append(caution)
            
        except Exception as e:
            erreurs.append(f"Ligne {row_num}: {str(e)}")
    
    return cautions, erreurs

def extraire_effets_commerces(worksheet, etablissement_cnef, fichier_import):
    """Extrait les effets de commerce d'une feuille Excel"""
    effets = []
    erreurs = []
    
    # Trouver la ligne d'en-tête
    header_row = None
    for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10), start=1):
        if any(cell.value for cell in row):
            header_row = row_num
            break
    
    if not header_row:
        return [], ["Aucune ligne d'en-tête trouvée"]
    
    # Extraction des données
    for row_num, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values = [cell.value for cell in row]
        
        if not any(values):
            continue
        
        try:
            if len(values) < 15:
                erreurs.append(f"Ligne {row_num}: Nombre de colonnes insuffisant ({len(values)}/15)")
                continue
            
            date_mise_place = convertir_date(values[2])
            date_echeance = convertir_date(values[3])
            
            if not date_mise_place:
                erreurs.append(f"Ligne {row_num}: Date de mise en place invalide")
                continue
            
            taux_nominal = convertir_decimal_safe(values[9])
            if taux_nominal and taux_nominal > 1:
                taux_nominal /= 100
            
            teg = convertir_decimal_safe(values[14]) if len(values) > 14 else 0
            if teg and teg > 1:
                teg /= 100
            
            montant_effet = convertir_entier_safe(values[10])
            montant_commission = convertir_entier_safe(values[12])
            autres_frais = convertir_entier_safe(values[13])
            duree_effet = convertir_entier_safe(values[4])
            
            # Calcul du TEG_effet
            try:
                if montant_effet and montant_effet != 0 and duree_effet and duree_effet != 0:
                    montant_net = (montant_effet - 
                                 (montant_commission or 0) - 
                                 (autres_frais or 0))
                    if montant_net != 0:
                        cout_interets = ((taux_nominal * montant_effet * duree_effet) / 360)
                        teg_effet = round(((cout_interets / montant_net) * (360 / duree_effet)) * 100, 2)
                    else:
                        teg_effet = 0.0
                else:
                    teg_effet = 0.0
            except Exception:
                teg_effet = 0.0
            
            donnees_effet = {
                'etablissement': etablissement_cnef,
                'fichier_import': fichier_import,
                'SIGLE_I01': str(nettoyer_valeur(values[0]) or ''),
                'CODE_BANQUE_I02': str(nettoyer_valeur(values[1]) or ''),
                'DATE_MISE_PLACE_I03': date_mise_place,
                'DATE_ECHEANCE_I04': date_echeance,
                'DUREE_EFFET_I05': duree_effet,
                'BENEFICAIRE_I06': str(nettoyer_valeur(values[5]) or ''),
                'CATEGORIE_BENEF_I07': str(nettoyer_valeur(values[6]) or ''),
                'LIEU_RESIDENCE_I08': str(nettoyer_valeur(values[7]) or ''),
                'SECT_ACT_I09': str(nettoyer_valeur(values[8]) or ''),
                'TAUX_NOMINAL_I10': taux_nominal,
                'MONTANT_EFFET_I11': montant_effet,
                'MONTANT_FRAIS_DOSSIERS_I12': convertir_entier_safe(values[11]),
                'MONTANT_COMMISSION_I13': montant_commission,
                'AUTRES_FRA_I14': autres_frais,
                'TEG_I15': teg,
                'TEG_effet': teg_effet
            }
            
            effet = creer_objet_avec_gestion_erreurs(Effets_commerces, **donnees_effet)
            effets.append(effet)
            
        except Exception as e:
            erreurs.append(f"Ligne {row_num}: {str(e)}")
    
    return effets, erreurs

def extraire_spot(worksheet, etablissement_cnef, fichier_import):
    """Extrait les spots d'une feuille Excel"""
    spots = []
    erreurs = []
    
    # Trouver la ligne d'en-tête
    header_row = None
    for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10), start=1):
        if any(cell.value for cell in row):
            header_row = row_num
            break
    
    if not header_row:
        return [], ["Aucune ligne d'en-tête trouvée"]
    
    # Extraction des données
    for row_num, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values = [cell.value for cell in row]
        
        # Ignorer les lignes vides
        if not any(values):
            continue
        
        try:
            # Vérifier le nombre minimum de colonnes
            if len(values) < 26:
                erreurs.append(f"Ligne {row_num}: Nombre de colonnes insuffisant ({len(values)}/26)")
                continue
            
            # Extraction et nettoyage des valeurs
            date_mep = convertir_date(values[2])
            if not date_mep:
                erreurs.append(f"Ligne {row_num}: Date de mise en place invalide")
                continue
            
            taux_nominal = convertir_decimal_safe(values[16])
            if taux_nominal and taux_nominal > 1:
                taux_nominal /= 100
            
            teg = convertir_decimal_safe(values[25]) if len(values) > 25 else 0
            if teg and teg > 1:
                teg /= 100
            
            montant_pret = convertir_decimal_safe(values[12])
            duree = convertir_entier_safe(values[13])
            montant_echeance = convertir_decimal_safe(values[22])
            frais_dossier = convertir_decimal_safe(values[17])
            montant_assurance = convertir_decimal_safe(values[19])
            frais_annexe = convertir_decimal_safe(values[20])
            
            # Calcul du TEG_spot
            try:
                if montant_pret and montant_pret != 0 and duree and duree != 0 and montant_echeance is not None:
                    taux_periodique = ((montant_echeance / montant_pret - 1) * 12) / duree
                    charges = ((frais_dossier or 0) + 
                             (montant_assurance or 0) + 
                             (frais_annexe or 0))
                    ratio_charges = charges / montant_pret
                    teg_spot = round((taux_periodique + ratio_charges) * 100, 2)
                else:
                    teg_spot = 0.0
            except Exception:
                teg_spot = 0.0
            
            donnees_spot = {
                'etablissement': etablissement_cnef,
                'fichier_import': fichier_import,
                'ETABLISSEMENT_I01': str(nettoyer_valeur(values[0]) or ''),
                'CODE_ETAB_I02': str(nettoyer_valeur(values[1]) or ''),
                'DATE_MEP_I03': date_mep,
                'CHA_ORI_I04': str(nettoyer_valeur(values[3]) or ''),
                'NATURE_PRET_I05': str(nettoyer_valeur(values[4]) or ''),
                'BENEFICIAIRE_I06': str(nettoyer_valeur(values[5]) or ''),
                'CATEGORIE_BENEF_I07': str(nettoyer_valeur(values[6]) or ''),
                'LIEU_RESIDENCE_I08': str(nettoyer_valeur(values[7]) or ''),
                'SECT_ACT_I09': str(nettoyer_valeur(values[8]) or ''),
                'MONTANT_CHAF_I10': convertir_decimal_safe(values[9]),
                'EFFECTIF_I11': convertir_entier_safe(values[10]),
                'PROFESSION_I12': str(nettoyer_valeur(values[11]) or ''),
                'MONTANT_PRET_I13': montant_pret,
                'DUREE_I14': duree,
                'DUREE_DIFFERRE_I15': convertir_entier_safe(values[14]),
                'FREQ_REMB_I16': str(nettoyer_valeur(values[15]) or ''),
                'TAUX_NOMINAL_I17': taux_nominal,
                'FRAIS_DOSSIER_I18': frais_dossier,
                'MODALITEPAIEMENT_ASS_I19': str(nettoyer_valeur(values[18]) or ''),
                'MONTANTASSURANCE_I20': montant_assurance,
                'FRAIS_ANNEXE_I21': frais_annexe,
                'MODEREMBOURSEMENT_I22': str(nettoyer_valeur(values[21]) or ''),
                'MONTANT_ECHEANCE_I23': montant_echeance,
                'MODE_DEBLOCAGE_I24': str(nettoyer_valeur(values[23]) or ''),
                'SITUATION_CREANCE_I25': str(nettoyer_valeur(values[24]) or ''),
                'TEG_I26': teg,
                'TEG_spot': teg_spot
            }
            
            spot_obj = creer_objet_avec_gestion_erreurs(Spot, **donnees_spot)
            spots.append(spot_obj)
            
        except Exception as e:
            erreurs.append(f"Ligne {row_num}: {str(e)}")
    
    return spots, erreurs

def identifier_type_feuille(sheet_name, sheet_mapping):
    """
    Identifie le type de feuille basé sur son nom avec une logique de correspondance exacte
    prioritaire pour éviter les confusions
    """
    sheet_name_clean = sheet_name.lower().strip()
    
    # Score de correspondance pour chaque type
    scores = {
        'credits': 0,
        'decouverts': 0,
        'affacturages': 0,
        'cautions': 0,
        'effets': 0,
        'spot': 0
    }
    
    # Calculer les scores pour chaque type
    for sheet_type, noms_possibles in sheet_mapping.items():
        for nom in noms_possibles:
            nom_lower = nom.lower()
            
            # Correspondance exacte = score maximum
            if sheet_name_clean == nom_lower:
                scores[sheet_type] = 1000
                break
            
            # Correspondance avec le mot complet = score élevé
            if sheet_name_clean in nom_lower or nom_lower in sheet_name_clean:
                # Vérifier si c'est un mot complet (pas une sous-chaîne)
                import re
                pattern = r'\b' + re.escape(nom_lower) + r'\b'
                if re.search(pattern, sheet_name_clean):
                    scores[sheet_type] += 100
                else:
                    scores[sheet_type] += 50
    
    # Retourner le type avec le score le plus élevé
    max_score = max(scores.values())
    
    if max_score == 0:
        return None
    
    # Trouver le type avec le score maximum
    for sheet_type, score in scores.items():
        if score == max_score:
            return sheet_type
    
    return None

def previsualiser_fichier_excel(fichier_import):
    """
    Prévisualise un fichier Excel sans enregistrer les données dans la base
    """
    resultat = {
        'success': False,
        'message': '',
        'credits': 0,
        'decouverts': 0,
        'affacturages': 0,
        'cautions': 0,
        'effets': 0,
        'spot': 0,
        'total_lignes': 0,
        'erreurs': []
    }
    
    try:
        logger.debug(f"Début de la prévisualisation du fichier {fichier_import.nom_fichier}")
        workbook = openpyxl.load_workbook(fichier_import.fichier.path, data_only=True)
        etablissement = fichier_import.etablissement_cnef
        
        sheet_mapping = {
            'credits': [
                'credits amortissables', 'credit amortissable', 'crédits amortissables',
                'crédit amortissable', 'credits_amortissables', 'credit_amortissable',
                'ca', 'credits', 'credit',
            ],
            'decouverts': [
                'découverts bancaires', 'decouvert bancaire', 'découverts',
                'decouverts', 'découvert', 'decouvert', 'dec',
            ],
            'affacturages': [
                'affacturage commercial', 'affacturages', 'affacturage', 'aff',
            ],
            'cautions': [
                'cautions bancaires', 'caution bancaire', 'cautions', 'caution', 'cau',
            ],
            'effets': [
                'effets de commerce', 'effet de commerce', 'effets_commerce', 'effet_commerce', 'effets commerciaux', 'effet commercial',
                'effets', 'effet', 'ec',
            ],
            'spot': [
                'spot', 'spots', 'spot_fx', 'spots_fx', 'spot-fx', 'spots-fx', 'spot fx', 'spots fx', 'cours spot', 'cours_spot', 'cours-spot', 'taux spot', 'taux_spot', 'taux-spot','valeur spot', 'valeur_spot', 'valeur-spot',
                'prix spot', 'prix_spot', 'prix-spot','sp',  
            ]
        }
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_type = identifier_type_feuille(sheet_name, sheet_mapping)
            
            if not sheet_type:
                resultat['erreurs'].append(f"Type de feuille non reconnu: '{sheet_name}'")
                logger.warning(f"Type de feuille non reconnu: {sheet_name}")
                continue
            
            logger.debug(f"Prévisualisation de la feuille '{sheet_name}' → {sheet_type}")
            
            # Appeler les fonctions d'extraction sans le paramètre preview
            if sheet_type == 'credits':
                credits, erreurs = extraire_credits_amortissables(worksheet, etablissement, fichier_import)
                resultat['credits'] += len(credits)
                resultat['erreurs'].extend(erreurs)
            
            elif sheet_type == 'decouverts':
                decouverts, erreurs = extraire_decouverts(worksheet, etablissement, fichier_import)
                resultat['decouverts'] += len(decouverts)
                resultat['erreurs'].extend(erreurs)
            
            elif sheet_type == 'affacturages':
                affacturages, erreurs = extraire_affacturages(worksheet, etablissement, fichier_import)
                resultat['affacturages'] += len(affacturages)
                resultat['erreurs'].extend(erreurs)
            
            elif sheet_type == 'cautions':
                cautions, erreurs = extraire_cautions(worksheet, etablissement, fichier_import)
                resultat['cautions'] += len(cautions)
                resultat['erreurs'].extend(erreurs)
            
            elif sheet_type == 'effets':
                effets, erreurs = extraire_effets_commerces(worksheet, etablissement, fichier_import)
                resultat['effets'] += len(effets)
                resultat['erreurs'].extend(erreurs)
            
            elif sheet_type == 'spot':
                spots, erreurs = extraire_spot(worksheet, etablissement, fichier_import)
                resultat['spot'] += len(spots)
                resultat['erreurs'].extend(erreurs)
        
        resultat['total_lignes'] = (
            resultat['credits'] + 
            resultat['decouverts'] + 
            resultat['affacturages'] + 
            resultat['cautions'] + 
            resultat['effets']+
            resultat['spot']
        )
        
        if resultat['total_lignes'] > 0:
            resultat['success'] = True
            resultat['message'] = f"Prévisualisation réussie: {resultat['total_lignes']} lignes détectées"
        else:
            resultat['message'] = "Aucune donnée valide trouvée dans le fichier"
            resultat['erreurs'].append("Aucune donnée valide détectée")
        
        logger.info(f"Prévisualisation réussie du fichier {fichier_import.nom_fichier}: {resultat['total_lignes']} lignes détectées")
        
    except Exception as e:
        resultat['success'] = False
        resultat['message'] = f"Erreur lors de la prévisualisation du fichier: {str(e)}"
        resultat['erreurs'].append(str(e))
        logger.error(f"Erreur lors de la prévisualisation du fichier {fichier_import.nom_fichier}: {str(e)}")
    
    return resultat

def traiter_fichier_excel(fichier_import):
    """
    Traite un fichier Excel et importe les données dans la base
    """
    resultat = {
        'success': False,
        'message': '',
        'credits': 0,
        'decouverts': 0,
        'affacturages': 0,
        'cautions': 0,
        'effets': 0,
        'spot': 0,
        'total_lignes': 0,
        'erreurs': []
    }
    
    # Mettre à jour le statut initial
    with transaction.atomic():
        fichier_import.statut = 'EN_COURS'
        fichier_import.erreurs = ''
        fichier_import.save()
    
    try:
        # Ouvrir le fichier Excel
        logger.debug(f"Début du traitement du fichier {fichier_import.nom_fichier}")
        workbook = openpyxl.load_workbook(fichier_import.fichier.path, data_only=True)
        etablissement = fichier_import.etablissement_cnef
        
        # Mapping étendu des noms de feuilles avec ordre de priorité
        sheet_mapping = {
            'credits': [
                'credits amortissables', 'credit amortissable', 'crédits amortissables',
                'crédit amortissable', 'credits_amortissables', 'credit_amortissable',
                'ca', 'credits', 'credit',
            ],
            'decouverts': [
                'découverts bancaires', 'decouvert bancaire', 'découverts',
                'decouverts', 'découvert', 'decouvert', 'dec',
            ],
            'affacturages': [
                'affacturage commercial', 'affacturages', 'affacturage', 'aff',
            ],
            'cautions': [
                'cautions bancaires', 'caution bancaire', 'cautions', 'caution', 'cau',
            ],
            'effets': [
                'effets de commerce', 'effet de commerce', 'effets_commerce',
                'effet_commerce', 'effets commerciaux', 'effet commercial',
                'effets', 'effet', 'ec',
            ],
            'spot': [
                'spot', 'spots',
                'spot_fx', 'spots_fx',
                'spot-fx', 'spots-fx',
                'spot fx', 'spots fx',
                'cours spot', 'cours_spot', 'cours-spot',
                'taux spot', 'taux_spot', 'taux-spot',
                'valeur spot', 'valeur_spot', 'valeur-spot',
                'prix spot', 'prix_spot', 'prix-spot',
                'sp',  
            ]
        }
        
        # Traiter chaque feuille
        feuilles_traitees = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_type = identifier_type_feuille(sheet_name, sheet_mapping)
            
            if not sheet_type:
                resultat['erreurs'].append(f"Type de feuille non reconnu: '{sheet_name}'")
                logger.warning(f"Type de feuille non reconnu: {sheet_name}")
                continue
            
            logger.debug(f"Traitement de la feuille '{sheet_name}' → {sheet_type}")
            feuilles_traitees.append(f" '{sheet_name}' → {sheet_type}")
            
            # Utiliser une transaction atomique pour chaque type de données
            with transaction.atomic():
                try:
                    if sheet_type == 'credits':
                        credits, erreurs = extraire_credits_amortissables(worksheet, etablissement, fichier_import)
                        if credits:
                            Credit_Amortissables.objects.bulk_create(credits, batch_size=500)
                            resultat['credits'] += len(credits)
                        resultat['erreurs'].extend(erreurs)
                        
                    elif sheet_type == 'spot':
                        spots, erreurs = extraire_spot(worksheet, etablissement, fichier_import)
                        if spots:
                            Spot.objects.bulk_create(spots, batch_size=500)
                            resultat['spot'] += len(spots)
                        resultat['erreurs'].extend(erreurs)
                    
                    elif sheet_type == 'decouverts':
                        decouverts, erreurs = extraire_decouverts(worksheet, etablissement, fichier_import)
                        if decouverts:
                            Decouverts.objects.bulk_create(decouverts, batch_size=500)
                            resultat['decouverts'] += len(decouverts)
                        resultat['erreurs'].extend(erreurs)
                    
                    elif sheet_type == 'affacturages':
                        affacturages, erreurs = extraire_affacturages(worksheet, etablissement, fichier_import)
                        if affacturages:
                            Affacturage.objects.bulk_create(affacturages, batch_size=500)
                            resultat['affacturages'] += len(affacturages)
                        resultat['erreurs'].extend(erreurs)
                    
                    elif sheet_type == 'cautions':
                        cautions, erreurs = extraire_cautions(worksheet, etablissement, fichier_import)
                        if cautions:
                            Cautions.objects.bulk_create(cautions, batch_size=500)
                            resultat['cautions'] += len(cautions)
                        resultat['erreurs'].extend(erreurs)
                    
                    elif sheet_type == 'effets':
                        effets, erreurs = extraire_effets_commerces(worksheet, etablissement, fichier_import)
                        if effets:
                            Effets_commerces.objects.bulk_create(effets, batch_size=500)
                            resultat['effets'] += len(effets)
                        resultat['erreurs'].extend(erreurs)
                    
                except (IntegrityError, DatabaseError) as e:
                    resultat['erreurs'].append(f"Erreur dans la feuille {sheet_name}: {str(e)}")
                    logger.error(f"Erreur dans la feuille {sheet_name}: {str(e)}")
                    raise  # Relancer pour annuler la transaction atomique interne
        
        # Calculer le total
        resultat['total_lignes'] = (
            resultat['credits'] + 
            resultat['decouverts'] + 
            resultat['affacturages'] + 
            resultat['cautions'] + 
            resultat['effets']+
            resultat['spot']
        )
        
        # Mettre à jour le fichier import dans une nouvelle transaction
        with transaction.atomic():
            fichier_import.nb_credits_importes = resultat['credits']
            fichier_import.nb_decouverts_importes = resultat['decouverts']
            fichier_import.nb_affacturages_importes = resultat['affacturages']
            fichier_import.nb_cautions_importes = resultat['cautions']
            fichier_import.nb_effets_importes = resultat['effets']
            fichier_import.nb_spots_importes = resultat['spot']
            
            if resultat['total_lignes'] > 0:
                fichier_import.statut = 'REUSSI'
                resultat['success'] = True
                resultat['message'] = f"Import réussi: {resultat['total_lignes']} lignes importées"
                
                details_parts = []
                if feuilles_traitees:
                    details_parts.append("Feuilles: " + ", ".join(feuilles_traitees))
                
                stats = []
                if resultat['credits'] > 0:
                    stats.append(f"{resultat['credits']} crédits")
                if resultat['decouverts'] > 0:
                    stats.append(f"{resultat['decouverts']} découverts")
                if resultat['affacturages'] > 0:
                    stats.append(f"{resultat['affacturages']} affacturages")
                if resultat['cautions'] > 0:
                    stats.append(f"{resultat['cautions']} cautions")
                if resultat['effets'] > 0:
                    stats.append(f"{resultat['effets']} effets")
                if resultat['spot'] > 0:
                    stats.append(f"{resultat['spot']} spot")
                
                if stats:
                    details_parts.append("Données: " + " | ".join(stats))
                
                fichier_import.details = "\n".join(details_parts)
            else:
                fichier_import.statut = 'ERREUR'
                resultat['message'] = "Aucune donnée valide trouvée dans le fichier"
            
            if resultat['erreurs']:
                fichier_import.erreurs = "\n".join(resultat['erreurs'][:20])
                if len(resultat['erreurs']) > 20:
                    fichier_import.erreurs += f"\n... et {len(resultat['erreurs']) - 20} autres erreurs"
            
            fichier_import.save()
        
        logger.info(f"Traitement réussi du fichier {fichier_import.nom_fichier}: {resultat['total_lignes']} lignes importées")
        
    except Exception as e:
        resultat['success'] = False
        resultat['message'] = f"Erreur lors du traitement du fichier: {str(e)}"
        resultat['erreurs'].append(str(e))
        logger.error(f"Erreur lors du traitement du fichier {fichier_import.nom_fichier}: {str(e)}")
        
        # Mettre à jour le fichier import dans une nouvelle transaction
        with transaction.atomic():
            fichier_import.statut = 'ERREUR'
            fichier_import.erreurs = "\n".join(resultat['erreurs'][:20])
            if len(resultat['erreurs']) > 20:
                fichier_import.erreurs += f"\n... et {len(resultat['erreurs']) - 20} autres erreurs"
            fichier_import.save()
    
    return resultat


from django.core.cache import cache
import numpy as np


def precalculer_teg_fichier(fichier_import):
    """
    Pré-calcule tous les TEG pour un fichier importé sans l'enregistrer en base
    """
    try:
        logger.debug(f"Début du pré-calcul TEG pour {fichier_import.nom_fichier}")
        workbook = openpyxl.load_workbook(fichier_import.fichier.path, data_only=True)
        etablissement = fichier_import.etablissement_cnef
        
        sheet_mapping = {
            'credits': ['credits amortissables', 'credit amortissable', 'crédits amortissables'],
            'decouverts': ['découverts bancaires', 'decouvert bancaire', 'découverts'],
            'affacturages': ['affacturage commercial', 'affacturages', 'affacturage'],
            'cautions': ['cautions bancaires', 'caution bancaire', 'cautions'],
            'effets': ['effets de commerce', 'effet de commerce', 'effets commerciaux'],
            'spot': ['spot', 'spots', 'cours spot']
        }
        
        resultats_precalcul = {}
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_type = identifier_type_feuille(sheet_name, sheet_mapping)
            
            if not sheet_type:
                continue
            
            # Extraire les données avec calcul des TEG
            if sheet_type == 'credits':
                credits, _ = extraire_credits_amortissables(worksheet, etablissement, fichier_import)
                resultats_precalcul['credits'] = credits
            elif sheet_type == 'decouverts':
                decouverts, _ = extraire_decouverts(worksheet, etablissement, fichier_import)
                resultats_precalcul['decouverts'] = decouverts
            elif sheet_type == 'affacturages':
                affacturages, _ = extraire_affacturages(worksheet, etablissement, fichier_import)
                resultats_precalcul['affacturages'] = affacturages
            elif sheet_type == 'cautions':
                cautions, _ = extraire_cautions(worksheet, etablissement, fichier_import)
                resultats_precalcul['cautions'] = cautions
            elif sheet_type == 'effets':
                effets, _ = extraire_effets_commerces(worksheet, etablissement, fichier_import)
                resultats_precalcul['effets'] = effets
            elif sheet_type == 'spot':
                spots, _ = extraire_spot(worksheet, etablissement, fichier_import)
                resultats_precalcul['spots'] = spots
        
        # Stocker les résultats en cache pour utilisation ultérieure
        cache_key = f"precalcul_teg_{fichier_import.id}"
        cache.set(cache_key, resultats_precalcul, timeout=3600)  # 1 heure
        
        logger.info(f"Pré-calcul TEG terminé pour {fichier_import.nom_fichier}")
        return resultats_precalcul
        
    except Exception as e:
        logger.error(f"Erreur lors du pré-calcul TEG pour {fichier_import.nom_fichier}: {str(e)}")
        return {}
    
def verifier_teg_simplifiee(fichier_import):
    """
    Version simplifiée et robuste de la vérification TEG
    """
    resultat = {
        'credits': {'conformes': 0, 'non_conformes': 0, 'total': 0},
        'decouverts': {'conformes': 0, 'non_conformes': 0, 'total': 0},
        'affacturages': {'conformes': 0, 'non_conformes': 0, 'total': 0},
        'cautions': {'conformes': 0, 'non_conformes': 0, 'total': 0},
        'effets': {'conformes': 0, 'non_conformes': 0, 'total': 0},
        'spots': {'conformes': 0, 'non_conformes': 0, 'total': 0},
        'erreurs': []
    }

    TOLERANCE = 0.01  # 1% de tolérance

    # Configuration simple pour chaque type
    model_config = {
        'credits': {
            'model': Credit_Amortissables,
            'teg_calcule': 'TEG_annualise',
            'teg_original': 'TEG_I26'
        },
        'decouverts': {
            'model': Decouverts,
            'teg_calcule': 'TEG_decouvert',
            'teg_original': 'TEG_I17'
        },
        'affacturages': {
            'model': Affacturage,
            'teg_calcule': 'TEG_affacturage',
            'teg_original': 'TEG_I14'
        },
        'cautions': {
            'model': Cautions,
            'teg_calcule': 'TEG_caution',
            'teg_original': 'TEG_I14'
        },
        'effets': {
            'model': Effets_commerces,
            'teg_calcule': 'TEG_effet',
            'teg_original': 'TEG_I15'
        },
        'spots': {
            'model': Spot,
            'teg_calcule': 'TEG_spot',
            'teg_original': 'TEG_I26'
        }
    }

    try:
        for key, config in model_config.items():
            try:
                # Récupérer les données pour ce fichier
                queryset = config['model'].objects.filter(fichier_import=fichier_import)
                
                for obj in queryset:
                    try:
                        teg_calcule = getattr(obj, config['teg_calcule'], 0) or 0
                        teg_original = getattr(obj, config['teg_original'], 0) or 0
                        
                        # Conversion en float
                        teg_calcule_val = float(teg_calcule)
                        teg_original_val = float(teg_original)
                        
                        # Ignorer les valeurs nulles
                        if teg_calcule_val == 0 and teg_original_val == 0:
                            continue
                        
                        # Vérifier la conformité
                        difference = abs(teg_calcule_val - teg_original_val)
                        
                        if difference <= TOLERANCE:
                            resultat[key]['conformes'] += 1
                        else:
                            resultat[key]['non_conformes'] += 1
                            
                        resultat[key]['total'] += 1
                        
                    except (ValueError, TypeError):
                        continue
                        
            except Exception as e:
                resultat['erreurs'].append(f"Erreur {key}: {str(e)}")
                continue

    except Exception as e:
        resultat['erreurs'].append(f"Erreur générale: {str(e)}")

    return resultat   


import logging
from decimal import Decimal
from typing import Dict, List, Tuple

logger = logging.getLogger(__name__)

# ========================================
# FONCTION PRINCIPALE - EXTRACTION + CALCUL
# ========================================

def extraire_et_calculer_teg(fichier_path, etablissement_cnef) -> Dict:
    """
    Extrait les données ET calcule les TEG EN UNE SEULE PASSE
    Retourne un dictionnaire avec toutes les données nécessaires
    """
    import openpyxl
    
    resultats = {
        'credits': [],
        'decouverts': [],
        'affacturages': [],
        'cautions': [],
        'effets': [],
        'spots': [],
        'erreurs': []
    }
    
    try:
        workbook = openpyxl.load_workbook(fichier_path, data_only=True)
        
        sheet_mapping = {
            'credits': ['credits amortissables', 'credit amortissable', 'crédits amortissables'],
            'decouverts': ['découverts bancaires', 'decouvert', 'découverts'],
            'affacturages': ['affacturage', 'affacturages'],
            'cautions': ['cautions', 'caution'],
            'effets': ['effets de commerce', 'effet'],
            'spots': ['spot', 'spots', 'cours spot']
        }
        
        for sheet_name in workbook.sheetnames:
            sheet_type = identifier_type_feuille(sheet_name, sheet_mapping)
            
            if not sheet_type:
                continue
            
            worksheet = workbook[sheet_name]
            
            # Extraction ET calcul selon le type
            if sheet_type == 'credits':
                resultats['credits'] = extraire_credits_avec_teg(worksheet)
            elif sheet_type == 'decouverts':
                resultats['decouverts'] = extraire_decouverts_avec_teg(worksheet)
            elif sheet_type == 'affacturages':
                resultats['affacturages'] = extraire_affacturages_avec_teg(worksheet)
            elif sheet_type == 'cautions':
                resultats['cautions'] = extraire_cautions_avec_teg(worksheet)
            elif sheet_type == 'effets':
                resultats['effets'] = extraire_effets_avec_teg(worksheet)
            elif sheet_type == 'spots':
                resultats['spots'] = extraire_spots_avec_teg(worksheet)
        
        logger.info(f"✅ Extraction terminée: {sum(len(v) for k, v in resultats.items() if k != 'erreurs')} lignes")
        
    except Exception as e:
        resultats['erreurs'].append(f"Erreur extraction: {str(e)}")
        logger.error(f"❌ Erreur extraction: {e}")
    
    return resultats


# ========================================
# FONCTIONS D'EXTRACTION AVEC CALCUL TEG
# ========================================

def extraire_credits_avec_teg(worksheet) -> List[Dict]:
    """Extrait les crédits ET calcule leurs TEG"""
    credits = []
    
    # Trouver l'en-tête
    header_row = trouver_entete(worksheet)
    if not header_row:
        return credits
    
    for row_num, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values = [cell.value for cell in row]
        
        if not any(values) or len(values) < 26:
            continue
        
        try:
            # Extraction des valeurs
            montant_pret = convertir_decimal_safe(values[12])
            duree = convertir_entier_safe(values[13])
            montant_echeance = convertir_decimal_safe(values[22])
            taux_nominal = convertir_decimal_safe(values[16])
            frais_dossier = convertir_decimal_safe(values[17])
            montant_assurance = convertir_decimal_safe(values[19])
            frais_annexe = convertir_decimal_safe(values[20])
            teg_original = convertir_decimal_safe(values[25])
            freq_remb = str(values[15] or '1').strip().lower()
            
            # Normalisation
            if taux_nominal > 1:
                taux_nominal /= 100
            if teg_original > 1:
                teg_original /= 100
            
            # CALCUL DU TEG
            teg_mensuel, teg_annualise = calculer_teg_credit(
                montant_pret, duree, montant_echeance,
                frais_dossier, montant_assurance, frais_annexe,
                freq_remb
            )
            
            # Comparaison
            conforme = verifier_conformite_teg(teg_annualise / 100, teg_original)
            
            credits.append({
                'montant_pret': montant_pret,
                'duree': duree,
                'taux_nominal': taux_nominal,
                'teg_calcule': teg_annualise,  # En %
                'teg_original': teg_original,  # En décimal
                'conforme': conforme,
                'ligne': row_num
            })
            
        except Exception as e:
            logger.debug(f"Ligne {row_num}: {e}")
            continue
    
    return credits


def extraire_decouverts_avec_teg(worksheet) -> List[Dict]:
    """Extrait les découverts ET calcule leurs TEG"""
    decouverts = []
    
    header_row = trouver_entete(worksheet)
    if not header_row:
        return decouverts
    
    for row_num, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values = [cell.value for cell in row]
        
        if not any(values) or len(values) < 17:
            continue
        
        try:
            montant_decouvert = convertir_decimal_safe(values[7])
            taux_nominal = convertir_decimal_safe(values[9])
            frais_dossiers = convertir_decimal_safe(values[10])
            couts_assurance = convertir_decimal_safe(values[11])
            frais_annexes = convertir_decimal_safe(values[12])
            teg_original = convertir_decimal_safe(values[16])
            
            if taux_nominal > 1:
                taux_nominal /= 100
            if teg_original > 1:
                teg_original /= 100
            
            # CALCUL DU TEG
            teg_calcule = calculer_teg_decouvert(
                montant_decouvert, taux_nominal,
                frais_dossiers, couts_assurance, frais_annexes
            )
            
            conforme = verifier_conformite_teg(teg_calcule / 100, teg_original)
            
            decouverts.append({
                'montant': montant_decouvert,
                'taux_nominal': taux_nominal,
                'teg_calcule': teg_calcule,
                'teg_original': teg_original,
                'conforme': conforme,
                'ligne': row_num
            })
            
        except Exception as e:
            continue
    
    return decouverts


def extraire_affacturages_avec_teg(worksheet) -> List[Dict]:
    """Extrait les affacturages ET calcule leurs TEG"""
    affacturages = []
    
    header_row = trouver_entete(worksheet)
    if not header_row:
        return affacturages
    
    for row_num, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values = [cell.value for cell in row]
        
        if not any(values) or len(values) < 14:
            continue
        
        try:
            duree = convertir_entier_safe(values[4])
            montant_creance = convertir_entier_safe(values[9])
            montant_com_affacturage = convertir_decimal_safe(values[10])
            montant_comm_financement = convertir_entier_safe(values[11])
            montant_frais_annexes = convertir_entier_safe(values[12])
            teg_original = convertir_decimal_safe(values[13])
            
            if teg_original > 1:
                teg_original /= 100
            
            # CALCUL DU TEG
            teg_calcule = calculer_teg_affacturage(
                montant_creance, duree,
                montant_com_affacturage, montant_comm_financement,
                montant_frais_annexes
            )
            
            conforme = verifier_conformite_teg(teg_calcule / 100, teg_original)
            
            affacturages.append({
                'montant_creance': montant_creance,
                'duree': duree,
                'teg_calcule': teg_calcule,
                'teg_original': teg_original,
                'conforme': conforme,
                'ligne': row_num
            })
            
        except Exception as e:
            continue
    
    return affacturages


def extraire_cautions_avec_teg(worksheet) -> List[Dict]:
    """Extrait les cautions ET calcule leurs TEG"""
    cautions = []
    
    header_row = trouver_entete(worksheet)
    if not header_row:
        return cautions
    
    for row_num, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values = [cell.value for cell in row]
        
        if not any(values) or len(values) < 14:
            continue
        
        try:
            duree = convertir_entier_safe(values[4])
            montant_caution = convertir_entier_safe(values[9])
            taux_caution = convertir_decimal_safe(values[10])
            frais_comm = convertir_entier_safe(values[11])
            frais_annexes = convertir_entier_safe(values[12])
            teg_original = convertir_decimal_safe(values[13])
            
            if taux_caution > 1:
                taux_caution /= 100
            if teg_original > 1:
                teg_original /= 100
            
            # CALCUL DU TEG
            teg_calcule = calculer_teg_caution(
                montant_caution, duree, taux_caution,
                frais_comm, frais_annexes
            )
            
            conforme = verifier_conformite_teg(teg_calcule / 100, teg_original)
            
            cautions.append({
                'montant': montant_caution,
                'duree': duree,
                'teg_calcule': teg_calcule,
                'teg_original': teg_original,
                'conforme': conforme,
                'ligne': row_num
            })
            
        except Exception as e:
            continue
    
    return cautions


def extraire_effets_avec_teg(worksheet) -> List[Dict]:
    """Extrait les effets ET calcule leurs TEG"""
    effets = []
    
    header_row = trouver_entete(worksheet)
    if not header_row:
        return effets
    
    for row_num, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values = [cell.value for cell in row]
        
        if not any(values) or len(values) < 15:
            continue
        
        try:
            duree = convertir_entier_safe(values[4])
            taux_nominal = convertir_decimal_safe(values[9])
            montant_effet = convertir_entier_safe(values[10])
            montant_commission = convertir_entier_safe(values[12])
            autres_frais = convertir_entier_safe(values[13])
            teg_original = convertir_decimal_safe(values[14])
            
            if taux_nominal > 1:
                taux_nominal /= 100
            if teg_original > 1:
                teg_original /= 100
            
            # CALCUL DU TEG
            teg_calcule = calculer_teg_effet(
                montant_effet, duree, taux_nominal,
                montant_commission, autres_frais
            )
            
            conforme = verifier_conformite_teg(teg_calcule / 100, teg_original)
            
            effets.append({
                'montant': montant_effet,
                'duree': duree,
                'teg_calcule': teg_calcule,
                'teg_original': teg_original,
                'conforme': conforme,
                'ligne': row_num
            })
            
        except Exception as e:
            continue
    
    return effets


def extraire_spots_avec_teg(worksheet) -> List[Dict]:
    """Extrait les spots ET calcule leurs TEG"""
    spots = []
    
    header_row = trouver_entete(worksheet)
    if not header_row:
        return spots
    
    for row_num, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values = [cell.value for cell in row]
        
        if not any(values) or len(values) < 26:
            continue
        
        try:
            montant_pret = convertir_decimal_safe(values[12])
            duree = convertir_entier_safe(values[13])
            montant_echeance = convertir_decimal_safe(values[22])
            frais_dossier = convertir_decimal_safe(values[17])
            montant_assurance = convertir_decimal_safe(values[19])
            frais_annexe = convertir_decimal_safe(values[20])
            teg_original = convertir_decimal_safe(values[25])
            
            if teg_original > 1:
                teg_original /= 100
            
            # CALCUL DU TEG
            teg_calcule = calculer_teg_spot(
                montant_pret, duree, montant_echeance,
                frais_dossier, montant_assurance, frais_annexe
            )
            
            conforme = verifier_conformite_teg(teg_calcule / 100, teg_original)
            
            spots.append({
                'montant_pret': montant_pret,
                'duree': duree,
                'teg_calcule': teg_calcule,
                'teg_original': teg_original,
                'conforme': conforme,
                'ligne': row_num
            })
            
        except Exception as e:
            continue
    
    return spots


# ========================================
# FONCTIONS DE CALCUL TEG
# ========================================

def calculer_teg_credit(montant_pret, duree, montant_echeance,
                        frais_dossier, montant_assurance, frais_annexe,
                        freq_remb) -> Tuple[float, float]:
    """Calcule le TEG mensuel et annualisé pour un crédit"""
    import numpy_financial as npf
    
    try:
        if not all([montant_pret, duree, montant_echeance]):
            return 0.0, 0.0
        
        montant_net = montant_pret - (frais_dossier or 0) - (montant_assurance or 0) - (frais_annexe or 0)
        
        if montant_net <= 0:
            return 0.0, 0.0
        
        # TEG mensuel
        teg_mensuel = npf.rate(
            nper=duree,
            pmt=-montant_echeance,
            pv=montant_net,
            fv=0
        ) * 100
        
        # Multiplicateur selon fréquence
        freq_mapping = {
            '1': 12, 'mensuel': 12, 'mois': 12,
            '2': 4, 'trimestriel': 4,
            '3': 2, 'semestriel': 2,
            '4': 1, 'annuel': 1,
            '5': 52, 'hebdomadaire': 52
        }
        
        multiplicateur = 12
        for key, value in freq_mapping.items():
            if key in freq_remb:
                multiplicateur = value
                break
        
        teg_annualise = teg_mensuel * multiplicateur
        
        return round(teg_mensuel, 4), round(teg_annualise, 2)
        
    except Exception:
        return 0.0, 0.0


def calculer_teg_decouvert(montant_decouvert, taux_nominal,
                          frais_dossiers, couts_assurance, frais_annexes) -> float:
    """Calcule le TEG pour un découvert"""
    try:
        if not montant_decouvert or montant_decouvert == 0:
            return 0.0
        
        interets = montant_decouvert * taux_nominal
        frais_totaux = (frais_dossiers or 0) + (couts_assurance or 0) + (frais_annexes or 0)
        
        teg = ((interets + frais_totaux) / montant_decouvert) * 100
        
        return round(teg, 2)
        
    except Exception:
        return 0.0


def calculer_teg_affacturage(montant_creance, duree,
                            montant_com_affacturage, montant_comm_financement,
                            montant_frais_annexes) -> float:
    """Calcule le TEG pour un affacturage"""
    try:
        if not montant_creance or montant_creance == 0 or not duree or duree == 0:
            return 0.0
        
        frais_totaux = ((montant_com_affacturage or 0) + 
                       (montant_comm_financement or 0) + 
                       (montant_frais_annexes or 0)) / montant_creance
        
        teg = (frais_totaux * 360 / duree) * 100
        
        return round(teg, 2)
        
    except Exception:
        return 0.0


def calculer_teg_caution(montant_caution, duree, taux_caution,
                        frais_comm, frais_annexes) -> float:
    """Calcule le TEG pour une caution"""
    try:
        if not montant_caution or montant_caution == 0 or not duree or duree == 0:
            return 0.0
        
        montant_net = montant_caution - (frais_comm or 0) - (frais_annexes or 0)
        
        if montant_net == 0:
            return 0.0
        
        cout_annualise = (montant_caution * taux_caution * duree) / 360
        teg = (cout_annualise / montant_net) * (360 / duree) * 100
        
        return round(teg, 2)
        
    except Exception:
        return 0.0


def calculer_teg_effet(montant_effet, duree, taux_nominal,
                      montant_commission, autres_frais) -> float:
    """Calcule le TEG pour un effet de commerce"""
    try:
        if not montant_effet or montant_effet == 0 or not duree or duree == 0:
            return 0.0
        
        montant_net = montant_effet - (montant_commission or 0) - (autres_frais or 0)
        
        if montant_net == 0:
            return 0.0
        
        cout_interets = (taux_nominal * montant_effet * duree) / 360
        teg = ((cout_interets / montant_net) * (360 / duree)) * 100
        
        return round(teg, 2)
        
    except Exception:
        return 0.0


def calculer_teg_spot(montant_pret, duree, montant_echeance,
                     frais_dossier, montant_assurance, frais_annexe) -> float:
    """Calcule le TEG pour un spot"""
    try:
        if not montant_pret or montant_pret == 0 or not duree or duree == 0:
            return 0.0
        
        taux_periodique = ((montant_echeance / montant_pret - 1) * 12) / duree
        charges = ((frais_dossier or 0) + (montant_assurance or 0) + (frais_annexe or 0))
        ratio_charges = charges / montant_pret
        
        teg = (taux_periodique + ratio_charges) * 100
        
        return round(teg, 2)
        
    except Exception:
        return 0.0


# ========================================
# FONCTION DE VÉRIFICATION
# ========================================

def verifier_conformite_teg(teg_calcule: float, teg_original: float, 
                           tolerance: float = 0.001) -> bool:
    """
    Vérifie si le TEG calculé est conforme au TEG original
    
    Args:
        teg_calcule: TEG calculé (en décimal, ex: 0.125 pour 12.5%)
        teg_original: TEG original (en décimal)
        tolerance: Tolérance acceptable (par défaut 0.1%)
    
    Returns:
        True si conforme, False sinon
    """
    try:
        if teg_calcule == 0 and teg_original == 0:
            return True
        
        difference = abs(teg_calcule - teg_original)
        return difference <= tolerance
        
    except Exception:
        return False


# ========================================
# FONCTION UTILITAIRE
# ========================================

def trouver_entete(worksheet, max_rows: int = 10) -> int:
    """Trouve la ligne d'en-tête dans une feuille"""
    for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_rows), start=1):
        if any(cell.value for cell in row):
            return row_num
    return None


# ========================================
# GÉNÉRATION DES STATISTIQUES
# ========================================

def generer_statistiques_teg(donnees_extraites: Dict) -> Dict:
    """
    Génère les statistiques de conformité TEG
    
    Args:
        donnees_extraites: Dictionnaire retourné par extraire_et_calculer_teg()
    
    Returns:
        Dictionnaire avec les statistiques détaillées
    """
    stats = {
        'credits': {'conformes': 0, 'non_conformes': 0, 'total': 0},
        'decouverts': {'conformes': 0, 'non_conformes': 0, 'total': 0},
        'affacturages': {'conformes': 0, 'non_conformes': 0, 'total': 0},
        'cautions': {'conformes': 0, 'non_conformes': 0, 'total': 0},
        'effets': {'conformes': 0, 'non_conformes': 0, 'total': 0},
        'spots': {'conformes': 0, 'non_conformes': 0, 'total': 0}
    }
    
    for type_produit in stats.keys():
        if type_produit in donnees_extraites:
            for item in donnees_extraites[type_produit]:
                stats[type_produit]['total'] += 1
                if item.get('conforme', False):
                    stats[type_produit]['conformes'] += 1
                else:
                    stats[type_produit]['non_conformes'] += 1
    
    # Statistiques globales
    total_global = sum(s['total'] for s in stats.values())
    conformes_global = sum(s['conformes'] for s in stats.values())
    non_conformes_global = sum(s['non_conformes'] for s in stats.values())
    
    stats['global'] = {
        'total': total_global,
        'conformes': conformes_global,
        'non_conformes': non_conformes_global,
        'taux_conformite': round((conformes_global / total_global * 100), 2) if total_global > 0 else 0
    }
    
    return stats
